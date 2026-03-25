## Functions are in py

import pandas as pd
import xlrd
import openpyxl
from pandas import ExcelWriter
from datetime import datetime
import os
import numpy as np
#from arcgis.gis import GIS #DOESN'T Work on Mac
#from arcgis.features import SpatialDataFrame #DOESN'T Work on Mac
from shutil import copyfile
from openpyxl.utils.cell import get_column_letter
import urllib.request, json


def detect_enumerator(questionnaire_file):
    enumerator = ""
    if "kobo" in questionnaire_file.lower():
        enumerator = "kobo"
    elif "geopoll" in questionnaire_file.lower():
        enumerator = "geopoll"
    else:
        print("Please include Geopoll or Kobo inside the filename: %s " % questionnaire_file)
        enumerator = "Please include Geopoll or Kobo inside the filename"
    return enumerator

def detect_language(file_name):
    language = ""
    if "_en_" in file_name.lower():
        language = "en"
    elif "_fr_" in file_name.lower():
        language = "fr"
    elif "_es_" in file_name.lower():
        language = "es"
    elif "_ar_" in file_name.lower():
        language = "ar"
    else:
        print("Please include the language in the filename: %s " % file_name)
        #"Please include the language inside the filename"
    return language

def detect_template(template_version,file_name):
    language = detect_language(file_name)
    if language == "en":
        template_questionnaire_file = r'household_questionnaire_kobo_EN_template_' + template_version + '_ISO3_F2F.xlsx'
    elif language == "fr":
        template_questionnaire_file = r'household_questionnaire_kobo_FR_template_' + template_version + '_ISO3_F2F.xlsx'
    elif language == "es":
        template_questionnaire_file = r'household_questionnaire_kobo_ES_template_' + template_version + '_ISO3_F2F.xlsx'
    elif language == "ar":
        template_questionnaire_file = r'household_questionnaire_kobo_AR_template_' + template_version + '_ISO3_F2F.xlsx'
    else:
        print("Please include the language in the filename: %s " % country_template)
    return template_questionnaire_file


def count_number_of_all_question_name(questionnaire_file):
    questionnaire_df = pd.read_excel(pd.ExcelFile(questionnaire_file), sheet_name='survey')
    questionnaire_df = questionnaire_df[['name']]

    list_questions = questionnaire_df.dropna().values.tolist()
    list_questions = [item for sublist in list_questions for item in sublist]
    list_questions = [s.strip() for s in list_questions]
    n_of_questions = len(list_questions)
    return list_questions, n_of_questions


#The below function is not used
def count_number_of_questions_name(questionnaire_file):
    questionnaire_df = pd.read_excel(pd.ExcelFile(questionnaire_file), sheet_name='survey')
    #print(questionnaire_df)
    questionnaire_df = questionnaire_df.loc[((questionnaire_df["type"] != "start") &
                                             (questionnaire_df["type"] != "end") &
                                             (questionnaire_df["type"] != "today") &
                                             (questionnaire_df["type"] != "deviceid") &
                                             (questionnaire_df["type"] != "calculate") &
                                             (questionnaire_df["type"] != "note") &
                                             (questionnaire_df["type"] != "begin_group") &
                                             (questionnaire_df["type"] != "begin group") &
                                             (questionnaire_df["type"] != "end_group")) , ['name'] ]
    #print(questionnaire_df.head())
    questionnaire_df = questionnaire_df[['name']]

    list_questions = questionnaire_df.dropna().values.tolist()
    list_questions = [item for sublist in list_questions for item in sublist]
    list_questions = [s.strip() for s in list_questions]
    n_of_questions = len(list_questions)
    return list_questions, n_of_questions


def update_question_label(questionnaire_file):
    # Assume detect_language function is defined elsewhere
    language = detect_language(questionnaire_file)
    workbook = openpyxl.load_workbook(questionnaire_file)
    worksheet = workbook['survey']

    name_col = None
    label_col_en = None
    label_col_fr = None
    label_col_ar = None
    label_col_es = None

    # Identify columns for each language
    for i in range(1, worksheet.max_column + 1):
        header_value = str(worksheet[f"{get_column_letter(i)}1"].value)
        if header_value.strip() == 'name':
            name_col = i
        elif header_value.strip() == 'label::English (en)':
            label_col_en = i
        elif header_value.strip() == 'label::French (fr)':
            label_col_fr = i
        elif header_value.strip() == 'label::Arabic (ar)':
            label_col_ar = i
        elif header_value.strip() == 'label::Spanish (es)':
            label_col_es = i

    if language == "en" and name_col and label_col_en:
        update_labels(worksheet, name_col, label_col_en, 'en')
    elif language == "fr" and name_col and label_col_fr:
        update_labels(worksheet, name_col, label_col_fr, 'fr')
    elif language == "ar" and name_col and label_col_ar:
        update_labels(worksheet, name_col, label_col_ar, 'ar')
    elif language == "es" and name_col and label_col_es:
        update_labels(worksheet, name_col, label_col_es, 'es')

    workbook.save(questionnaire_file)
    print("Workbook saved after updates.")

def update_labels(worksheet, name_col, label_col, language):
    # Define the text replacements for each identified question
    questions_updates = {
        "adm1_pcode": {
            'en': "Currently, in which #ADMIN1# does your household reside?",
            'fr': "Actuellement, dans quel #ADMIN1# votre ménage vit-il ?",
            'ar': "حاليًا ، في أي #ADMIN1 AR# تقيم أسرتك؟",
            'es': "Actualmente, ¿en qué #ADMIN1 ES# reside su hogar?"
        },
        "adm2_pcode ": {
            'en': "Currently, in which #ADMIN2# does your household reside?",
            'fr': "Actuellement, dans quel #ADMIN2# votre ménage vit-il ?",
            'ar': "حاليًا ، في أي #ADMIN2 AR# تقيم أسرتك؟",
            'es': "Actualmente, ¿en qué #ADMIN2 ES# reside su hogar?"
        },
        "income_main_amount": {
            'en': "How much was earned (in #currency#) from your main income source, ${incomemain}, in the past 3 months ?",
            'fr': "Quelle somme (en #currency# a rapporté votre principale source de revenu, ${incomemain}, au cours des trois derniers mois ?",
            'ar': "ما المبلغ الذي كسبته أسرتك (بالعملة #currency AR#) من مصدر دخلها الرئيسي ، ${incomemain} ، في الأشهر الثلاثة الماضية؟",
            'es': "¿Cuánto ganó (en #currency#) por su principal fuente de ingresos, ${incomemain}, en los últimos 3 meses?"
        },
        "income_main_comp": {
            'en': "Has your household's income from ${incomemain} changed in the last 3 months compared to the same period of #reference year#?",
            'fr': "Le revenu de votre ménage provenant de ${incomemain} a-t-il changé au cours des trois derniers mois par rapport à la même periode de #reference year FR# ?",
            'ar': "هل تغير دخل أسرتك من ${incomemain} في الأشهر الثلاثة الماضية مقارنة بنفس الفترة من #reference year AR#؟",
            'es': "En los últimos 3 meses, ¿han cambiado los ingresos del hogar provenientes de la ${incomemain} comparado con el mismo periodo en #reference year ES#?"
        },
        "income_sec_amount": {
            'en': "How much was earned (in #currency#) from your second main income source, ${incomesec}, in the past 3 months ?",
            'fr': "Quelle somme (en #currency#) a rapporté votre deuxième source de revenu, ${incomesec}, au cours des trois derniers mois ?",
            'ar': "ما المبلغ الذي كسبته أسرتك (بالعملة #currency AR#) من ثاني أهم مصدر دخل لها ، ${incomesec} ، في الأشهر الثلاثة الماضية؟",
            'es': "¿Cuánto ganó (en #currency#) por la segunda fuente de ingresos, ${incomesec}, en los últimos 3 meses?"
        },
        "income_sec_comp": {
            'en': "Has your household's income from ${incomesec} changed in the last 3 months compared to the same period of #reference year#?",
            'fr': "Le revenu de votre ménage provenant de ${incomesec} a-t-il changé au cours des trois derniers mois par rapport à la même période de #reference year FR# ?",
            'ar': "هل تغير دخل أسرتك من ${incomesec} في الأشهر الثلاثة الماضية مقارنة بنفس الفترة من #reference year AR#؟",
            'es': "En los últimos 3 meses, ¿han cambiado los ingresos del hogar provenientes de la ${incomesec} comparado con el mismo periodo en #reference year ES#?"
        },
        "income_third_amount": {
            'en': "How much was earned (in #currency#) from your third main income source, ${incomethird}, in the past 3 months ?",
            'fr': "Quelle somme (en #currency#) a rapporté votre troisième source de revenu, ${incomethird}, au cours des trois derniers mois ?",
            'ar': "ما المبلغ الذي كسبته أسرتك (بالعملة #currency AR#) من ثالث أهم مصدر دخل لها ، ${incomethird} ، في الأشهر الثلاثة الماضية؟",
            'es': "¿Cuánto ganó (en #currency#) por la tercera fuente de ingresos, ${incomethird}, en los últimos 3 meses?"
        },
        "income_third_comp": {
            'en': "Has your household's income from ${incomethird} changed in the last 3 months compared to the same period of #reference year#?",
            'fr': "Le revenu de votre ménage provenant de ${incomethird} a-t-il changé au cours des trois derniers mois par rapport à la même période de #reference year FR# ?",
            'ar': "هل تغير دخل أسرتك من ${incomethird} في الأشهر الثلاثة الماضية مقارنة بنفس الفترة من #reference year AR#؟",
            'es': "En los últimos 3 meses, ¿han cambiado los ingresos del hogar provenientes de la ${incomethird} comparado con el mismo periodo en #reference year ES#?"
        },
        "crp_intro": {
            'en': "We will now ask you some questions about crop production which apply to you and anyone in your household that is involved in crop production or sales. All questions refer to the crops grown during #season#.",
            'fr': "Nous allons maintenant vous poser quelques questions sur la production végétale. Ces questions vous concernent ainsi que toute personne de votre ménage impliquée dans la production ou la vente de cultures. Toutes les questions se rapportent aux cultures pratiquées pendant la #season FR#. ",
            'ar': "سنطرح عليك الآن بعض الأسئلة حول إنتاج المحاصيل التي تنطبق عليك وعلى أي شخص في أسرتك يشارك في إنتاج المحاصيل أو بيعها. تشير جميع الأسئلة إلى المحاصيل المزروعة خلال #season AR#. ",
            'es': "Ahora me gustaría preguntarle sobre la producción de cultivos. Estas preguntas aplican para usted o cualquier miembro de su hogar que esté involucrado/a en la producción de cultivos. Todas las preguntas se refieren a lo producido durante #season ES#."
        },
        "crp_main": {
            'en': "What is the main crop that your household grows for food or income generation in #season#, if any?",
            'fr': "Quelle est la principale culture pratiquée par votre ménage pour se nourrir ou générer des revenus pendant la #season FR#, si applicable ?",
            'ar': "ما هو المحصول الرئيسي الذي تزرعه أسرتك من أجل الغذاء وتوليد الدخل في موسم #season AR# ، إن وجد؟",
            'es': "De haber alguno, ¿cuál es el principal cultivo que su hogar produce para la alimentación o la generación de ingresos en #season ES#?"
        },
        "crp_main_check": {
            'en': "Why did you not produce any crops in #season#?",
            'fr': "Pourquoi n'avez-vous pas produit de cultures pendant la #season FR# ?",
            'ar': "لماذا لم تنتج أي محاصيل في #season AR#؟",
            'es': "¿Por qué no produjo ningún cultivo en #season ES#"
        },
        "crp_area_change": {
            'en': "How did the ***area planted*** for ${crpmain} from #season# compare to #reference year#?",
            'fr': "Comment estimez-vous la **superficie plantée** de  ${crpmain} pour la #season FR# par rapport à #reference year FR# ?",
            'ar': "كيف يمكن مقارنة ***المساحة المزروعة*** ل ${crpmain} في #season AR# نسبةً ل #reference year AR#؟",
            'es': "¿Cómo se compara el ***área sembrada*** de ${crpmain} en #season ES# con respecto a #reference year ES#?"
        },
        "crp_harv_change": {
            'en': "How does the #expected or nothing# ***harvest*** of ${crpmain} in the #season# compare to #reference year#? ",
            'fr': "Comment estimez-vous la **récolte** #expected or nothing FR# de ${crpmain} pour la #season FR# par rapport à #reference year FR# ?",
            'ar': "كيف يمكن مقارنة ***الحصاد***  #expected or nothing AR# لـ ${crpmain} في #season AR# نسبةً ل #reference year AR#؟",
            'es': "¿Cómo compara la ***cosecha*** #expected or nothing ES# para su ${crpmain} en #season ES# comparado con #reference year ES#?"
        },
        "crp_harv_vol": {
            'en': "How much (threshed) ${crpmain} did you harvest this #season#?",
            'fr': "Quelle quantité (battue) de ${crpmain} avez-vous récoltée cette #season FR# ?",
            'ar': " ما هي كمية ال ${crpmain}  (المدروس/ة) التي قمت بحصادها هذا الموسم #season AR#؟",
            'es': "¿Cuánto ${crp_main} (trillado) cosechó en #season ES#"
        },
        "crp_proddif": {
            'en': "Has your household faced any significant difficulty ***in terms of crop production*** in #season#?",
            'fr': "Votre ménage a-t-il rencontré des difficultés importantes en matière de **production végétale** au cours de #season FR# ?",
            'ar': "هل واجهت أسرتك أي صعوبة ***في إنتاج المحاصيل*** في #season AR#؟",
            'es': "¿Ha enfrentado su hogar algún tipo de dificultad con la ***producción de cultivos*** en #season ES#? "
        },
        "crp_proddif_": {
            'en': "Which difficulties has your household faced with crop production in #season#?",
            'fr': "Quelles difficultés votre ménage a-t-il rencontrées dans la production végétale au cours de #season FR# ?",
            'ar': "ما الصعوبات التي واجهتها أسرتك في إنتاج المحاصيل في #season AR#؟",
            'es': "¿Cuáles dificultades ha enfrentado su hogar con la producción de cultivos en #season ES#?"
        },
        "crp_salesprice": {
            'en': "How does the current price of your household's ${crp_sales_main} compare to the average price in the same period of #reference year#?",
            'fr': "Comment évaluez-vous le prix actuel de ${crp_sales_main} par rapport au prix moyen de la même periode de #reference year FR# ?",
            'ar': "كيف تقارن السعر الحالي ل${crp_sales_main} لأسرتك بمتوسط ​​السعر في نفس الفترة من #reference year AR#؟",
            'es': "¿Cómo compararía el precio actual de la cosecha principal de su hogar ${crp_sales_main} con respecto al precio promedio en #reference year ES#?"
        },
        "ls_salesprice": {
            'en': "How does the current price of your household's ${lssalesmain} compare to the average price in the same period of #reference year#?",
            'fr': "Comment évaluez-vous le prix actuel de ${lssalesmain} par rapport au prix moyen sur la même période de #reference year FR# ?",
            'ar': "كيف تقارن السعر الحالي ل ${lssalesmain} لأسرتك بمتوسط ​​السعر في نفس الفترة من #reference year AR#؟",
            'es': "¿Cómo se compara el precio actual del ${lssalesmain} con el precio promedio para la misma época del año en #reference year ES#?"
        },
        "fish_salesprice": {
            'en': "How does the current price of your household's ${fishsalesmain} compare to the average price in the same period/harvest of #reference year#? ",
            'fr': "Comment évaluez-vous le prix actuel de vos ${fishsalesmain} vendus par votre ménage par rapport au prix moyen sur la même période [récolte] de #reference year FR# ? ",
            'ar': "كيف تقارن السعر الحالي لبيع  ${fishsalesmain}  لأسرتك بمتوسط السعر في نفس الفترة من #reference year AR#؟",
            'es': "¿Cómo se compara el precio actual de ${fishsalesmain} vendido por su hogar, con respecto con el precio promedio para el mismo periodo/cosecha de #reference year ES#?"
        },
        "fcs_vegetables_days": {
            'en': "How many days over *the last 7 days*, did members of your household eat **vegetables or leaves** such as #local vegetables# and/or other leaves/vegetables?",
            'fr': "Combien de fois (en nombre de jours) au cours des *7 derniers jours*, les membres de votre foyer ont-ils mangé des **légumes ou des feuilles** tels que des #légumes locaux# et/ou d'autres feuilles/légumes ?",
            'ar': "كم يومًا خلال *السبعة أيام الماضية* ، أكل أفراد أسرتك **خضروات أو خضار ورقية** مثل  #local vegetables AR# و / أو خضار ورقية / خضروات أخرى؟",
            'es': "¿Durante cuántos días en los *últimos 7 días* consumió usted o su hogar **vegetales y hojas** como #local vegetables ES# y/o otras hojas/verduras?"
        },
        "fcs_fruit_days": {
            'en': "How many days over *the last 7 days*, did members of your household eat **fruit** such as #local fruits# and/or other fruits?",
            'fr': "Combien de fois (en nombre de jours) au cours des *7 derniers jours*, les membres de votre foyer ont-ils mangé des **fruits** tels que des #fruits locaux# et/ou d'autres fruits ?",
            'ar': "كم يومًا خلال *السبعة أيام الماضية* ، تناول أفراد أسرتك **فواكه** مثل #local fruits AR# و / أو غيرها من الفواكه؟",
            'es': "¿Durante cuántos días en los *últimos 7 días* consumió usted o su hogar **frutas** como #local fruits ES# y/o otras frutas?"
        },
    }

    # Process each row in the 'name' column
    for row in range(2, worksheet.max_row + 1):  # Skip header row
        question_key = worksheet[f"{get_column_letter(name_col)}{row}"].value
        if question_key and question_key in questions_updates:
            replacement_text = questions_updates[question_key][language]
            worksheet[f"{get_column_letter(label_col)}{row}"].value = replacement_text
            print(f"Updated row {row} for '{question_key}' in column {get_column_letter(label_col)}")

def find_and_replace_strings_in_df(questionnaire_file):
    
    language = detect_language(questionnaire_file)
    replacing_table_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Additional information', skiprows=1)
    workbook = openpyxl.load_workbook(questionnaire_file)
    workbook.sheetnames
    worksheet = workbook["survey"]
    number_of_rows = worksheet.max_row #Number of Rows
    number_of_columns = worksheet.max_column #Number of Columns

    worksheet_choices = workbook["choices"]
    number_of_rows_choices = worksheet_choices.max_row #Number of Rows
    number_of_columns_choices = worksheet_choices.max_column #Number of Columns

    
    if language == "en":
        print("English")
        
        replacing_table_df = replacing_table_df[['Original ','Replacement']]
        replacement_list = list(replacing_table_df['Original '].values.tolist())
        replacing_table_df.set_index("Original ", inplace = True)
        


        expected_or_nothing = ""

        if (replacing_table_df.loc['season phase']['Replacement'] == "Not yet in season" or replacing_table_df.loc['season phase']['Replacement'] == "Land preparation" or replacing_table_df.loc['season phase']['Replacement'] == "Planting" or replacing_table_df.loc['season phase']['Replacement'] == "Early growing" or replacing_table_df.loc['season phase']['Replacement'] == "Growing" or replacing_table_df.loc['season phase']['Replacement'] == "Maturing"):
            expected_or_nothing = "expected"
        elif (replacing_table_df.loc['season phase']['Replacement'] == "Harvesting" or replacing_table_df.loc['season phase']['Replacement'] == "Recently finished"):
            expected_or_nothing = ""
        else:
            expected_or_nothing = ""

        # Saison and attendue ou rien added on the replacementTextKeyPairs
        replacementTextKeyPairs = {'#phone number#': replacing_table_df.loc["phone number"]["Replacement"],
                                   '#number of digits#': len(str(replacing_table_df.loc["phone number"]["Replacement"])),
                                   '#age#': replacing_table_df.loc["age"]["Replacement"],
                                   '#ADMIN1#': replacing_table_df.loc["ADMIN1"]["Replacement"],
                                   '#ADMIN2#': replacing_table_df.loc["ADMIN2"]["Replacement"],
                                   '#reference year#': replacing_table_df.loc["reference year"]["Replacement"],
                                   '#season#': replacing_table_df.loc["season"]["Replacement"],
                                   '#season phase#': replacing_table_df.loc['season phase']['Replacement'],
                                   '#expected or nothing#': expected_or_nothing,
                                   '#local measurement units#': replacing_table_df.loc["local measurement unit"]["Replacement"],
                                   '#currency#': replacing_table_df.loc["currency"]["Replacement"],
                                   '#MIN AMOUNT#': replacing_table_df.loc["MIN AMOUNT"]["Replacement"],
                                   #'#MAX SALARY#': replacing_table_df.loc[5]['Replacement'],
                                   '#THRESHOLD#': replacing_table_df.loc["THRESHOLD"]["Replacement"],
                                   '#local vegetables#': replacing_table_df.loc["local vegetables"]["Replacement"],
                                   '#local fruits#': replacing_table_df.loc["local fruits"]["Replacement"]
                                  }

        if pd.isna(replacing_table_df.loc["local measurement unit"]["Replacement"]) == True:
            print("\nLocal measurement unit is null, the last answer option (#5) has to be deleted manually.")


        #print("replacing_tableagain")
        #print(replacing_table_df)

        # Iterate over the columns and rows, search
        # for the text and replace
        for i in range(number_of_columns):
            for k in range(number_of_rows):
                cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        #newCellValue = replacementTextKeyPairs.get(key)
                        worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue



        # Iterate over the columns and rows, search
        # for the text and replace
        for i in range(number_of_columns_choices):
            for k in range(number_of_rows_choices):
                cellValue = str(worksheet_choices[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        #newCellValue = replacementTextKeyPairs.get(key)
                        worksheet_choices[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue
                        
                                     
    elif language == "fr":
        print("French")
        replacing_table_df = replacing_table_df[['Original','Replacement (EN)','Replacement (FR)']]
        replacement_list = list(replacing_table_df['Original'].values.tolist())
        replacing_table_df.set_index("Original", inplace = True)
        


        expected_or_nothing = ""
        attendue_ou_rien = ""
        if (replacing_table_df.loc['season phase']['Replacement (EN)'] == "Not yet in season" or replacing_table_df.loc['season phase']['Replacement (EN)'] == "Land preparation" or replacing_table_df.loc['season phase']['Replacement (EN)'] == "Planting" or replacing_table_df.loc['season phase']['Replacement (EN)'] == "Early growing" or replacing_table_df.loc['season phase']['Replacement (EN)'] == "Growing" or replacing_table_df.loc['season phase']['Replacement (EN)'] == "Maturing"):
            expected_or_nothing = "expected"
            attendue_ou_rien = "attendue"
        elif (replacing_table_df.loc['season phase']['Replacement (EN)'] == "Harvesting" or replacing_table_df.loc['season phase']['Replacement (EN)'] == "Recently finished"):
            expected_or_nothing = ""
            attendue_ou_rien = ""
        else:
            expected_or_nothing = ""
            attendue_ou_rien = ""

        # Saison and attendue ou rien added on the replacementTextKeyPairs
        replacementTextKeyPairs = {'#phone number#': replacing_table_df.loc["phone number"]["Replacement (FR)"],
                                   '#number of digits#': len(str(replacing_table_df.loc["phone number"]["Replacement (FR)"])),
                                   '#age#': replacing_table_df.loc["age"]["Replacement (FR)"],
                                   '#ADMIN1#': replacing_table_df.loc["ADMIN1"]["Replacement (FR)"],
                                   '#ADMIN2#': replacing_table_df.loc["ADMIN2"]["Replacement (FR)"],
                                   '#reference year#': replacing_table_df.loc["reference year"]["Replacement (EN)"],
                                   '#reference year FR#': replacing_table_df.loc["reference year"]["Replacement (FR)"],
                                   '#season#': replacing_table_df.loc["season"]["Replacement (EN)"],
                                   '#season FR#': replacing_table_df.loc["season"]["Replacement (FR)"],
                                   '#season phase#': replacing_table_df.loc['season phase']['Replacement (FR)'],
                                   '#expected or nothing#': expected_or_nothing,
                                   '#expected or nothing FR#': attendue_ou_rien,
                                   '#local measurement units#': replacing_table_df.loc["local measurement unit"]["Replacement (FR)"],
                                   '#currency#': replacing_table_df.loc["currency"]["Replacement (FR)"],
                                   '#devise#': replacing_table_df.loc["currency"]["Replacement (FR)"],
                                   '#MIN AMOUNT#': replacing_table_df.loc["MIN AMOUNT"]["Replacement (FR)"],
                                   #'#MAX SALARY#': replacing_table_df.loc[5]['Replacement'],
                                   '#THRESHOLD#': replacing_table_df.loc["THRESHOLD"]["Replacement (FR)"],
                                   '#local vegetables#': replacing_table_df.loc["local vegetables"]["Replacement (EN)"],
                                   '#local fruits#': replacing_table_df.loc["local fruits"]["Replacement (EN)"],
                                   
                                   '#légumes locaux#': replacing_table_df.loc["local vegetables"]["Replacement (FR)"],
                                   '#fruits locaux#': replacing_table_df.loc["local fruits"]["Replacement (FR)"],
                                   
                                  }

        if pd.isna(replacing_table_df.loc["local measurement unit"]["Replacement (EN)"]) == True:
            print("\nLocal measurement unit is null, the last answer option (#5) has to be deleted manually.")


        #print("replacing_tableagain")
        #print(replacing_table_df)

        # Iterate over the columns and rows, search
        # for the text and replace
        for i in range(number_of_columns):
            for k in range(number_of_rows):
                cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        #newCellValue = replacementTextKeyPairs.get(key)
                        worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue



        # Iterate over the columns and rows, search
        # for the text and replace
        for i in range(number_of_columns_choices):
            for k in range(number_of_rows_choices):
                cellValue = str(worksheet_choices[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        #newCellValue = replacementTextKeyPairs.get(key)
                        worksheet_choices[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue
           
    
    elif language == "ar":
        print("Arabic")

        replacing_table_df = replacing_table_df[['Original','Replacement (EN)','Remplacement (AR)']]
        replacement_list = list(replacing_table_df['Original'].values.tolist())
        replacing_table_df.set_index("Original", inplace = True)

        expected_or_nothing = ""
        expected_or_nothing_ar = ""

        if (replacing_table_df.loc['season phase']["Remplacement (AR)"] == "ليس في الموسم بعد" or replacing_table_df.loc["season phase"]["Remplacement (AR)"] == "إعداد الأرض" or replacing_table_df.loc["season phase"]["Remplacement (AR)"] == "الزرع" or replacing_table_df.loc["season phase"]["Remplacement (AR)"] == "النمو المبكر" or replacing_table_df.loc["season phase"]["Remplacement (AR)"] == "النمو" or replacing_table_df.loc["season phase"]["Remplacement (AR)"] == "النضوج"):
            expected_or_nothing = "expected"
            expected_or_nothing_ar = "متوقع"
        elif replacing_table_df.loc["season phase"]["Remplacement (AR)"] == "الحصاد" or replacing_table_df.loc["season phase"]["Remplacement (AR)"] == "انتهى مؤخرا":
            expected_or_nothing = ""
            expected_or_nothing_ar = ""
        else:
            expected_or_nothing = ""
            expected_or_nothing_ar = ""

        # Saison and attendue ou rien added on the replacementTextKeyPairs
        replacementTextKeyPairs = {'#phone number#': replacing_table_df.loc["phone number"]["Remplacement (AR)"],
                                   '#number of digits#': len(str(replacing_table_df.loc["phone number"]["Remplacement (AR)"])),
                                   '#age#': replacing_table_df.loc["age"]["Remplacement (AR)"],
                                   '#ADMIN1#': replacing_table_df.loc["ADMIN1"]["Replacement (EN)"],
                                   '#ADMIN2#': replacing_table_df.loc["ADMIN2"]["Replacement (EN)"],
                                   '#ADMIN1 AR#': replacing_table_df.loc["ADMIN1"]["Remplacement (AR)"],
                                   '#ADMIN2 AR#': replacing_table_df.loc["ADMIN2"]["Remplacement (AR)"],
                                   #'#reference year#': replacing_table_df.loc["reference year"]["Replacement (EN)"],
                                   #'#reference year AR#': replacing_table_df.loc["reference year"]["Remplacement (AR)"],
                                   '#season#': replacing_table_df.loc["season"]["Replacement (EN)"],
                                   '#season AR#': replacing_table_df.loc["season"]["Remplacement (AR)"],
                                   '#season phase#': replacing_table_df.loc["season phase"]["Replacement (EN)"],
                                   '#expected or nothing#': expected_or_nothing,
                                   '#expected or nothing AR#': expected_or_nothing_ar,
                                   '#local measurement units#': replacing_table_df.loc["local measurement unit "]["Replacement (EN)"],
                                   '#local measurement units AR#': replacing_table_df.loc["local measurement unit "]["Remplacement (AR)"],
                                   '#currency#': replacing_table_df.loc["currency"]["Replacement (EN)"],
                                   '#currency AR#': replacing_table_df.loc["currency"]["Remplacement (AR)"],
                                   '#MIN AMOUNT#': replacing_table_df.loc["MIN AMOUNT"]["Remplacement (AR)"],
                                   '#THRESHOLD#': replacing_table_df.loc["THRESHOLD"]["Remplacement (AR)"],
                                   '#local vegetables#': replacing_table_df.loc["local vegetables"]["Replacement (EN)"],
                                   '#local vegetables AR#': replacing_table_df.loc["local vegetables"]["Remplacement (AR)"],
                                   '#local fruits#': replacing_table_df.loc["local fruits "]["Replacement (EN)"],
                                   '#local fruits AR#': replacing_table_df.loc["local fruits "]["Remplacement (AR)"]
                                  }

        if pd.isna(replacing_table_df.loc["local measurement unit "]["Replacement (EN)"]) == True or pd.isna(replacing_table_df.loc["local measurement unit "]["Remplacement (AR)"]) == True:
            print("\nLocal measurement unit is null, the last answer option (#5) has to be deleted manually.")


        #print("replacing_tableagain")
        #print(replacing_table_df)

        # Iterate over the columns and rows, search
        # for the text and replace
        for i in range(number_of_columns):
            for k in range(number_of_rows):
                cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        #newCellValue = replacementTextKeyPairs.get(key)
                        worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue



        # Iterate over the columns and rows, search
        # for the text and replace
        for i in range(number_of_columns_choices):
            for k in range(number_of_rows_choices):
                cellValue = str(worksheet_choices[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        #newCellValue = replacementTextKeyPairs.get(key)
                        worksheet_choices[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue
    
    
    
    
    
              
    
    elif language == "es":
        print("Spanish")
        replacing_table_df = replacing_table_df[['Original','Replacement (EN)','Replacement (ES)']]
        replacement_list = list(replacing_table_df['Original'].values.tolist())
        replacing_table_df.set_index("Original", inplace = True)

        expected_or_nothing = ""
        expected_or_nothing_es = ""

        if (replacing_table_df.loc['season phase']["Replacement (ES)"] == "Aún no en temporada" or replacing_table_df.loc["season phase"]["Replacement (ES)"] == "Preparación de tierra" or replacing_table_df.loc["season phase"]["Replacement (ES)"] == "plantar" or replacing_table_df.loc["season phase"]["Replacement (ES)"] == "desarrollo temprano" or replacing_table_df.loc["season phase"]["Replacement (ES)"] == "el crecimiento" or replacing_table_df.loc["season phase"]["Replacement (ES)"] == "Madurez"):
            expected_or_nothing = "expected"
            expected_or_nothing_es = "esperado"
        elif replacing_table_df.loc["season phase"]["Replacement (ES)"] == "cosecha" or replacing_table_df.loc["season phase"]["Replacement (ES)"] == "Recientemente terminado":
            expected_or_nothing = ""
            expected_or_nothing_es = ""
        else:
            expected_or_nothing = ""
            expected_or_nothing_es = ""

        # Saison and attendue ou rien added on the replacementTextKeyPairs
        replacementTextKeyPairs = {'#phone number#': replacing_table_df.loc["phone number"]["Replacement (EN)"],
                                   '#number of digits#': len(str(replacing_table_df.loc["phone number"]["Replacement (EN)"])),
                                   '#age#': replacing_table_df.loc["age"]["Replacement (EN)"],
                                   '#ADMIN1#': replacing_table_df.loc["ADMIN1"]["Replacement (EN)"],
                                   '#ADMIN2#': replacing_table_df.loc["ADMIN2"]["Replacement (EN)"],
                                   '#ADMIN1#': replacing_table_df.loc["ADMIN1"]["Replacement (ES)"],
                                   '#ADMIN2#': replacing_table_df.loc["ADMIN2"]["Replacement (ES)"],
                                   '#reference year#': replacing_table_df.loc["reference year"]["Replacement (EN)"],
                                   '#reference year ES#': replacing_table_df.loc["reference year"]["Replacement (ES)"],
                                   '#season#': replacing_table_df.loc["season"]["Replacement (EN)"],
                                   '#season ES#': replacing_table_df.loc["season"]["Replacement (ES)"],
                                   '#season phase#': replacing_table_df.loc["season phase"]["Replacement (EN)"],
                                   '#expected or nothing#': expected_or_nothing,
                                   '#expected or nothing ES#': expected_or_nothing_es,
                                   '#local measurement units#': replacing_table_df.loc["local measurement unit"]["Replacement (EN)"],
                                   '#local measurement units ES#': replacing_table_df.loc["local measurement unit"]["Replacement (ES)"],
                                   '#currency#': replacing_table_df.loc["currency"]["Replacement (EN)"],
                                   '#currency#': replacing_table_df.loc["currency"]["Replacement (ES)"],
                                   '#MIN AMOUNT#': replacing_table_df.loc["MIN AMOUNT"]["Replacement (ES)"],
                                   '#THRESHOLD#': replacing_table_df.loc["THRESHOLD"]["Replacement (ES)"],
                                   '#local vegetables#': replacing_table_df.loc["local vegetables"]["Replacement (EN)"],
                                   '#local vegetables ES#': replacing_table_df.loc["local vegetables"]["Replacement (ES)"],
                                   '#local fruits#': replacing_table_df.loc["local fruits"]["Replacement (EN)"],
                                   '#local fruits ES#': replacing_table_df.loc["local fruits"]["Replacement (ES)"]
                                  }

        if pd.isna(replacing_table_df.loc["local measurement unit"]["Replacement (EN)"]) == True or pd.isna(replacing_table_df.loc["local measurement unit"]["Replacement (ES)"]) == True:
            print("\nLocal measurement unit is null, the last answer option (#5) has to be deleted manually.")


        #print("replacing_tableagain")
        #print(replacing_table_df)

        # Iterate over the columns and rows, search
        # for the text and replace
        for i in range(number_of_columns):
            for k in range(number_of_rows):
                cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        #newCellValue = replacementTextKeyPairs.get(key)
                        worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue



        # Iterate over the columns and rows, search
        # for the text and replace
        for i in range(number_of_columns_choices):
            for k in range(number_of_rows_choices):
                cellValue = str(worksheet_choices[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        #newCellValue = replacementTextKeyPairs.get(key)
                        worksheet_choices[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue
    
    
                        
    workbook.save(questionnaire_file)


def sort_crop_list_by_selection(questionnaire_file):
    crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)

    columns = list(crop_list_df.columns.values.tolist())

    n_of_choices = 0

    language = detect_language(questionnaire_file)

    # if-elif added for handleing lables in French. if section I just copied the previous version. I haven't done anything yet
    # on the elif partof the code.


    if language == "en":

        crop_list_df = crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)']]

        template_crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)
        template_crop_list_df = template_crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)']]

        choices_name = "crop"

        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["choices"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns

        crop_list = crop_list_df['Dataset code'].dropna().values.tolist()
        n_of_crops = len(crop_list)

        # Adding choices
        data = [[np.nan,'777','No crop production'], [np.nan,'888','Don\'t know'], [np.nan,'999','Refused']]
        df = pd.DataFrame(data, columns = ['Select top 10 crops ','Dataset code', 'Label (EN)'])
        crop_list_df = pd.concat([crop_list_df, df], ignore_index=True)

        template_crop_list = template_crop_list_df['Dataset code'].dropna().values.tolist()
        template_n_of_crops = len(template_crop_list)

        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)

        if template_n_of_crops == n_of_crops:

            sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','Dataset code'])

            sorted_crop_list_datasetcodes = sorted_crop_list_df[['Dataset code']].values.tolist()
            sorted_crop_list_label = sorted_crop_list_df[['Label (EN)']].values.tolist()


            workbook = openpyxl.load_workbook(questionnaire_file)
            workbook.sheetnames
            worksheet = workbook["choices"]

            number_of_rows = worksheet.max_row #Number of Rows
            number_of_columns = worksheet.max_column #Number of Columns

            max_needed_number_rows = number_of_rows + n_of_crops + 1 + 3
            j = 0
            k = number_of_rows + 1

            for k in range(max_needed_number_rows):

                if (k > number_of_rows) and j <= n_of_crops + 3:

                    datasetcodes = str(*sorted_crop_list_datasetcodes[j])
                    label = str(*sorted_crop_list_label[j])

                    worksheet[get_column_letter(1)+str(k+1)] = str(choices_name)
                    worksheet[get_column_letter(2)+str(k+1)] = str(datasetcodes)
                    worksheet[get_column_letter(3)+str(k+1)] = str(label)

                    j += 1
                    k += 1

            workbook.save(questionnaire_file)


    elif language == "fr":

        crop_list_df = crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)','Label (FR)']]

        template_crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)
        template_crop_list_df = template_crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)','Label (FR)']]

        choices_name = "crop"

        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["choices"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns

        crop_list = crop_list_df['Dataset code'].dropna().values.tolist()
        n_of_crops = len(crop_list)

        # Adding choices
        data = [[np.nan,'777','No crop production','Pas de production végétale'], [np.nan,'888','Don\'t know','Je ne sais pas'], [np.nan,'999','Refused','Refusé']]
        df = pd.DataFrame(data, columns = ['Select top 10 crops ','Dataset code', 'Label (EN)','Label (FR)'])
        crop_list_df = pd.concat([crop_list_df, df], ignore_index=True)
        #print(crop_list_df)

        template_crop_list = template_crop_list_df['Dataset code'].dropna().values.tolist()
        template_n_of_crops = len(template_crop_list)

        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)

        if template_n_of_crops == n_of_crops:

            sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','Dataset code'])

            sorted_crop_list_datasetcodes = sorted_crop_list_df[['Dataset code']].values.tolist()
            sorted_crop_list_label_EN = sorted_crop_list_df[['Label (EN)']].values.tolist()
            sorted_crop_list_label_FR = sorted_crop_list_df[['Label (FR)']].values.tolist()

            workbook = openpyxl.load_workbook(questionnaire_file)
            workbook.sheetnames
            worksheet = workbook["choices"]

            number_of_rows = worksheet.max_row #Number of Rows
            number_of_columns = worksheet.max_column #Number of Columns

            max_needed_number_rows = number_of_rows + n_of_crops + 1 + 3
            j = 0
            k = number_of_rows + 1

            for k in range(max_needed_number_rows):

                if (k > number_of_rows) and j <= n_of_crops + 3:

                    datasetcodes = str(*sorted_crop_list_datasetcodes[j])
                    label_EN = str(*sorted_crop_list_label_EN[j])
                    label_FR = str(*sorted_crop_list_label_FR[j])

                    worksheet[get_column_letter(1)+str(k+1)] = str(choices_name)
                    worksheet[get_column_letter(2)+str(k+1)] = str(datasetcodes)
                    worksheet[get_column_letter(3)+str(k+1)] = str(label_EN)
                    worksheet[get_column_letter(4)+str(k+1)] = str(label_FR)


                    j += 1
                    k += 1

            workbook.save(questionnaire_file)
            
            

    elif language == "es":

        crop_list_df = crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)','Label (ES)']]

        template_crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)
        template_crop_list_df = template_crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)','Label (ES)']]

        choices_name = "crop"

        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["choices"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns

        crop_list = crop_list_df['Dataset code'].dropna().values.tolist()
        n_of_crops = len(crop_list)

        # Adding choices
        data = [[np.nan,'777','No crop production','Sin producción de cultivos'], [np.nan,'888','Don\'t know','no sé'], [np.nan,'999','Refused','Rechazado']]
        df = pd.DataFrame(data, columns = ['Select top 10 crops ','Dataset code', 'Label (EN)','Label (ES)'])
        crop_list_df = pd.concat([crop_list_df, df], ignore_index=True)
        #print(crop_list_df)

        template_crop_list = template_crop_list_df['Dataset code'].dropna().values.tolist()
        template_n_of_crops = len(template_crop_list)

        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)

        if template_n_of_crops == n_of_crops:

            sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','Dataset code'])

            sorted_crop_list_datasetcodes = sorted_crop_list_df[['Dataset code']].values.tolist()
            sorted_crop_list_label_EN = sorted_crop_list_df[['Label (EN)']].values.tolist()
            sorted_crop_list_label_FR = sorted_crop_list_df[['Label (ES)']].values.tolist()

            workbook = openpyxl.load_workbook(questionnaire_file)
            workbook.sheetnames
            worksheet = workbook["choices"]

            number_of_rows = worksheet.max_row #Number of Rows
            number_of_columns = worksheet.max_column #Number of Columns

            max_needed_number_rows = number_of_rows + n_of_crops + 1 + 3
            j = 0
            k = number_of_rows + 1

            for k in range(max_needed_number_rows):

                if (k > number_of_rows) and j <= n_of_crops + 3:

                    datasetcodes = str(*sorted_crop_list_datasetcodes[j])
                    label_EN = str(*sorted_crop_list_label_EN[j])
                    label_FR = str(*sorted_crop_list_label_FR[j])

                    worksheet[get_column_letter(1)+str(k+1)] = str(choices_name)
                    worksheet[get_column_letter(2)+str(k+1)] = str(datasetcodes)
                    worksheet[get_column_letter(3)+str(k+1)] = str(label_EN)
                    worksheet[get_column_letter(4)+str(k+1)] = str(label_FR)


                    j += 1
                    k += 1

            workbook.save(questionnaire_file)
            
            
        

    elif language == "ar":

        crop_list_df = crop_list_df[['Select top 10 crops / حدد أكبر 10 محاصيل ','Dataset code','Label (EN)','Label (AR)']]

        template_crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)
        template_crop_list_df = template_crop_list_df[['Select top 10 crops / حدد أكبر 10 محاصيل ','Dataset code','Label (EN)','Label (AR)']]

        choices_name = "crop"

        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["choices"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns

        crop_list = crop_list_df['Dataset code'].dropna().values.tolist()
        n_of_crops = len(crop_list)

        # Adding choices
        data = [[np.nan,'777','No crop production','لا إنتاج للمحصول'], [np.nan,'888','Don\'t know','لا أعرف'], [np.nan,'999','Refused','رفض الإجابة']]
        df = pd.DataFrame(data, columns = ['Select top 10 crops / حدد أكبر 10 محاصيل ','Dataset code', 'Label (EN)','Label (AR)'])
        crop_list_df = pd.concat([crop_list_df, df], ignore_index=True)
        #print(crop_list_df)

        template_crop_list = template_crop_list_df['Dataset code'].dropna().values.tolist()
        template_n_of_crops = len(template_crop_list)

        list_choices = crop_list_df['Select top 10 crops / حدد أكبر 10 محاصيل '].dropna().values.tolist()
        n_of_choices = len(list_choices)

        if template_n_of_crops == n_of_crops:

            sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops / حدد أكبر 10 محاصيل ','Dataset code'])

            sorted_crop_list_datasetcodes = sorted_crop_list_df[['Dataset code']].values.tolist()
            sorted_crop_list_label_EN = sorted_crop_list_df[['Label (EN)']].values.tolist()
            sorted_crop_list_label_AR = sorted_crop_list_df[['Label (AR)']].values.tolist()

            workbook = openpyxl.load_workbook(questionnaire_file)
            workbook.sheetnames
            worksheet = workbook["choices"]

            number_of_rows = worksheet.max_row #Number of Rows
            number_of_columns = worksheet.max_column #Number of Columns

            max_needed_number_rows = number_of_rows + n_of_crops + 1 + 3
            j = 0
            k = number_of_rows + 1

            for k in range(max_needed_number_rows):

                if (k > number_of_rows) and j <= n_of_crops + 3:

                    datasetcodes = str(*sorted_crop_list_datasetcodes[j])
                    label_EN = str(*sorted_crop_list_label_EN[j])
                    label_AR = str(*sorted_crop_list_label_AR[j])

                    worksheet[get_column_letter(1)+str(k+1)] = str(choices_name)
                    worksheet[get_column_letter(2)+str(k+1)] = str(datasetcodes)
                    worksheet[get_column_letter(3)+str(k+1)] = str(label_EN)
                    worksheet[get_column_letter(4)+str(k+1)] = str(label_AR)


                    j += 1
                    k += 1

            workbook.save(questionnaire_file)
            
            
    


    return n_of_choices


def sort_crop2_list_by_selection(questionnaire_file):
    crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)

    columns = list(crop_list_df.columns.values.tolist())

    n_of_choices = 0

    language = detect_language(questionnaire_file)

    # if-elif added for handleing lables in French. if section I just copied the previous version. I haven't done anything yet
    # on the elif partof the code.


    if language == "en":

        crop_list_df = crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)']]

        template_crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)
        template_crop_list_df = template_crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)']]

        choices_name = "crop2"

        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["choices"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns

        crop_list = crop_list_df['Dataset code'].dropna().values.tolist()
        n_of_crops = len(crop_list)

        # Adding choices
        data = [[np.nan,'666','No other crop'], [np.nan,'888','Don\'t know'], [np.nan,'999','Refused']]
        df = pd.DataFrame(data, columns = ['Select top 10 crops ','Dataset code', 'Label (EN)'])
        crop_list_df = pd.concat([crop_list_df, df], ignore_index=True)

        template_crop_list = template_crop_list_df['Dataset code'].dropna().values.tolist()
        template_n_of_crops = len(template_crop_list)

        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)

        if template_n_of_crops == n_of_crops:

            sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','Dataset code'])

            sorted_crop_list_datasetcodes = sorted_crop_list_df[['Dataset code']].values.tolist()
            sorted_crop_list_label = sorted_crop_list_df[['Label (EN)']].values.tolist()


            workbook = openpyxl.load_workbook(questionnaire_file)
            workbook.sheetnames
            worksheet = workbook["choices"]

            number_of_rows = worksheet.max_row #Number of Rows
            number_of_columns = worksheet.max_column #Number of Columns

            max_needed_number_rows = number_of_rows + n_of_crops + 1 + 3
            j = 0
            k = number_of_rows + 1

            for k in range(max_needed_number_rows):

                if (k > number_of_rows) and j <= n_of_crops + 3:

                    datasetcodes = str(*sorted_crop_list_datasetcodes[j])
                    label = str(*sorted_crop_list_label[j])

                    worksheet[get_column_letter(1)+str(k+1)] = str(choices_name)
                    worksheet[get_column_letter(2)+str(k+1)] = str(datasetcodes)
                    worksheet[get_column_letter(3)+str(k+1)] = str(label)

                    j += 1
                    k += 1

            workbook.save(questionnaire_file)


    elif language == "fr":

        crop_list_df = crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)','Label (FR)']]

        template_crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)
        template_crop_list_df = template_crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)','Label (FR)']]

        choices_name = "crop2"

        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["choices"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns

        crop_list = crop_list_df['Dataset code'].dropna().values.tolist()
        n_of_crops = len(crop_list)

        # Adding choices
        data = [[np.nan,'666','No other crop','Aucune autre culture'], [np.nan,'888','Don\'t know','Je ne sais pas'], [np.nan,'999','Refused','Refusé']]
        df = pd.DataFrame(data, columns = ['Select top 10 crops ','Dataset code', 'Label (EN)','Label (FR)'])
        crop_list_df = pd.concat([crop_list_df, df], ignore_index=True)
        #print(crop_list_df)

        template_crop_list = template_crop_list_df['Dataset code'].dropna().values.tolist()
        template_n_of_crops = len(template_crop_list)

        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)

        if template_n_of_crops == n_of_crops:

            sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','Dataset code'])

            sorted_crop_list_datasetcodes = sorted_crop_list_df[['Dataset code']].values.tolist()
            sorted_crop_list_label_EN = sorted_crop_list_df[['Label (EN)']].values.tolist()
            sorted_crop_list_label_FR = sorted_crop_list_df[['Label (FR)']].values.tolist()


            workbook = openpyxl.load_workbook(questionnaire_file)
            workbook.sheetnames
            worksheet = workbook["choices"]

            number_of_rows = worksheet.max_row #Number of Rows
            number_of_columns = worksheet.max_column #Number of Columns

            max_needed_number_rows = number_of_rows + n_of_crops + 1 + 3
            j = 0
            k = number_of_rows + 1

            for k in range(max_needed_number_rows):

                if (k > number_of_rows) and j <= n_of_crops + 3:

                    datasetcodes = str(*sorted_crop_list_datasetcodes[j])
                    label_EN = str(*sorted_crop_list_label_EN[j])
                    label_FR = str(*sorted_crop_list_label_FR[j])

                    worksheet[get_column_letter(1)+str(k+1)] = str(choices_name)
                    worksheet[get_column_letter(2)+str(k+1)] = str(datasetcodes)
                    worksheet[get_column_letter(3)+str(k+1)] = str(label_EN)
                    worksheet[get_column_letter(4)+str(k+1)] = str(label_FR)


                    j += 1
                    k += 1

            workbook.save(questionnaire_file)

    elif language == "ar":

        crop_list_df = crop_list_df[['Select top 10 crops / حدد أكبر 10 محاصيل ','Dataset code','Label (EN)','Label (AR)']]

        template_crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)
        template_crop_list_df = template_crop_list_df[['Select top 10 crops / حدد أكبر 10 محاصيل ','Dataset code','Label (EN)','Label (AR)']]

        choices_name = "crop2"

        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["choices"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns

        crop_list = crop_list_df['Dataset code'].dropna().values.tolist()
        n_of_crops = len(crop_list)

        # Adding choices
        data = [[np.nan,'666','No other crop','لا محاصيل أخرى'], [np.nan,'888','Don\'t know','لا أعرف'], [np.nan,'999','Refused','رفض الإجابة']]
        df = pd.DataFrame(data, columns = ['Select top 10 crops / حدد أكبر 10 محاصيل ','Dataset code', 'Label (EN)','Label (AR)'])
        crop_list_df = pd.concat([crop_list_df, df], ignore_index=True)
        #print(crop_list_df)

        template_crop_list = template_crop_list_df['Dataset code'].dropna().values.tolist()
        template_n_of_crops = len(template_crop_list)

        list_choices = crop_list_df['Select top 10 crops / حدد أكبر 10 محاصيل '].dropna().values.tolist()
        n_of_choices = len(list_choices)

        if template_n_of_crops == n_of_crops:

            sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops / حدد أكبر 10 محاصيل ','Dataset code'])

            sorted_crop_list_datasetcodes = sorted_crop_list_df[['Dataset code']].values.tolist()
            sorted_crop_list_label_EN = sorted_crop_list_df[['Label (EN)']].values.tolist()
            sorted_crop_list_label_AR = sorted_crop_list_df[['Label (AR)']].values.tolist()

            workbook = openpyxl.load_workbook(questionnaire_file)
            workbook.sheetnames
            worksheet = workbook["choices"]

            number_of_rows = worksheet.max_row #Number of Rows
            number_of_columns = worksheet.max_column #Number of Columns

            max_needed_number_rows = number_of_rows + n_of_crops + 1 + 3
            j = 0
            k = number_of_rows + 1

            for k in range(max_needed_number_rows):

                if (k > number_of_rows) and j <= n_of_crops + 3:

                    datasetcodes = str(*sorted_crop_list_datasetcodes[j])
                    label_EN = str(*sorted_crop_list_label_EN[j])
                    label_AR = str(*sorted_crop_list_label_AR[j])

                    worksheet[get_column_letter(1)+str(k+1)] = str(choices_name)
                    worksheet[get_column_letter(2)+str(k+1)] = str(datasetcodes)
                    worksheet[get_column_letter(3)+str(k+1)] = str(label_EN)
                    worksheet[get_column_letter(4)+str(k+1)] = str(label_AR)


                    j += 1
                    k += 1

            workbook.save(questionnaire_file) 
            
    elif language == "es":

        crop_list_df = crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)','Label (ES)']]

        template_crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)
        template_crop_list_df = template_crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)','Label (ES)']]

        choices_name = "crop2"

        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["choices"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns

        crop_list = crop_list_df['Dataset code'].dropna().values.tolist()
        n_of_crops = len(crop_list)

        # Adding choices
        data = [[np.nan,'666','No other crop','Ningún otro cultivo'], [np.nan,'888','Don\'t know','no sé'], [np.nan,'999','Refused','Rechazado']]
        df = pd.DataFrame(data, columns = ['Select top 10 crops ','Dataset code', 'Label (EN)','Label (ES)'])
        crop_list_df = pd.concat([crop_list_df, df], ignore_index=True)
        #print(crop_list_df)

        template_crop_list = template_crop_list_df['Dataset code'].dropna().values.tolist()
        template_n_of_crops = len(template_crop_list)

        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)

        if template_n_of_crops == n_of_crops:

            sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','Dataset code'])

            sorted_crop_list_datasetcodes = sorted_crop_list_df[['Dataset code']].values.tolist()
            sorted_crop_list_label_EN = sorted_crop_list_df[['Label (EN)']].values.tolist()
            sorted_crop_list_label_FR = sorted_crop_list_df[['Label (ES)']].values.tolist()


            workbook = openpyxl.load_workbook(questionnaire_file)
            workbook.sheetnames
            worksheet = workbook["choices"]

            number_of_rows = worksheet.max_row #Number of Rows
            number_of_columns = worksheet.max_column #Number of Columns

            max_needed_number_rows = number_of_rows + n_of_crops + 1 + 3
            j = 0
            k = number_of_rows + 1

            for k in range(max_needed_number_rows):

                if (k > number_of_rows) and j <= n_of_crops + 3:

                    datasetcodes = str(*sorted_crop_list_datasetcodes[j])
                    label_EN = str(*sorted_crop_list_label_EN[j])
                    label_FR = str(*sorted_crop_list_label_FR[j])

                    worksheet[get_column_letter(1)+str(k+1)] = str(choices_name)
                    worksheet[get_column_letter(2)+str(k+1)] = str(datasetcodes)
                    worksheet[get_column_letter(3)+str(k+1)] = str(label_EN)
                    worksheet[get_column_letter(4)+str(k+1)] = str(label_FR)


                    j += 1
                    k += 1

            workbook.save(questionnaire_file)

    
            
            

def sort_crop3_list_by_selection(questionnaire_file):
    crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)

    columns = list(crop_list_df.columns.values.tolist())

    n_of_choices = 0

    language = detect_language(questionnaire_file)

    # if-elif added for handleing lables in French. if section I just copied the previous version. I haven't done anything yet
    # on the elif partof the code.


    if language == "en":

        crop_list_df = crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)']]

        template_crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)
        template_crop_list_df = template_crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)']]

        choices_name = "crop3"

        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["choices"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns

        crop_list = crop_list_df['Dataset code'].dropna().values.tolist()
        n_of_crops = len(crop_list)

        # Adding choices
        data = [[np.nan,'888','Don\'t know'], [np.nan,'999','Refused']]
        df = pd.DataFrame(data, columns = ['Select top 10 crops ','Dataset code', 'Label (EN)'])
        crop_list_df = pd.concat([crop_list_df, df], ignore_index=True)

        template_crop_list = template_crop_list_df['Dataset code'].dropna().values.tolist()
        template_n_of_crops = len(template_crop_list)

        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)

        if template_n_of_crops == n_of_crops:

            sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','Dataset code'])

            sorted_crop_list_datasetcodes = sorted_crop_list_df[['Dataset code']].values.tolist()
            sorted_crop_list_label = sorted_crop_list_df[['Label (EN)']].values.tolist()


            workbook = openpyxl.load_workbook(questionnaire_file)
            workbook.sheetnames
            worksheet = workbook["choices"]

            number_of_rows = worksheet.max_row #Number of Rows
            number_of_columns = worksheet.max_column #Number of Columns

            max_needed_number_rows = number_of_rows + n_of_crops + 1 + 2
            j = 0
            k = number_of_rows + 1

            for k in range(max_needed_number_rows):

                if (k > number_of_rows) and j <= n_of_crops + 2:

                    datasetcodes = str(*sorted_crop_list_datasetcodes[j])
                    label = str(*sorted_crop_list_label[j])

                    worksheet[get_column_letter(1)+str(k+1)] = str(choices_name)
                    worksheet[get_column_letter(2)+str(k+1)] = str(datasetcodes)
                    worksheet[get_column_letter(3)+str(k+1)] = str(label)

                    j += 1
                    k += 1

            workbook.save(questionnaire_file)


    elif language == "fr":

        crop_list_df = crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)','Label (FR)']]

        template_crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)
        template_crop_list_df = template_crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)','Label (FR)']]

        choices_name = "crop3"

        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["choices"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns

        crop_list = crop_list_df['Dataset code'].dropna().values.tolist()
        n_of_crops = len(crop_list)

        # Adding choices
        data = [[np.nan,'888','Don\'t know','Je ne sais pas'], [np.nan,'999','Refused','Refusé']]
        df = pd.DataFrame(data, columns = ['Select top 10 crops ','Dataset code', 'Label (EN)','Label (FR)'])
        crop_list_df = pd.concat([crop_list_df, df], ignore_index=True)
        #print(crop_list_df)

        template_crop_list = template_crop_list_df['Dataset code'].dropna().values.tolist()
        template_n_of_crops = len(template_crop_list)

        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)

        if template_n_of_crops == n_of_crops:

            sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','Dataset code'])

            sorted_crop_list_datasetcodes = sorted_crop_list_df[['Dataset code']].values.tolist()
            sorted_crop_list_label_EN = sorted_crop_list_df[['Label (EN)']].values.tolist()
            sorted_crop_list_label_FR = sorted_crop_list_df[['Label (FR)']].values.tolist()

            workbook = openpyxl.load_workbook(questionnaire_file)
            workbook.sheetnames
            worksheet = workbook["choices"]

            number_of_rows = worksheet.max_row #Number of Rows
            number_of_columns = worksheet.max_column #Number of Columns

            max_needed_number_rows = number_of_rows + n_of_crops + 1 + 2
            j = 0
            k = number_of_rows + 1

            for k in range(max_needed_number_rows):

                if (k > number_of_rows) and j <= n_of_crops + 2:

                    datasetcodes = str(*sorted_crop_list_datasetcodes[j])
                    label_EN = str(*sorted_crop_list_label_EN[j])
                    label_FR = str(*sorted_crop_list_label_FR[j])

                    worksheet[get_column_letter(1)+str(k+1)] = str(choices_name)
                    worksheet[get_column_letter(2)+str(k+1)] = str(datasetcodes)
                    worksheet[get_column_letter(3)+str(k+1)] = str(label_EN)
                    worksheet[get_column_letter(4)+str(k+1)] = str(label_FR)

                    j += 1
                    k += 1

            workbook.save(questionnaire_file)
            
            
            

    elif language == "ar":

        crop_list_df = crop_list_df[['Select top 10 crops / حدد أكبر 10 محاصيل ','Dataset code','Label (EN)','Label (AR)']]

        template_crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)
        template_crop_list_df = template_crop_list_df[['Select top 10 crops / حدد أكبر 10 محاصيل ','Dataset code','Label (EN)','Label (AR)']]

        choices_name = "crop3"

        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["choices"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns

        crop_list = crop_list_df['Dataset code'].dropna().values.tolist()
        n_of_crops = len(crop_list)

        # Adding choices
        data = [[np.nan,'888','Don\'t know','لا أعرف'], [np.nan,'999','Refused','رفض الإجابة']]
        df = pd.DataFrame(data, columns = ['Select top 10 crops / حدد أكبر 10 محاصيل ','Dataset code', 'Label (EN)','Label (AR)'])
        crop_list_df = pd.concat([crop_list_df, df], ignore_index=True)
        #print(crop_list_df)

        template_crop_list = template_crop_list_df['Dataset code'].dropna().values.tolist()
        template_n_of_crops = len(template_crop_list)

        list_choices = crop_list_df['Select top 10 crops / حدد أكبر 10 محاصيل '].dropna().values.tolist()
        n_of_choices = len(list_choices)

        if template_n_of_crops == n_of_crops:

            sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops / حدد أكبر 10 محاصيل ','Dataset code'])

            sorted_crop_list_datasetcodes = sorted_crop_list_df[['Dataset code']].values.tolist()
            sorted_crop_list_label_EN = sorted_crop_list_df[['Label (EN)']].values.tolist()
            sorted_crop_list_label_AR = sorted_crop_list_df[['Label (AR)']].values.tolist()

            workbook = openpyxl.load_workbook(questionnaire_file)
            workbook.sheetnames
            worksheet = workbook["choices"]

            number_of_rows = worksheet.max_row #Number of Rows
            number_of_columns = worksheet.max_column #Number of Columns

            max_needed_number_rows = number_of_rows + n_of_crops + 1 + 2
            j = 0
            k = number_of_rows + 1

            for k in range(max_needed_number_rows):

                if (k > number_of_rows) and j <= n_of_crops + 2:

                    datasetcodes = str(*sorted_crop_list_datasetcodes[j])
                    label_EN = str(*sorted_crop_list_label_EN[j])
                    label_AR = str(*sorted_crop_list_label_AR[j])

                    worksheet[get_column_letter(1)+str(k+1)] = str(choices_name)
                    worksheet[get_column_letter(2)+str(k+1)] = str(datasetcodes)
                    worksheet[get_column_letter(3)+str(k+1)] = str(label_EN)
                    worksheet[get_column_letter(4)+str(k+1)] = str(label_AR)


                    j += 1
                    k += 1

            workbook.save(questionnaire_file)



    elif language == "es":

        crop_list_df = crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)','Label (ES)']]

        template_crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)
        template_crop_list_df = template_crop_list_df[['Select top 10 crops ','Dataset code','Label (EN)','Label (ES)']]

        choices_name = "crop3"

        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["choices"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns

        crop_list = crop_list_df['Dataset code'].dropna().values.tolist()
        n_of_crops = len(crop_list)

        # Adding choices
        data = [[np.nan,'888','Don\'t know','no sé'], [np.nan,'999','Refused','Rechazado']]
        df = pd.DataFrame(data, columns = ['Select top 10 crops ','Dataset code', 'Label (EN)','Label (ES)'])
        crop_list_df = pd.concat([crop_list_df, df], ignore_index=True)
        #print(crop_list_df)
        template_crop_list = template_crop_list_df['Dataset code'].dropna().values.tolist()
        template_n_of_crops = len(template_crop_list)

        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)

        if template_n_of_crops == n_of_crops:

            sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','Dataset code'])

            sorted_crop_list_datasetcodes = sorted_crop_list_df[['Dataset code']].values.tolist()
            sorted_crop_list_label_EN = sorted_crop_list_df[['Label (EN)']].values.tolist()
            sorted_crop_list_label_FR = sorted_crop_list_df[['Label (ES)']].values.tolist()

            workbook = openpyxl.load_workbook(questionnaire_file)
            workbook.sheetnames
            worksheet = workbook["choices"]

            number_of_rows = worksheet.max_row #Number of Rows
            number_of_columns = worksheet.max_column #Number of Columns

            max_needed_number_rows = number_of_rows + n_of_crops + 1 + 2
            j = 0
            k = number_of_rows + 1

            for k in range(max_needed_number_rows):

                if (k > number_of_rows) and j <= n_of_crops + 2:

                    datasetcodes = str(*sorted_crop_list_datasetcodes[j])
                    label_EN = str(*sorted_crop_list_label_EN[j])
                    label_FR = str(*sorted_crop_list_label_FR[j])

                    worksheet[get_column_letter(1)+str(k+1)] = str(choices_name)
                    worksheet[get_column_letter(2)+str(k+1)] = str(datasetcodes)
                    worksheet[get_column_letter(3)+str(k+1)] = str(label_EN)
                    worksheet[get_column_letter(4)+str(k+1)] = str(label_FR)

                    j += 1
                    k += 1

            workbook.save(questionnaire_file)
            
            


def insert_adm_reference(questionnaire_file,adm0_iso3):

    admin0_url = "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/Administrative_Boundaries_Reference_(view_layer)/FeatureServer/2/query?where=adm0_ISO3%20%3D%20'" + adm0_iso3 + "'&outFields=adm0_name,adm0_name_local,adm0_ISO3_2d,adm0_ISO3&returnGeometry=false&outSR=4326&f=json"
    admin1_url = "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/Administrative_Boundaries_Reference_(view_layer)/FeatureServer/1/query?where=adm0_ISO3%20%3D%20'" + adm0_iso3 + "'&outFields=adm1_name,adm1_name_local,adm1_pcode,adm0_name,adm0_ISO3&returnGeometry=false&outSR=4326&f=json"
    admin2_url = "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/Administrative_Boundaries_Reference_(view_layer)/FeatureServer/0/query?where=adm0_ISO3%20%3D%20'" + adm0_iso3 + "'&outFields=adm2_name,adm2_name_local,adm2_pcode,adm1_name,adm1_pcode,adm0_name,adm0_ISO3&returnGeometry=false&outSR=4326&f=json"
    admin3_url = "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/Reference_Admin_3/FeatureServer/0/query?where=adm0_ISO3%20%3D%20'"+ adm0_iso3 + "'&outFields=adm3_name,adm3_name_local,adm3_pcode,adm2_name,adm2_pcode,adm1_name,adm1_pcode,adm0_name,adm0_ISO3&returnGeometry=false&outSR=4326&f=json"

    with urllib.request.urlopen(admin0_url) as admin0_url:
        data_adm0 = json.loads(admin0_url.read().decode())
        print(type(data_adm0))
    with urllib.request.urlopen(admin1_url) as admin1_url:
        data_adm1 = json.loads(admin1_url.read().decode())
        print(type(data_adm1))
    with urllib.request.urlopen(admin2_url) as admin2_url:
        data_adm2 = json.loads(admin2_url.read().decode())
        print(type(data_adm2))
    with urllib.request.urlopen(admin3_url) as admin3_url:
        data_adm3 = json.loads(admin3_url.read().decode())
        print(type(data_adm3))



    adm0_df = pd.json_normalize(data_adm0['features'])
    #print (adm0_df.head())
    adm1_df = pd.json_normalize(data_adm1['features'])
    #print (adm1_df.head())
    adm2_df = pd.json_normalize(data_adm2['features'])
    #print (adm2_df.head())
    adm3_df = pd.json_normalize(data_adm3['features'])
    #print (adm3_df.head())


    #adm0_df.rename(columns={'attributes.adm0_name': 'adm0_name',
                           #'attributes.adm0_name_local': 'adm0_name_local',
                           #'attributes.adm0_ISO3_2d': 'adm0_ISO3_2d',
                           #'attributes.adm0_ISO3': 'adm0_ISO3'}, inplace=True)

    adm1_df.rename(columns={'attributes.adm0_name': 'adm0_name',
                           'attributes.adm0_ISO3': 'adm0_ISO3',
                           'attributes.adm1_name': 'adm1_name',
                           'attributes.adm1_name_local': 'adm1_name_local',
                           'attributes.adm1_pcode': 'adm1_pcode'}, inplace=True)


    adm2_df.rename(columns={'attributes.adm0_name': 'adm0_name',
                           'attributes.adm0_ISO3': 'adm0_ISO3',
                           'attributes.adm1_name': 'adm1_name',
                           'attributes.adm1_pcode': 'adm1_pcode',
                           'attributes.adm2_name': 'adm2_name',
                           'attributes.adm2_name_local': 'adm2_name_local',
                           'attributes.adm2_pcode': 'adm2_pcode'}, inplace=True)


    adm3_df.rename(columns={'attributes.adm0_name': 'adm0_name',
                           'attributes.adm0_ISO3': 'adm0_ISO3',
                           'attributes.adm1_name': 'adm1_name',
                           'attributes.adm1_pcode': 'adm1_pcode',
                           'attributes.adm2_name': 'adm2_name',
                           'attributes.adm2_pcode': 'adm2_pcode',
                           'attributes.adm3_name': 'adm3_name',
                           'attributes.adm3_name_local': 'adm3_name_local',
                           'attributes.adm3_pcode': 'adm3_pcode'}, inplace=True)


    workbook = openpyxl.load_workbook(questionnaire_file)
    workbook.sheetnames
    worksheet = workbook["choices"]

    number_of_rows = worksheet.max_row #Number of Rows
    number_of_columns = worksheet.max_column #Number of Columns


    adm1 = adm1_df['adm1_pcode'].dropna().values.tolist()
    nb_adm1 = len(adm1)
    sorted_adm1_df = adm1_df.sort_values(by = ['adm1_pcode'])
    sorted_adm1_list_adm1_pcode = sorted_adm1_df[['adm1_pcode']].values.tolist()
    sorted_adm1_list_adm1_name = sorted_adm1_df[['adm1_name']].values.tolist()


    adm2 = adm2_df['adm2_pcode'].dropna().values.tolist()
    nb_adm2 = len(adm2)
    sorted_adm2_df = adm2_df.sort_values(by = ['adm2_pcode'])
    sorted_adm2_list_adm2_pcode = sorted_adm2_df[['adm2_pcode']].values.tolist()
    sorted_adm2_list_adm2_name = sorted_adm2_df[['adm2_name']].values.tolist()
    sorted_adm2_list_adm1_pcode = sorted_adm2_df[['adm1_pcode']].values.tolist()



    max_needed_number_rows_adm1 = number_of_rows + nb_adm1 + 1
    j = 0
    k = number_of_rows + 1

    for k in range(max_needed_number_rows_adm1):

        if (k > number_of_rows) and j <= nb_adm1:

            adm1_pcode = str(*sorted_adm1_list_adm1_pcode[j])
            adm1_name = str(*sorted_adm1_list_adm1_name[j])

            worksheet[get_column_letter(1)+str(k+1)] = str("admin1")
            worksheet[get_column_letter(2)+str(k+1)] = str(adm1_pcode)
            worksheet[get_column_letter(3)+str(k+1)] = str(adm1_name)
            j += 1
            k += 1

    max_needed_number_rows_adm2 = number_of_rows + nb_adm1 + nb_adm2 + 2
    i = 0
    k = max_needed_number_rows_adm1 + 1


    for k in range(max_needed_number_rows_adm2):

        if (k > max_needed_number_rows_adm1) and i <= nb_adm2:

            adm1_pcode = str(*sorted_adm2_list_adm1_pcode[i])
            adm2_pcode = str(*sorted_adm2_list_adm2_pcode[i])
            adm2_name = str(*sorted_adm2_list_adm2_name[i])


            worksheet[get_column_letter(1)+str(k+1)] = str("admin2")
            worksheet[get_column_letter(2)+str(k+1)] = str(adm2_pcode)
            worksheet[get_column_letter(3)+str(k+1)] = str(adm2_name)
            worksheet[get_column_letter(5)+str(k+1)] = str(adm1_pcode)

            i += 1
            k += 1
    workbook.save(questionnaire_file)

def check_all_domains(questionnaire_file, template_file, result_file):
    questionnaire_df = pd.read_excel(pd.ExcelFile(questionnaire_file), sheet_name='choices').dropna(subset=['name'])
    template_df = pd.read_excel(pd.ExcelFile(template_file), sheet_name='choices').dropna(subset=['name'])
    writer = pd.ExcelWriter(result_file, engine = 'xlsxwriter')
    
    outer_join = pd.merge(template_df, questionnaire_df, how='outer', on=['name', 'name'])

    comparaison_result = pd.DataFrame()
    comparaison_result['name'] = outer_join['name']

    comparaison_result_list_name = pd.DataFrame()
    comparaison_result_list_name['name'] = outer_join['name']
    comparaison_result_list_name['list_name_x'] = outer_join['list_name_x']
    comparaison_result_list_name['list_name_y'] = outer_join['list_name_y']
    comparaison_result_list_name['list_name_match'] = np.where(outer_join['list_name_x'] == outer_join['list_name_y'], 'True', 'False')
    comparaison_result_list_name = comparaison_result_list_name.dropna(subset=['list_name_x','list_name_y'], how='all')
    comparaison_result_list_name.to_excel(writer, sheet_name='list_name_match', engine='xlsxwriter')
    comparaison_result_list_name_list = list(comparaison_result_list_name.loc[((comparaison_result_list_name["list_name_match"] == "False")),['name']].values.tolist())
    if len(comparaison_result_list_name_list) > 0:
        print("\nDETECTING DIFFERENCES IN DOMAINS: list_name")
        print(comparaison_result_list_name_list)

    comparaison_result_label = pd.DataFrame()
    comparaison_result_label['name'] = outer_join['name']
    comparaison_result_label['label::English (en)_x'] = outer_join['label::English (en)_x']
    comparaison_result_label['label::English (en)_y'] = outer_join['label::English (en)_y']
    comparaison_result_label['label::English (en)_match'] = np.where(outer_join['label::English (en)_x'] == outer_join['label::English (en)_y'], 'True', 'False')
    comparaison_result_label = comparaison_result_label.dropna(subset=['label::English (en)_x','label::English (en)_y'], how='all')
    comparaison_result_label.to_excel(writer, sheet_name='labelEnglish (en)_match', engine='xlsxwriter')
    comparaison_result_label_list = list(comparaison_result_label.loc[((comparaison_result_label["label::English (en)_match"] == "False")),['name']].values.tolist())
    if len(comparaison_result_label_list) > 0:
        print("\nDETECTING DIFFERENCES IN DOMAINS: label::English (en)")
        print(comparaison_result_label_list)


    writer.save()

#OLD function that we need to replace with a new one:
def OLD_check_all_questions(questionnaire_file, template_file, result_file):
    questionnaire_df = pd.read_excel(pd.ExcelFile(questionnaire_file), sheet_name='survey').dropna(subset=['name'])
    template_df = pd.read_excel(pd.ExcelFile(template_file), sheet_name='survey').dropna(subset=['name'])
    writer = pd.ExcelWriter(result_file, engine = 'xlsxwriter')
    language = detect_language(questionnaire_file)
    result_brief = ""
    result_details = ""

    outer_join = pd.merge(template_df, questionnaire_df, how='inner', on=['name', 'name'])

    comparaison_result = pd.DataFrame()
    comparaison_result['question_name'] = outer_join['name']

    comparaison_result_type = pd.DataFrame()
    comparaison_result_type['name'] = outer_join['name']
    comparaison_result_type['type_x'] = outer_join['type_x']
    comparaison_result_type['type_y'] = outer_join['type_y']
    comparaison_result_type['type_match'] = np.where(outer_join['type_x'] == outer_join['type_y'], 'True', 'False')
    comparaison_result_type = comparaison_result_type.dropna(subset=['type_x','type_y'], how='all')
    comparaison_result_type.to_excel(writer, sheet_name='type_match', engine='xlsxwriter')
    comparaison_result_type_list = list(comparaison_result_type.loc[((comparaison_result_type["type_match"] == "False")),['name']].values.tolist())

    if len(comparaison_result_type_list) > 0:
        message = "\nDETECTING DIFFERENCES IN QUESTIONS: TYPE\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_type_list)


    comparaison_result_label = pd.DataFrame()
    comparaison_result_label['name'] = outer_join['name']
    comparaison_result_label['label::English (en)_x'] = outer_join['label::English (en)_x']
    comparaison_result_label['label::English (en)_y'] = outer_join['label::English (en)_y']
    comparaison_result_label['label::English (en)_match'] = np.where(outer_join['label::English (en)_x'] == outer_join['label::English (en)_y'], 'True', 'False')
    comparaison_result_label = comparaison_result_label.dropna(subset=['label::English (en)_x','label::English (en)_y'], how='all')
    comparaison_result_label.to_excel(writer, sheet_name='labelEnglish (en)_match', engine='xlsxwriter')
    comparaison_result_label_list = list(comparaison_result_label.loc[((comparaison_result_label["label::English (en)_match"] == "False")),['name']].values.tolist())
    if len(comparaison_result_label_list) > 0:
        message = "\nDETECTING DIFFERENCES IN QUESTIONS: label::English (en)\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_label_list)

        
        
    if language == "ar":
        comparaison_result_label2 = pd.DataFrame()
        comparaison_result_label2['name'] = outer_join['name']
        comparaison_result_label2['label::Arabic (ar)_x'] = outer_join['label::Arabic (ar)_x']
        comparaison_result_label2['label::Arabic (ar)_y'] = outer_join['label::Arabic (ar)_y']
        comparaison_result_label2['label::Arabic (ar)_match'] = np.where(outer_join['label::Arabic (ar)_x'] == outer_join['label::Arabic (ar)_y'], 'True', 'False')
        comparaison_result_label2 = comparaison_result_label2.dropna(subset=['label::Arabic (ar)_x','label::Arabic (ar)_y'], how='all')
        comparaison_result_label2.to_excel(writer, sheet_name='labelArabic_match', engine='xlsxwriter')
        comparaison_result_label2_list = list(comparaison_result_label2.loc[((comparaison_result_label2["label::Arabic (ar)_match"] == "False")),['name']].values.tolist())
        if len(comparaison_result_label2_list) > 0:
            message = "\nDETECTING DIFFERENCES IN QUESTIONS: label::Arabic (ar)\n"
            result_brief = result_brief + message
            result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_label2_list)
            
    elif language == "fr":
        comparaison_result_label2 = pd.DataFrame()
        comparaison_result_label2['name'] = outer_join['name']
        comparaison_result_label2['label::French (fr)_x'] = outer_join['label::French (fr)_x']
        comparaison_result_label2['label::French (fr)_y'] = outer_join['label::French (fr)_y']
        comparaison_result_label2['label::French (fr)_match'] = np.where(outer_join['label::French (fr)_x'] == outer_join['label::French (fr)_y'], 'True', 'False')
        comparaison_result_label2 = comparaison_result_label2.dropna(subset=['label::French (fr)_x','label::French (fr)_y'], how='all')
        comparaison_result_label2.to_excel(writer, sheet_name='labelFrench_match', engine='xlsxwriter')
        comparaison_result_label2_list = list(comparaison_result_label2.loc[((comparaison_result_label2["label::French (fr)_match"] == "False")),['name']].values.tolist())
        if len(comparaison_result_label2_list) > 0:
            message = "\nDETECTING DIFFERENCES IN QUESTIONS: label::French (fr)\n"
            result_brief = result_brief + message
            result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_label2_list)
            
    #else:
        
        #comparaison_result_label2 = pd.DataFrame()
        #comparaison_result_label2['name'] = outer_join['name']
        #comparaison_result_label2['label::_x'] = outer_join['label::_x']
        #comparaison_result_label2['label::_y'] = outer_join['label::_y']
        #comparaison_result_label2['label::_match'] = np.where(outer_join['label::_x'] == outer_join['label::_y'], 'True', 'False')
        #comparaison_result_label2 = comparaison_result_label2.dropna(subset=['label::_x','label::_y'], how='all')
        #comparaison_result_label2.to_excel(writer, sheet_name='label2_match', engine='xlsxwriter')
        #comparaison_result_label2_list = list(comparaison_result_label2.loc[((comparaison_result_label2["label::_match"] == "False")),['name']].values.tolist())
        #if len(comparaison_result_label2_list) > 0:
        #    message = "\nDETECTING DIFFERENCES IN QUESTIONS: label::\n"
        #    result_brief = result_brief + message
        #    result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_label2_list)


    comparaison_result_hintEnglish = pd.DataFrame()
    comparaison_result_hintEnglish['name'] = outer_join['name']
    comparaison_result_hintEnglish['hint::English (en)_x'] = outer_join['hint::English (en)_x']
    comparaison_result_hintEnglish['hint::English (en)_y'] = outer_join['hint::English (en)_y']
    comparaison_result_hintEnglish['hint::English (en)_match'] = np.where(outer_join['hint::English (en)_x'] == outer_join['hint::English (en)_y'], 'True', 'False')
    comparaison_result_hintEnglish = comparaison_result_hintEnglish.dropna(subset=['hint::English (en)_x','hint::English (en)_y'], how='all')
    comparaison_result_hintEnglish.to_excel(writer, sheet_name='hintEnglish_match', engine='xlsxwriter')
    comparaison_result_hintEnglish_list = list(comparaison_result_hintEnglish.loc[((comparaison_result_hintEnglish["hint::English (en)_match"] == "False")),['name']].values.tolist())
    if len(comparaison_result_hintEnglish_list) > 0:
        message = "\nDETECTING DIFFERENCES IN QUESTIONS: hint::English (en)\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_hintEnglish_list)

        
    if language == "ar":
        comparaison_result_hint = pd.DataFrame()
        comparaison_result_hint['name'] = outer_join['name']
        comparaison_result_hint['hint::Arabic (ar)_x'] = outer_join['hint::Arabic (ar)_x']
        comparaison_result_hint['hint::Arabic (ar)_y'] = outer_join['hint::Arabic (ar)_y']
        comparaison_result_hint['hint::Arabic (ar)_match'] = np.where(outer_join['hint::Arabic (ar)_x'] == outer_join['hint::Arabic (ar)_y'], 'True', 'False')
        comparaison_result_hint = comparaison_result_hint.dropna(subset=['hint::Arabic (ar)_x','hint::Arabic (ar)_y'], how='all')
        comparaison_result_hint.to_excel(writer, sheet_name='hintArabic_match', engine='xlsxwriter')
        comparaison_result_hint_list = list(comparaison_result_hint.loc[((comparaison_result_hint["hint::Arabic (ar)_match"] == "False")),['name']].values.tolist())
        if len(comparaison_result_hint_list) > 0:
            message = "\nDETECTING DIFFERENCES IN QUESTIONS: hint::Arabic (ar)\n"
            result_brief = result_brief + message
            result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_hint_list)
        
    elif language == "fr":
        comparaison_result_hint = pd.DataFrame()
        comparaison_result_hint['name'] = outer_join['name']
        comparaison_result_hint['hint::French (fr)_x'] = outer_join['hint::French (fr)_x']
        comparaison_result_hint['hint::French (fr)_y'] = outer_join['hint::French (fr)_y']
        comparaison_result_hint['hint::French (fr)_match'] = np.where(outer_join['hint::French (fr)_x'] == outer_join['hint::French (fr)_y'], 'True', 'False')
        comparaison_result_hint = comparaison_result_hint.dropna(subset=['hint::French (fr)_x','hint::French (fr)_y'], how='all')
        comparaison_result_hint.to_excel(writer, sheet_name='hintFrench_match', engine='xlsxwriter')
        comparaison_result_hint_list = list(comparaison_result_hint.loc[((comparaison_result_hint["hint::French (fr)_match"] == "False")),['name']].values.tolist())
        if len(comparaison_result_hint_list) > 0:
            message = "\nDETECTING DIFFERENCES IN QUESTIONS: hint::French (fr)\n"
            result_brief = result_brief + message
            result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_hint_list)
        
        

    comparaison_result_required_match = pd.DataFrame()
    comparaison_result_required_match['name'] = outer_join['name']
    comparaison_result_required_match['required_x'] = outer_join['required_x']
    comparaison_result_required_match['required_y'] = outer_join['required_y']
    comparaison_result_required_match['required_match'] = np.where(outer_join['required_x'] == outer_join['required_y'], 'True', 'False')
    comparaison_result_required_match = comparaison_result_required_match.dropna(subset=['required_x','required_y'], how='all')
    comparaison_result_required_match.to_excel(writer, sheet_name='required_match', engine='xlsxwriter')
    comparaison_result_required_match_list = list(comparaison_result_required_match.loc[((comparaison_result_required_match["required_match"] == "False")),['name']].values.tolist())
    if len(comparaison_result_required_match_list) > 0:
        message = "\nDETECTING DIFFERENCES IN QUESTIONS: required\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_required_match_list)


    comparaison_result_appearance_match = pd.DataFrame()
    comparaison_result_appearance_match['name'] = outer_join['name']
    comparaison_result_appearance_match['appearance_x'] = outer_join['appearance_x']
    comparaison_result_appearance_match['appearance_y'] = outer_join['appearance_y']
    comparaison_result_appearance_match['appearance_match'] = np.where(outer_join['appearance_x'] == outer_join['appearance_y'], 'True', 'False')
    comparaison_result_appearance_match = comparaison_result_appearance_match.dropna(subset=['appearance_x','appearance_y'], how='all')
    comparaison_result_appearance_match.to_excel(writer, sheet_name='appearance_match', engine='xlsxwriter')
    comparaison_result_appearance_match_list = list(comparaison_result_appearance_match.loc[((comparaison_result_appearance_match["appearance_match"] == "False")),['name']].values.tolist())
    if len(comparaison_result_appearance_match_list) > 0:
        message = "\nDETECTING DIFFERENCES IN QUESTIONS: appearance\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_appearance_match_list)

    comparaison_result_constraint_match = pd.DataFrame()
    comparaison_result_constraint_match['name'] = outer_join['name']
    comparaison_result_constraint_match['constraint_x'] = outer_join['constraint_x']
    comparaison_result_constraint_match['constraint_y'] = outer_join['constraint_y']
    comparaison_result_constraint_match['constraint_match'] = np.where(outer_join['constraint_x'] == outer_join['constraint_y'], 'True', 'False')
    comparaison_result_constraint_match = comparaison_result_constraint_match.dropna(subset=['constraint_x','constraint_y'], how='all')
    comparaison_result_constraint_match.to_excel(writer, sheet_name='constraint_match', engine='xlsxwriter')
    comparaison_result_constraint_match_list = list(comparaison_result_constraint_match.loc[((comparaison_result_constraint_match["constraint_match"] == "False")),['name']].values.tolist())
    if len(comparaison_result_constraint_match_list) > 0:
        message = "\nDETECTING DIFFERENCES IN QUESTIONS: constraint\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_constraint_match_list)

    #comparaison_result_constraint_message_match = pd.DataFrame()
    #comparaison_result_constraint_message_match['name'] = outer_join['name']
    #comparaison_result_constraint_message_match['constraint_message_x'] = outer_join['constraint_message_x']
    #comparaison_result_constraint_message_match['constraint_message_y'] = outer_join['constraint_message_y']
    #comparaison_result_constraint_message_match['constraint_message_match'] = np.where(outer_join['constraint_message_x'] == outer_join['constraint_message_y'], 'True', 'False')
    #comparaison_result_constraint_message_match = comparaison_result_constraint_message_match.dropna(subset=['constraint_message_x','constraint_message_y'], how='all')
    #comparaison_result_constraint_message_match.to_excel(writer, sheet_name='constraint_message_match', engine='xlsxwriter')
    #comparaison_result_constraint_message_match_list = list(comparaison_result_constraint_message_match.loc[((comparaison_result_constraint_message_match["constraint_message_match"] == "False")),['name']].values.tolist())
    #if len(comparaison_result_constraint_message_match_list) > 0:
    #    message = "\nDETECTING DIFFERENCES IN QUESTIONS: constraint_message\n"
    #    result_brief = result_brief + message
    #    result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_constraint_message_match_list)

    #comparaison_result_relevant_match = pd.DataFrame()
    #comparaison_result_relevant_match['name'] = outer_join['name']
    #comparaison_result_relevant_match['relevant_x'] = outer_join['relevant_x']
    #comparaison_result_relevant_match['relevant_y'] = outer_join['relevant_y']
    #comparaison_result_relevant_match['relevant_match'] = np.where(outer_join['relevant_x'] == outer_join['relevant_y'], 'True', 'False')
    #comparaison_result_relevant_match = comparaison_result_relevant_match.dropna(subset=['relevant_x','relevant_y'], how='all')
    #comparaison_result_relevant_match.to_excel(writer, sheet_name='relevant_match', engine='xlsxwriter')
    #comparaison_result_relevant_match_list = list(comparaison_result_relevant_match.loc[((comparaison_result_relevant_match["relevant_match"] == "False")),['name']].values.tolist())
    #if len(comparaison_result_relevant_match_list) > 0:
    #    message =  "\nDETECTING DIFFERENCES IN QUESTIONS: relevant\n"
    #    result_brief = result_brief + message
    #    result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_relevant_match_list)


    comparaison_result_choice_filter_match = pd.DataFrame()
    comparaison_result_choice_filter_match['name'] = outer_join['name']
    comparaison_result_choice_filter_match['choice_filter_x'] = outer_join['choice_filter_x']
    comparaison_result_choice_filter_match['choice_filter_y'] = outer_join['choice_filter_y']
    comparaison_result_choice_filter_match['choice_filter_match'] = np.where(outer_join['choice_filter_x'] == outer_join['choice_filter_y'], 'True', 'False')
    comparaison_result_choice_filter_match = comparaison_result_choice_filter_match.dropna(subset=['choice_filter_x','choice_filter_y'], how='all')
    comparaison_result_choice_filter_match.to_excel(writer, sheet_name='choice_filter_match', engine='xlsxwriter')
    comparaison_result_choice_filter_match_list = list(comparaison_result_choice_filter_match.loc[((comparaison_result_choice_filter_match["choice_filter_match"] == "False")),['name']].values.tolist())
    if len(comparaison_result_choice_filter_match_list) > 0:
        message = "\nDETECTING DIFFERENCES IN QUESTIONS: choice_filter\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_choice_filter_match_list)

    #comparaison_result_given_name_match = pd.DataFrame()
    #comparaison_result_given_name_match['name'] = outer_join['name']
    #comparaison_result_given_name_match['$given_name_x'] = outer_join['$given_name_x']
    #comparaison_result_given_name_match['$given_name_y'] = outer_join['$given_name_y']
    #comparaison_result_given_name_match['$given_name_match'] = np.where(outer_join['$given_name_x'] == outer_join['$given_name_y'], 'True', 'False')
    #comparaison_result_given_name_match = comparaison_result_given_name_match.dropna(subset=['$given_name_x','$given_name_y'], how='all')
    #comparaison_result_given_name_match.to_excel(writer, sheet_name='given_name_match', engine='xlsxwriter')
    #comparaison_result_given_name_match_list = list(comparaison_result_given_name_match.loc[((comparaison_result_given_name_match["$given_name_match"] == "False")),['name']].values.tolist())
    #if len(comparaison_result_given_name_match_list) > 0:
    #    message = "\nDETECTING DIFFERENCES IN QUESTIONS: parameters\n"
    #    result_brief = result_brief + message
    #    result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_given_name_match_list)

    comparaison_result_calculation_match = pd.DataFrame()
    comparaison_result_calculation_match['name'] = outer_join['name']
    comparaison_result_calculation_match['calculation_x'] = outer_join['calculation_x']
    comparaison_result_calculation_match['calculation_y'] = outer_join['calculation_y']
    comparaison_result_calculation_match['calculation_match'] = np.where(outer_join['calculation_x'] == outer_join['calculation_y'], 'True', 'False')
    comparaison_result_calculation_match = comparaison_result_calculation_match.dropna(subset=['calculation_x','calculation_y'], how='all')
    comparaison_result_calculation_match.to_excel(writer, sheet_name='calculation_match', engine='xlsxwriter')
    comparaison_result_calculation_match_list = list(comparaison_result_calculation_match.loc[((comparaison_result_calculation_match["calculation_match"] == "False")),['name']].values.tolist())
    if len(comparaison_result_calculation_match_list) > 0:
        message = "\nDETECTING DIFFERENCES IN QUESTIONS: calculation\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_calculation_match_list)


    #comparaison_result['Unnamed: 14_match'] = np.where(outer_join['Unnamed: 14_x'] == outer_join['Unnamed: 14_y'], 'True', 'False')
    #comparaison_result['Unnamed: 15_match'] = np.where(outer_join['Unnamed: 15_x'] == outer_join['Unnamed: 15_y'], 'True', 'False')
    #comparaison_result['Unnamed: 16_match'] = np.where(outer_join['Unnamed: 16_x'] == outer_join['Unnamed: 16_y'], 'True', 'False')
    #comparaison_result['Unnamed: 17_match'] = np.where(outer_join['Unnamed: 17_x'] == outer_join['Unnamed: 17_y'], 'True', 'False')


    comparaison_result_Mandatory_match = pd.DataFrame()
    comparaison_result_Mandatory_match['name'] = outer_join['name']
    comparaison_result_Mandatory_match['Mandatory_x'] = outer_join['Mandatory _x']
    comparaison_result_Mandatory_match['Mandatory_y'] = outer_join['Mandatory _y']
    comparaison_result_Mandatory_match['Mandatory_match'] = np.where(outer_join['Mandatory _x'] == outer_join['Mandatory _y'], 'True', 'False')
    comparaison_result_Mandatory_match = comparaison_result_Mandatory_match.dropna(subset=['Mandatory_x','Mandatory_y'], how='all')
    comparaison_result_Mandatory_match.to_excel(writer, sheet_name='Mandatory_match', engine='xlsxwriter')
    comparaison_result_Mandatory_match_list = list(comparaison_result_Mandatory_match.loc[((comparaison_result_Mandatory_match["Mandatory_match"] == "False")),['name']].values.tolist())
    if len(comparaison_result_Mandatory_match_list) > 0:
        message =  "\nDETECTING DIFFERENCES IN QUESTIONS: Mandatory\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_Mandatory_match_list)

    #comparaison_result['Unnamed: 19_match'] = np.where(outer_join['Unnamed: 19_x'] == outer_join['Unnamed: 19_y'], 'True', 'False')


    #comparaison_result_Question_duration_match = pd.DataFrame()
    #comparaison_result_Question_duration_match['name'] = outer_join['name']
    #comparaison_result_Question_duration_match['Question avg duration (sec)_x'] = outer_join['Question avg duration (sec)_x']
    #comparaison_result_Question_duration_match['Question avg duration (sec)_y'] = outer_join['Question avg duration (sec)_y']
    #comparaison_result_Question_duration_match['Question avg duration (sec)_match'] = np.where(outer_join['Question avg duration (sec)_x'] == outer_join['Question avg duration (sec)_y'], 'True', 'False')
    #comparaison_result_Question_duration_match = comparaison_result_Question_duration_match.dropna(subset=['Question avg duration (sec)_x','Question avg duration (sec)_y'], how='all')
    #comparaison_result_Question_duration_match.to_excel(writer, sheet_name='Question avg duration_match', engine='xlsxwriter')
    #comparaison_result_Question_duration_match_list = list(comparaison_result_Question_duration_match.loc[((comparaison_result_Question_duration_match["Question avg duration (sec)_match"] == "False")),['name']].values.tolist())
    #if len(comparaison_result_Question_duration_match_list) > 0:
    #    message = "\nDETECTING DIFFERENCES IN QUESTIONS: Question avg duration (sec)\n"
    #    result_brief = result_brief + message
    #    result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_Question_duration_match_list)


    #comparaison_result_newQuestion_duration_match = pd.DataFrame()
    #comparaison_result_newQuestion_duration_match['name'] = outer_join['name']
    #comparaison_result_newQuestion_duration_match['New question duration (sec)_x'] = outer_join['New question duration (sec)_x']
    #comparaison_result_newQuestion_duration_match['New question duration (sec)_y'] = outer_join['New question duration (sec)_y']
    #comparaison_result_newQuestion_duration_match['New question duration (sec)_match'] = np.where(outer_join['New question duration (sec)_x'] == outer_join['New question duration (sec)_y'], 'True', 'False')
    #comparaison_result_newQuestion_duration_match = comparaison_result_newQuestion_duration_match.dropna(subset=['New question duration (sec)_x','New question duration (sec)_y'], how='all')
    #comparaison_result_newQuestion_duration_match.to_excel(writer, sheet_name='New question duration_match', engine='xlsxwriter')
    #comparaison_result_newQuestion_duration_match_list = list(comparaison_result_newQuestion_duration_match.loc[((comparaison_result_newQuestion_duration_match["New question duration (sec)_match"] == "False")),['name']].values.tolist())
    #if len(comparaison_result_newQuestion_duration_match_list) > 0:
    #    message = "\nDETECTING DIFFERENCES IN QUESTIONS: New question duration (sec)\n"
    #    result_brief = result_brief + message
    #    result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_newQuestion_duration_match_list)


    #comparaison_result_estimated_perc_repondents = pd.DataFrame()
    #comparaison_result_estimated_perc_repondents['name'] = outer_join['name']
    #comparaison_result_estimated_perc_repondents['Estimated percentage of repondents_x'] = outer_join['Estimated percentage of repondents_x']
    #comparaison_result_estimated_perc_repondents['Estimated percentage of repondents_y'] = outer_join['Estimated percentage of repondents_y']
    #comparaison_result_estimated_perc_repondents['Estimated percentage of repondents_match'] = np.where(outer_join['Estimated percentage of repondents_x'] == outer_join['Estimated percentage of repondents_y'], 'True', 'False')
    #comparaison_result_estimated_perc_repondents = comparaison_result_estimated_perc_repondents.dropna(subset=['Estimated percentage of repondents_x','Estimated percentage of repondents_y'], how='all')
    #comparaison_result_estimated_perc_repondents.to_excel(writer, sheet_name='Estimated perc repondents_match', engine='xlsxwriter')
    #comparaison_result_estimated_perc_repondents_list = list(comparaison_result_estimated_perc_repondents.loc[((comparaison_result_estimated_perc_repondents["Estimated percentage of repondents_match"] == "False")),['name']].values.tolist())
    #if len(comparaison_result_estimated_perc_repondents_list) > 0:
    #    message = "\nDETECTING DIFFERENCES IN QUESTIONS: Estimated percentage of repondents\n"
    #    result_brief = result_brief + message
    #    result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_estimated_perc_repondents_list)


    #comparaison_result_weighted_time_match = pd.DataFrame()
    #comparaison_result_weighted_time_match['name'] = outer_join['name']
    #comparaison_result_weighted_time_match['Weighted time_x'] = outer_join['Weighted time_x']
    #comparaison_result_weighted_time_match['Weighted time_y'] = outer_join['Weighted time_y']
    #comparaison_result_weighted_time_match['Weighted time_match'] = np.where(outer_join['Weighted time_x'] == outer_join['Weighted time_y'], 'True', 'False')
    #comparaison_result_weighted_time_match = comparaison_result_weighted_time_match.dropna(subset=['Weighted time_x','Weighted time_y'], how='all')
    #comparaison_result_weighted_time_match.to_excel(writer, sheet_name='Weighted time_match', engine='xlsxwriter')
    #comparaison_result_weighted_time_match_list = list(comparaison_result_weighted_time_match.loc[((comparaison_result_weighted_time_match["Weighted time_match"] == "False")),['name']].values.tolist())
    #if len(comparaison_result_weighted_time_match_list) > 0:
    #    message = "\nDETECTING DIFFERENCES IN QUESTIONS: Weighted time\n"
    #    result_brief = result_brief + message
    #    result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_weighted_time_match_list)


    #comparaison_result_weighted_time_core = pd.DataFrame()
    #comparaison_result_weighted_time_core['name'] = outer_join['name']
    #comparaison_result_weighted_time_core['Weighted time for core vars_x'] = outer_join['Weighted time for core vars_x']
    #comparaison_result_weighted_time_core['Weighted time for core vars_y'] = outer_join['Weighted time for core vars_y']
    #comparaison_result_weighted_time_core['Weighted time for core vars_match'] = np.where(outer_join['Weighted time for core vars_x'] == outer_join['Weighted time for core vars_y'], 'True', 'False')
    #comparaison_result_weighted_time_core = comparaison_result_weighted_time_core.dropna(subset=['Weighted time for core vars_x','Weighted time for core vars_y'], how='all')
    #comparaison_result_weighted_time_core.to_excel(writer, sheet_name='Weighted timecore vars', engine='xlsxwriter')
    #comparaison_result_weighted_time_core_list = list(comparaison_result_weighted_time_core.loc[((comparaison_result_weighted_time_core["Weighted time for core vars_match"] == "False")),['name']].values.tolist())
    #if len(comparaison_result_weighted_time_core_list) > 0:
    #    message = "\nDETECTING DIFFERENCES IN QUESTIONS: Weighted time for core vars\n"
    #    result_brief = result_brief + message
    #    result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_weighted_time_core_list)

    #comparaison_result['Unnamed: 25_match'] = np.where(outer_join['Unnamed: 25_x'] == outer_join['Unnamed: 25_y'], 'True', 'False')

    #writer.save()
    writer.close()
    return result_brief, result_details



def check_all_questions(questionnaire_file, template_file, result_file):
    questionnaire_df = pd.read_excel(pd.ExcelFile(questionnaire_file), sheet_name='survey').dropna(subset=['name'])
    template_df = pd.read_excel(pd.ExcelFile(template_file), sheet_name='survey').dropna(subset=['name'])
    writer = pd.ExcelWriter(result_file, engine='xlsxwriter')
    language = detect_language(questionnaire_file)
    result_brief = ""
    result_details = ""

    merged = pd.merge(template_df, questionnaire_df, how='inner', on='name', suffixes=('_x', '_y'))

    def compare_and_store(field, sheet_name=None):
        nonlocal result_brief, result_details
        if field + '_x' in merged.columns and field + '_y' in merged.columns:
            df = pd.DataFrame()
            df['name'] = merged['name']
            df[field + '_x'] = merged[field + '_x']
            df[field + '_y'] = merged[field + '_y']
            df[field + '_match'] = np.where(df[field + '_x'] == df[field + '_y'], 'True', 'False')
            df = df[df[field + '_match'] == 'False'].dropna(subset=[field + '_x', field + '_y'], how='all')
            if not df.empty:
                df.to_excel(writer, sheet_name=(sheet_name or field)[:31], index=False)
                message = f"\nDETECTING DIFFERENCES IN QUESTIONS: {field}\n"
                result_brief += message
                result_details += message + '; '.join(df['name'].astype(str).tolist())

    # Universal fields
    compare_and_store('type')
    compare_and_store('label::English (en)', 'labelEnglish_match')
    compare_and_store('hint::English (en)', 'hintEnglish_match')
    compare_and_store('required')
    compare_and_store('appearance')
    compare_and_store('constraint')
    compare_and_store('choice_filter')
    compare_and_store('calculation')
    compare_and_store('Mandatory ', 'Mandatory_match')

    # Language-specific comparisons
    if language == "ar":
        compare_and_store('label::Arabic (ar)', 'labelArabic_match')
        compare_and_store('hint::Arabic (ar)', 'hintArabic_match')
    elif language == "fr":
        compare_and_store('label::French (fr)', 'labelFrench_match')
        compare_and_store('hint::French (fr)', 'hintFrench_match')

    writer.close()
    return result_brief, result_details

















def check_questionnaire_duration(questionnaire_file,template_file):
    questionnaire_df = pd.read_excel(pd.ExcelFile(questionnaire_file), sheet_name='survey').dropna(subset=['name'])
    template_df = pd.read_excel(pd.ExcelFile(template_file), sheet_name='survey').dropna(subset=['name'])

    weighted_time_df = questionnaire_df.loc[((questionnaire_df["type"] != "start") &
                                             (questionnaire_df["type"] != "end") &
                                             (questionnaire_df["type"] != "today") &
                                             (questionnaire_df["type"] != "deviceid") &
                                             (questionnaire_df["type"] != "calculate") &
                                             (questionnaire_df["type"] != "begin_group") &
                                             (questionnaire_df["type"] != "begin group") &
                                             (questionnaire_df["type"] != "end_group")) , ['name', 'type','New question duration (sec)','Estimated percentage of repondents']]

    weighted_time_df = weighted_time_df.dropna(subset=['type'], how='all')
    weighted_time_df['Weighted time'] = weighted_time_df['New question duration (sec)'].astype(float) * weighted_time_df['Estimated percentage of repondents'].astype(float)
    weighted_time_df['Weighted time'] = weighted_time_df['Weighted time'].fillna(17)

    total_weighted_time = weighted_time_df['Weighted time'].astype(float).sum()

    return total_weighted_time