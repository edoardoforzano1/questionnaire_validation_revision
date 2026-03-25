import pandas as pd
import xlrd
#import arcpy
import openpyxl
from pandas import ExcelWriter
from datetime import datetime
import os
import numpy as np
#from arcgis.gis import GIS
#from arcgis.features import SpatialDataFrame
from shutil import copyfile
from openpyxl.utils.cell import get_column_letter
import urllib.request, json

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font



## Function to detect the language from the filename
def detect_language(file_name):
    # Initialize an empty string to store the detected language
    language = ""
    
    # Check for language indicators in the filename and set the language variable accordingly
    if "_en_" in file_name.lower():
        language = "en"
    elif "_fr_" in file_name.lower():
        language = "fr"
    elif "_es_" in file_name.lower():
        language = "es"
    elif "_ar_" in file_name.lower():
        language = "ar"
    elif "_pt_" in file_name.lower():
        language = "pt"        
        
    else:
        print("Please include the language in the filename: %s " % file_name)
        language = "en"#"Please include the language inside the filename"
    return language


# Function to detect the appropriate template based on the language and template version
def detect_template(template_version,file_name):
    # Use the detect_language function to get the language
    language = detect_language(file_name)
    
    # Create the template filename based on the detected language and template version
    if language == "en":
        template_questionnaire_file = r'household_questionnaire_geopoll_EN_template_' + template_version + '_ISO3.xlsx'
    elif language == "fr":
        template_questionnaire_file = r'household_questionnaire_geopoll_FR_template_' + template_version + '_ISO3.xlsx'
    elif language == "es":
        template_questionnaire_file = r'household_questionnaire_geopoll_ES_template_' + template_version + '_ISO3.xlsx'
    elif language == "ar":
        template_questionnaire_file = r'household_questionnaire_geopoll_AR_template_' + template_version + '_ISO3.xlsx'
    elif language == "pt":
        template_questionnaire_file = r'household_questionnaire_geopoll_PT_template_' + template_version + '_ISO3.xlsx'
    else:
        print("Please include the language in the filename: %s " % country_template)
    return template_questionnaire_file


def importallsheets(in_excel, out_gdb):
    # Print the maximum counter value; this limits the number of sheets to be imported
    print("max_counter: %s" % max_counter)
    
    # Initialize a counter variable to keep track of the number of sheets processed
    counter = 0
    # Open the Excel workbook using xlrd
    workbook = xlrd.open_workbook(in_excel)
    
    # Get the names of all sheets in the workbook
    sheets = [sheet.name for sheet in workbook.sheets()]

    print('{} sheets found: {}'.format(len(sheets), ','.join(sheets)))
    
    # Loop through each sheet in the workbook
    for sheet in sheets:
        # Increment the counter
        counter +=1
        
        # Only process sheets if the counter is less than or equal to the max_counter
        if counter <= max_counter:
            # Create the output table name by joining the Geodatabase path and the sheet name
            # Also validate the table name to ensure it meets Geodatabase naming conventions
            out_table = os.path.join(
                out_gdb,
                arcpy.ValidateTableName(
                    "{0}".format(sheet),
                    out_gdb))

            print('Converting {} to {}'.format(sheet, out_table))

            # Perform the conversion from Excel to table in the Geodatabase
            arcpy.ExcelToTable_conversion(in_excel, out_table, sheet)


# Function to create a dictionary based on a table in a Geodatabase (GDB)
def make_attribute_dict(fc, code_field, value_field):
    # Initialize an empty dictionary to store attribute mappings
    attdict = {}
    
    # Use ArcGIS Data Access (arcpy.da) SearchCursor to loop through the table
    with arcpy.da.SearchCursor(fc, [code_field, value_field]) as cursor:
        for row in cursor:
            # Populate the dictionary with code as the key and value as the value
            attdict[row[0]] = row[1]
    return attdict

# Function to improve and standardize the formatting of category descriptions
def fix_category_formatting(category):
    # Perform various string replacements and formatting adjustments
     return category.replace("[","(").replace("]",")").replace("(specify)","").replace("/ ",", ").capitalize().replace("adps","ADPs").replace("idp","IDP").replace("covid","COVID").replace(" , ",", ").replace("staplec","staple")

# Function to count the number of questions based on the 'Q Name' column in the 'survey' sheet of the questionnaire Excel file  count the number of questions based on the 'Q Name' column in the 'survey' sheet of the questionnaire Excel file
def count_number_of_questions_qname(questionnaire_file):
    # Read the 'survey' sheet from the Excel file and skip the first 2 rows
    questionnaire_df = pd.read_excel(pd.ExcelFile(questionnaire_file), sheet_name='survey',skiprows=2)
    # Select only the 'Q Name' column
    questionnaire_df = questionnaire_df[['Q Name']]
    # Drop any NaN values and convert the DataFrame to a list of lists
    list_questions = questionnaire_df.dropna().values.tolist()
    # Flatten the list and remove any leading or trailing whitespaces
    list_questions = [item for sublist in list_questions for item in sublist]
    list_questions = [s.strip() for s in list_questions]
    # Count the number of questions
    n_of_questions = len(list_questions)
    return list_questions, n_of_questions

def count_number_of_questions_sqname(questionnaire_file):
    # Read the Excel file and access the 'survey' sheet, skipping the first 2 rows
    questionnaire_df = pd.read_excel(pd.ExcelFile(questionnaire_file), sheet_name='survey',skiprows=2)
    # Filter the dataframe to only include the 'Suggested Qname' column
    questionnaire_df = questionnaire_df[['Suggested Qname']]
    # Remove rows with NaN values and convert the DataFrame into a list
    list_questions = questionnaire_df.dropna().values.tolist()
    # Flatten the list
    list_questions = [item for sublist in list_questions for item in sublist]
    # Remove any leading/trailing whitespace from the strings in the list
    list_questions = [s.strip() for s in list_questions]
    # Count the number of questions
    n_of_questions = len(list_questions)
    # Return the cleaned list of questions and the count of questions
    return list_questions, n_of_questions


def highlight_differences_in_qname(questionnaire_file, previousround_questionnaire_file):
    #green file
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    # Load current and previous data
    curr_df = pd.read_excel(questionnaire_file, sheet_name='survey', skiprows=2)
    prev_df = pd.read_excel(previousround_questionnaire_file, sheet_name='survey', skiprows=2)

    # Create lookup by 'Q Name'
    prev_lookup = prev_df.set_index('Q Name')[['Suggested Qname', 'English']].dropna(how='all')

    # Load workbook and worksheet
    wb = load_workbook(questionnaire_file)
    ws = wb['survey']

    # Define fill colors
    green_fill = PatternFill(start_color="9dff9d", end_color="9dff9d", fill_type="solid")
    yellow_fill = PatternFill(start_color="fff8b3", end_color="fff8b3", fill_type="solid")

    # Iterate through rows (starting from row 4 after skiprows+header)
    for row in range(4, ws.max_row + 1):
        qname = ws[f"B{row}"].value  # 'Q Name'
        suggested = ws[f"C{row}"]
        english = ws[f"D{row}"]

        if qname and qname in prev_lookup.index:
            prev_suggested = prev_lookup.at[qname, 'Suggested Qname']
            prev_english = prev_lookup.at[qname, 'English']

            # Compare Suggested Qname
            if str(suggested.value).strip() == str(prev_suggested).strip():
                suggested.fill = green_fill
            else:
                suggested.fill = yellow_fill

            # Compare English label
            if str(english.value).strip() == str(prev_english).strip():
                english.fill = green_fill
            else:
                english.fill = yellow_fill

    wb.save(questionnaire_file)
    print("✅ Highlighted English column, Suggested Qname, and Q Name (by lookup): Green = same, Yellow = changed.")

    #OLD PART OF THE SCRIPT
    # Read 'Q Name' and 'Suggested Qname' columns from both files into lists
    #questionnaire_df = pd.read_excel(pd.ExcelFile(questionnaire_file), sheet_name='survey', skiprows=2)
    #previousround_df = pd.read_excel(pd.ExcelFile(previousround_questionnaire_file), sheet_name='survey', skiprows=2)

    # Extract the 'Q Name' and 'Suggested Qname' columns into lists for the current questionnaire
    #questionnaire_qnames = questionnaire_df['Q Name'].dropna().tolist()
    #questionnaire_suggested_qnames = questionnaire_df['Suggested Qname'].dropna().tolist()
    
    # Extract the 'Q Name' and 'Suggested Qname' columns into lists for the previous round questionnaire
    #previousround_qnames = previousround_df['Q Name'].dropna().tolist()
    #previousround_suggested_qnames = previousround_df['Suggested Qname'].dropna().tolist()

    # Create lookup dictionaries using 'Q Name' as key, if 'name' column is missing
    #prev_qname_dict = dict(zip(previousround_df['Q Name'], previousround_df['Q Name']))
    #prev_sugg_dict = dict(zip(previousround_df['Q Name'], previousround_df['Suggested Qname']))
    
    # Open the questionnaire file with openpyxl
    #wb = load_workbook(filename=questionnaire_file)
    #ws = wb['survey']

    # Define a green color fill to highlight cells
    #green_fill = PatternFill(start_color="9dff9d", end_color="9dff9d", fill_type="solid")

    # Highlight cells that are in both questionnaire_qnames and previousround_qnames for 'Q Name' column
    #for row in ws.iter_rows(min_row=4, min_col=2, max_col=2):  # updated column number for 'Q Name'
    #    for cell in row:
    #        if cell.value in previousround_qnames:
    #            cell.fill = green_fill

    # Highlight cells that are in both questionnaire_suggested_qnames and previousround_suggested_qnames for 'Suggested Qname' column
    #for row in ws.iter_rows(min_row=4, min_col=3, max_col=3):  # updated column number for 'Suggested Qname'
    #    for cell in row:
    #        if cell.value in previousround_suggested_qnames:
    #            cell.fill = green_fill

    

from openpyxl.styles import Font

def create_question_changes_sheet(wb, unique_country_questionnaire_qname, unique_template_questionnaire_qname, unique_country_questionnaire, unique_template_questionnaire):
    # Create a new sheet called "Question changes"
    new_sheet = wb.create_sheet("Question changes")

    # Modify the values for A1 and apply bold and italic formatting
    new_sheet["A1"] = "***In the case of a Previous Questionnaire validation, the 'Template Questionnaire' refers to the Previous round questionnaire***"
    font = Font(bold=True, italic=True)
    new_sheet["A1"].font = font

    # Paste the outputs of Message 7 and 8 into separate columns in "Question changes"
    new_sheet["A3"] = "QName - Country questionnaire:"
    new_sheet["B3"] = "QName - Template questionnaire only:"

    # Apply bold formatting to cells A3 and B3
    bold_font = Font(bold=True)
    new_sheet["A3"].font = bold_font
    new_sheet["B3"].font = bold_font
    new_sheet["D3"].font = bold_font
    new_sheet["E3"].font = bold_font

    # Paste the values vertically in column A (QName Field) starting from row 4
    for i, question in enumerate(unique_country_questionnaire_qname, start=4):
        new_sheet.cell(row=i, column=1).value = question

    # Paste the values vertically in column B (QName Field) starting from row 4
    for i, question in enumerate(unique_template_questionnaire_qname, start=4):
        new_sheet.cell(row=i, column=2).value = question

    # Paste the outputs of Message 5 and 6 into separate columns in "Question changes"
    new_sheet["D3"] = "SuggestedQName - Country questionnaire:"
    new_sheet["E3"] = "SuggestedQName - Template questionnaire:"

    # Paste the values vertically in column D (SuggestedQName Field) starting from row 4
    for i, question in enumerate(unique_country_questionnaire, start=4):
        new_sheet.cell(row=i, column=4).value = question

    # Paste the values vertically in column E (SuggestedQName Field) starting from row 4
    for i, question in enumerate(unique_template_questionnaire, start=4):
        new_sheet.cell(row=i, column=5).value = question

    # Auto format column width for cells A1, A3, B3, D3, and E3
    columns_to_format = ['A', 'B', 'D', 'E']

    # Determine the maximum length for each column
    max_lengths = [max(len(str(new_sheet[column + '3'].value)), len(column)) for column in columns_to_format]

    # Auto format column width for cells A1, A3, B3, D3, and E3
    for column in columns_to_format:
        adjusted_width = (max_lengths[columns_to_format.index(column)] + 2) * 1.2
        new_sheet.column_dimensions[column].width = adjusted_width

    return wb


def read_questionnaire(input_questionnaire_file, output_filename):
    ##this section of the script reads the survey excel file and creates an excel file with multiple sheets:
    ### each sheet contains the coded value and description for a "Single choice" or "Open Ended-Select All That Apply" question.
    ###moreover, it creates and populates several lists that will be used later for defining each field of the final table names, types and domains

    #print("Opening questionnaire DF")
    quest_df = pd.read_excel(open(input_questionnaire_file, 'rb'), sheet_name='survey',skiprows=2)

# Check if the survey includes other languages besides English and print an appropriate message
    if 'French' in quest_df:
        print("Version in French and English")
        languages = ["French","English"]
    elif 'Arabic' in quest_df:
        print("Version in Arabic and English")
        languages = ["Arabic","English"]
    elif 'Spanish' in quest_df:
        print("Version in Spanish and English")
        languages = ["Spanish","English"]
    elif 'Portuguese' in quest_df:
        print("Version in Portuguese and English")
        languages = ["Portuguese","English"]
        
    else:
        print("Version in English only")
        languages = ["English"]
        
# Iterate through languages to create Excel files with appropriate naming convention        
    for language in languages:
        filename = output_filename[:-5] + "_" +  language.lower()[:2] + ".xlsx"
        writer = pd.ExcelWriter(filename, engine='xlsxwriter')
        #create a list of all possible numbering
        numbering = ["%s)" % n for n in range(1,200)] ## 1), 2), ... 200)
        
        # Initialize containers to store the results
        dict_derived_fieldnames = {} #this dict will group all derived fields in case of "Select All That Apply" type of questions
        field_names_list = [] ##this list will contain all fields of the final table
        text_type_fields = [] ##this list will contain all fields of the final table with TEXT type
        range_type_fields = [] ##this list will contain all fields of the final table storing RANGE data  (will be LONG type)
        double_type_fields = [] ##this list will contain all fields of the final table storing DOUBLE data
        ##iterate the following for each row (so each question of the questionnaire)
        all_derived_fieldnames = []
        all_answers_with_other_option = [] ##this list will contain all "Other: specify" fields
        
         # Filters the dataframe to rows where the 'Suggested Qname' is not null
        quest_df = quest_df[quest_df['Suggested Qname'].notna()]

        for index, row in quest_df.iterrows():
            question_name = []
            question_type = []
            try:
                first_derived_fieldname = "" #the name of the first derived field will be the main of the domain table
                derived_fieldnames = []
                codes_and_labels = []
                categories = str(row[language]).replace("\t","")
                question_name = row['Suggested Qname'].strip()  #Q Name
                question_type = row['Q Type']
                programming_instructions = row['Programming Instructions'] #this field contains coded values for crop_main
                #print("\n\n----%s----" % question_name)
                #only for questions with pre-defined categories need domains
                if question_type in ("StartRecording","Single Choice","Open Ended-Single Choice", "Open Ended - Single Choice", "Open Ended-Select All That Apply",
                                     "Select All That Apply","Open Ended - Select All That Apply "):
                    if question_name == 'crp_main': #for this question only, coded values should be taken from field programming_instructions
                        programming_lines = programming_instructions.splitlines()
                        for programming_line in programming_lines:
                            if ")" in programming_line:
                                #print(programming_line)
                                index, category = programming_line.split(")")
                                category = fix_category_formatting(category)
                                codes_and_labels.append([index, category])
                    else:
                        #find all numbering present in the category string
                        numbering_in_text = [n for n in numbering if n in categories]
                        #print(numbering_in_text)
                        ##the following loop creates a list "codes_and_labels" with all available codes&labels for each question
                        for index in range(0,len(numbering_in_text)):
                            start = categories.find(numbering_in_text[index]) + len(numbering_in_text[index])
                            try:
                                end = categories.find(numbering_in_text[index + 1])
                                substring = categories[start:end].strip()
                            except:
                                # it fails during the last loop -> the last option is usually at the end of the string
                                substring = categories[start:].strip()
                            #print(substring)
                            category = fix_category_formatting(substring)
                            codes_and_labels.append([index +1, category])

                    if question_type not in ["Open Ended-Select All That Apply","Select All That Apply","Open Ended - Select All That Apply "]:
                        #so questions with NO derived fields
                        field_names_list.append(question_name.strip())
                        codes_and_labels_df = pd.DataFrame(codes_and_labels, columns=['code', 'label'])
                        codes_and_labels_df.to_excel(writer, sheet_name=question_name)
                    else:
                        #so questions with derived fields
                        numbering_in_qname = [n for n in numbering if n in question_name]
                        for index in range(0, len(numbering_in_qname)):
                            start = question_name.find(numbering_in_qname[index]) + len(numbering_in_qname[index])
                            try:
                                end = question_name.find(numbering_in_qname[index + 1])
                                derived_field_name = question_name[start:end].strip()
                            except:
                                # it fails during the last loop -> the last option is usually at the end of the string
                                derived_field_name = question_name[start:].strip()
                            all_derived_fieldnames.append(derived_field_name)
                            derived_fieldnames.append(derived_field_name)
                            field_names_list.append(derived_field_name)
                            if derived_field_name[-6:] == "_other": ##this field will need to be STRING - with no domain (since it's a 'other specify')
                                all_answers_with_other_option.append(derived_field_name)
                            if index == 0:
                                first_derived_fieldname = derived_field_name
                                codes_and_labels_df = pd.DataFrame(codes_and_labels, columns=['code', 'label'])
                                #codes_and_labels_df.to_excel(writer, sheet_name=derived_field_name) #we don't need domain table for derived fields, since they will use YES NOT domain table
                        dict_derived_fieldnames[first_derived_fieldname] = all_derived_fieldnames
                elif question_type == "Range":
                    #these questions will be associated to LONG type fields
                    field_names_list.append(question_name.strip())
                    range_type_fields.append(question_name)
                elif question_type == "Open Ended":
                    #these questions will be associated to TEXT type fields
                    if not pd.isnull(question_name): #NaN rows we want to skip (i.e. OptIn question without a name in the survey)
                        field_names_list.append(question_name)
                        text_type_fields.append(question_name)
                else:
                    #print("QUESTION SKIPPED ---------", question_name, question_type)
                    pass
            except Exception as e:
                print("Failed question %s (type: %s) for error: %s " % (question_name, question_type,e))

        #adding Yes No table (for derived fields domain)
        d = {1:"Yes",0:"No"}
        yesno_df = pd.DataFrame(d.items(), columns=['code', 'label'])
        yesno_df.to_excel(writer, sheet_name='yes_no')
        #print("Saving codes and labels %s" % coded_values_file)

        #print(all_answers_with_other_option)

        #creating additional sheet with all derived fields, that will be used for a script that ensures that also these values are within the domains in the output table, in a later stage
        list_of_yes_no_fields = []
        for derived_field in all_derived_fieldnames:
            if derived_field not in all_answers_with_other_option:
                list_of_yes_no_fields.append(derived_field)

        derived_fields_df = pd.DataFrame(list_of_yes_no_fields)
        derived_fields_df.to_excel(writer, sheet_name='derived_fields')
        # Close the Pandas Excel writer and output the Excel file.
        #writer.save()
    return languages

def detect_enumerator(geopoll_or_kobo_template):
    enumerator = ""
    if "kobo" in geopoll_or_kobo_template.lower():
        country_df.columns = country_df.columns.str.replace("[/]", "")
        enumerator = "kobo"
    elif "geopoll" in geopoll_or_kobo_template.lower():
        enumerator = "geopoll"
    else:
        print("Please include Geopoll or Kobo inside the filename: %s " % geopoll_or_kobo_template)
        enumerator = "Please include Geopoll or Kobo inside the filename"
    return enumerator

# Function to insert a sheet with administrative level 2 reference data to the existing questionnaire Excel file
def insert_sheet_with_adm2_reference_old(questionnaire_file,adm0_iso3):
    # Connect to the ArcGIS online service
    gis = GIS("https://hqfao.maps.arcgis.com")
    # Retrieve the layer using its item ID
    item = gis.content.get('3596c3ad318849068eda21517ade30be')
    flayer = item.layers[0]
    # Query the layer to obtain the administrative level 2 data for the given ISO3 country code
    query = "adm0_ISO3 = '" + adm0_iso3 +"'"
    sdf = flayer.query(where=query).sdf
    
    # Remove unnecessary columns
    del sdf['OBJECTID']
    #del sdf['validity']
    del sdf['Shape__Area']
    del sdf['Shape__Length']
    del sdf['SHAPE']
    #print(sdf.head())
    excel_book = openpyxl.load_workbook(questionnaire_file)
    with pd.ExcelWriter(questionnaire_file, engine='openpyxl') as writer:
        writer.book = excel_book
        sdf.to_excel(writer, 'ADMIN info', index=False)
        writer.save()

    return sdf


def insert_sheet_with_adm_reference(questionnaire_file,admin_level,adm0_iso3):
    if admin_level == "Admin 0":
        admin_url = "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/Administrative_Boundaries_Reference_(view_layer)/FeatureServer/2/query?where=adm0_ISO3%20%3D%20'" + adm0_iso3 + "'&outFields=adm0_name,adm0_name_local,adm0_ISO3_2d,adm0_ISO3&returnGeometry=false&outSR=4326&f=json"
    elif admin_level == "Admin 1":
        admin_url = "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/Administrative_Boundaries_Reference_(view_layer)/FeatureServer/1/query?where=adm0_ISO3%20%3D%20'" + adm0_iso3 + "'&outFields=adm1_name,adm1_name_local,adm1_pcode,adm0_name,adm0_ISO3&returnGeometry=false&outSR=4326&f=json"
    elif admin_level == "Admin 2":
        admin_url = "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/Administrative_Boundaries_Reference_(view_layer)/FeatureServer/0/query?where=adm0_ISO3%20%3D%20'" + adm0_iso3 + "'&outFields=adm2_name,adm2_name_local,adm2_pcode,adm1_name,adm1_pcode,adm0_name,adm0_ISO3&returnGeometry=false&outSR=4326&f=json"
    elif admin_level == "Admin 3":
        admin_url ="https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/Reference_Admin_3/FeatureServer/0/query?where=adm0_ISO3%20%3D%20'"+ adm0_iso3 + "'&outFields=adm3_name,adm3_name_local,adm3_pcode,adm2_name,adm2_pcode,adm1_name,adm1_pcode,adm0_name,adm0_ISO3&returnGeometry=false&outSR=4326&f=json"

    with urllib.request.urlopen(admin_url) as admin_url:
        data = json.loads(admin_url.read().decode())
        #print(type(data))

    df = pd.json_normalize(data['features'])
    #print (df.head())

    if admin_level == "Admin 0":
        df.rename(columns={'attributes.adm0_name': 'adm0_name',
                           'attributes.adm0_name_local': 'adm0_name_local',
                           'attributes.adm0_ISO3_2d': 'adm0_ISO3_2d',
                           'attributes.adm0_ISO3': 'adm0_ISO3'}, inplace=True)
    elif admin_level == "Admin 1":
        df.rename(columns={'attributes.adm0_name': 'adm0_name',
                           'attributes.adm0_ISO3': 'adm0_ISO3',
                           'attributes.adm1_name': 'adm1_name',
                           'attributes.adm1_name_local': 'adm1_name_local',
                           'attributes.adm1_pcode': 'adm1_pcode'}, inplace=True)

    elif admin_level == "Admin 2":
        df.rename(columns={'attributes.adm0_name': 'adm0_name',
                           'attributes.adm0_ISO3': 'adm0_ISO3',
                           'attributes.adm1_name': 'adm1_name',
                           'attributes.adm1_pcode': 'adm1_pcode',
                           'attributes.adm2_name': 'adm2_name',
                           'attributes.adm2_name_local': 'adm2_name_local',
                           'attributes.adm2_pcode': 'adm2_pcode'}, inplace=True)

    elif admin_level == "Admin 3":
        df.rename(columns={'attributes.adm0_name': 'adm0_name',
                           'attributes.adm0_ISO3': 'adm0_ISO3',
                           'attributes.adm1_name': 'adm1_name',
                           'attributes.adm1_pcode': 'adm1_pcode',
                           'attributes.adm2_name': 'adm2_name',
                           'attributes.adm2_pcode': 'adm2_pcode',
                           'attributes.adm3_name': 'adm3_name',
                           'attributes.adm3_name_local': 'adm3_name_local',
                           'attributes.adm3_pcode': 'adm3_pcode'}, inplace=True)


    excel_book = openpyxl.load_workbook(questionnaire_file)
    with pd.ExcelWriter(questionnaire_file, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name=admin_level + ' info', index=False)
        #writer.save()
    return df


from openpyxl.utils import get_column_letter

def update_questionnaire(template_file, country_file):
     # Define the list of placeholder keys that are expected in the template and need to be replaced.
    keys = ['$ADMIN1$', '$ADMIN2$', '$reference year$', '$season$', '$season phase $',
            '$local measurement unit$', '$currency$', '$MIN AMOUNT$', '$MAX AMOUNT$',
            '$THRESHOLD$', '$local vegetables$', '$local fruits$', '$expected or nothing$']
    
    # Load the questionnaire template from the specified file, skipping the first 3 rows.
    template_data = pd.read_excel(template_file,'survey', header=2)
    # Extract only the 'Q Name' and 'English' columns and drop any rows with missing values.
    selected_data = template_data[['Q Name', 'English']]
    selected_data=selected_data.dropna()
    
    # Filter rows where the 'English' column contains any of the specified keys.
    mask = selected_data['English'].apply(lambda x: any(key in str(x) for key in keys))
    selected_data = selected_data[mask]
    #print(selected_data["Q Name"])
    
    # Load the country questionnaire with openpyxl
    workbook = openpyxl.load_workbook(country_file)
    workbook.sheetnames
    worksheet = workbook["survey"]
    
    # Determine the number of rows in the worksheet.
    number_of_rows = worksheet.max_row 
    
    # Iterate over each row in the worksheet.
    for k in range(number_of_rows):
        # Extract the question value from the second column of the current row.
        questionValue = str(worksheet[get_column_letter(2)+str(k+1)].value)
        
        # Compare the question value with the 'Q Name' column of the selected_data DataFrame.
        for index, row in selected_data.iterrows():
            if questionValue == row['Q Name']:

                # If there's a match, update the value in the fourth column with the 'English' value from the template.
                worksheet[get_column_letter(4)+str(k+1)] = str(row['English'])
    # Save the modified country-specific questionnaire back to the file
    workbook.save(country_file)




# Function to find and replace specific strings in a given questionnaire Excel file based on a language-specific table
def find_and_replace_strings_in_df(questionnaire_file):
    # Detect the language of the questionnaire using the 'detect_language' function
    language = detect_language(questionnaire_file)
    # Load the 'Additional information' sheet from the questionnaire Excel file into a DataFrame
    replacing_table_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Additional information', skiprows=1)
    #print(replacing_table_df)

    # Open the workbook and select the 'survey' worksheet
    workbook = openpyxl.load_workbook(questionnaire_file)
    workbook.sheetnames
    worksheet = workbook["survey"]
    # Get the number of rows and columns in the worksheet
    number_of_rows = worksheet.max_row #Number of Rows
    number_of_columns = worksheet.max_column #Number of Columns

    # Handle English language replacements
    if language == "en":
        # Extract 'Original' and 'Replacement' columns from the DataFrame
        replacing_table_df = replacing_table_df[['Original','Replacement']]
        # Create a list of original strings to be replaced
        replacement_list = list(replacing_table_df['Original'].values.tolist())
        #print(replacement_list)

        # Determine the 'expected_or_nothing' value based on the 'Replacement' value at index 4
        expected_or_nothing = ""
        if replacing_table_df.iloc[4]['Replacement'] == "Not yet in season" or replacing_table_df.iloc[4]['Replacement'] == "Land preparation" or replacing_table_df.iloc[4]['Replacement'] == "Planting" or replacing_table_df.iloc[4]['Replacement'] == "Early growing" or replacing_table_df.iloc[4]['Replacement'] == "Growing" or replacing_table_df.iloc[4]['Replacement'] == "Maturing":
            expected_or_nothing = "expected"
        elif replacing_table_df.iloc[4]['Replacement'] == "Harvesting" or replacing_table_df.iloc[4]['Replacement'] == "Recently finished":
            expected_or_nothing = ""
        else:
            expected_or_nothing = ""

        replacementTextKeyPairs = {'$ADMIN1$': replacing_table_df.iloc[0]['Replacement'],
                                   '$ADMIN2$': replacing_table_df.iloc[1]['Replacement'],
                                   '$reference year$': replacing_table_df.iloc[2]['Replacement'],
                                   '$season$': replacing_table_df.iloc[3]['Replacement'],
                                   '$season phase $': replacing_table_df.iloc[4]['Replacement'],
                                   '$local measurement unit$': replacing_table_df.iloc[5]['Replacement'],
                                   '$currency$': replacing_table_df.iloc[7]['Replacement'],
                                   '$MIN AMOUNT$': replacing_table_df.iloc[8]['Replacement'],
                                   '$MAX AMOUNT$': replacing_table_df.iloc[9]['Replacement'],
                                   '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement'],
                                   '$local vegetables$': replacing_table_df.iloc[11]['Replacement'],
                                   '$local fruits$': replacing_table_df.iloc[12]['Replacement'],
                                   '$expected or nothing$': expected_or_nothing
                              }

        

    elif language == "fr":
        
        print("FRENCH")

        replacing_table_df = replacing_table_df[['Original','Replacement (EN)','Replacement (FR)']]
        replacement_list = list(replacing_table_df['Original'].values.tolist())
        
        #print(replacing_table_df)
        #print(replacement_list)
        
        
        ##Translating values
        
        #Translating Reference Year
        reference_year_EN = ""
        if replacing_table_df.iloc[2]['Replacement (FR)'] == "l'année dernière":
            reference_year_EN = "last year"
        elif replacing_table_df.iloc[2]['Replacement (FR)'] == "une année normale":
            reference_year_EN = "a normal year"
        else:
            reference_year_EN = replacing_table_df.iloc[2]['Replacement (FR)']
        

        #Translating season phase
        season_phase_EN = ""
        if replacing_table_df.iloc[4]['Replacement (FR)'] == "Pas encore en saison":
            season_phase_EN = "Not yet in season"
        elif replacing_table_df.iloc[4]['Replacement (FR)'] == "Préparation du terrain":
            season_phase_EN = "Land preparation"
        elif replacing_table_df.iloc[4]['Replacement (FR)'] == "Semis / plantation":
            season_phase_EN = "Planting"            
        elif replacing_table_df.iloc[4]['Replacement (FR)'] == "Début de la croissance":
            season_phase_EN = "Early growing"
        elif replacing_table_df.iloc[4]['Replacement (FR)'] == "Croissance":
            season_phase_EN = "Growing"
        elif replacing_table_df.iloc[4]['Replacement (FR)'] == "Proche de la récolte":
            season_phase_EN = "Maturing"
        elif replacing_table_df.iloc[4]['Replacement (FR)'] == "Récolte":
            season_phase_EN = "Harvesting"
        elif replacing_table_df.iloc[4]['Replacement (FR)'] == "Post-récolte":
            season_phase_EN = "Recently finished"
        else:
            season_phase_EN = replacing_table_df.iloc[4]['Replacement (FR)']
            
            
        expected_or_nothing = ""
        attendue_ou_rien = ""
        if replacing_table_df.iloc[4]['Replacement (EN)'] == "Not yet in season" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Land preparation" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Planting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Early growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Maturing":
            expected_or_nothing = "expected"
            attendue_ou_rien = "attendue"
        elif replacing_table_df.iloc[4]['Replacement (EN)'] == "Harvesting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Recently finished":
            expected_or_nothing = ""
            attendue_ou_rien = ""
        else:
            expected_or_nothing = ""
            attendue_ou_rien = ""


        replacementTextKeyPairs = {'$ADMIN1$': replacing_table_df.iloc[0]['Replacement (FR)'],
                                   '$ADMIN2$': replacing_table_df.iloc[1]['Replacement (FR)'],
                                   '$reference year$': reference_year_EN,
                                   '$reference year FR$': replacing_table_df.iloc[2]['Replacement (FR)'],
                                   '$season$': replacing_table_df.iloc[3]['Replacement (EN)'],
                                   '$saison$': replacing_table_df.iloc[3]['Replacement (FR)'],
                                   '$season phase $': replacing_table_df.iloc[4]['Replacement (FR)'],
                                   '$local measurement unit$': replacing_table_df.iloc[5]['Replacement (EN)'],
                                   '$unité de mesure locale$': replacing_table_df.iloc[5]['Replacement (FR)'],
                                   '$currency$': replacing_table_df.iloc[7]['Replacement (FR)'],
                                   '$MIN AMOUNT$': replacing_table_df.iloc[8]['Replacement (FR)'],
                                   '$MAX AMOUNT$': replacing_table_df.iloc[9]['Replacement (FR)'],
                                   '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement (FR)'],
                                   '$local vegetables$': replacing_table_df.iloc[11]['Replacement (FR)'],
                                   '$local fruits$': replacing_table_df.iloc[12]['Replacement (FR)'],
                                   '$expected or nothing$': expected_or_nothing,
                                   '$attendue ou rien$': attendue_ou_rien
                                    }
        


    elif language == "ar":

        expected_or_nothing = ""
        if replacing_table_df.iloc[4]['Replacement (EN)'] == "Not yet in season" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Land preparation" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Planting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Early growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Maturing":
            expected_or_nothing = "expected"
        elif replacing_table_df.iloc[4]['Replacement (EN)'] == "Harvesting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Recently finished":
            expected_or_nothing = ""
        else:
            expected_or_nothing = ""

        attendue_ou_rien = ""
        if replacing_table_df.iloc[4]['Replacement (AR)'] == "ليس في الموسم بعد" or replacing_table_df.iloc[4]['Replacement (AR)'] == "إعداد الأرض" or replacing_table_df.iloc[4]['Replacement (AR)'] == "الزرع" or replacing_table_df.iloc[4]['Replacement (AR)'] == "النمو المبكر" or replacing_table_df.iloc[4]['Replacement (AR)'] == "النمو" or replacing_table_df.iloc[4]['Replacement (AR)'] == "النضوج":
            attendue_ou_rien = "مُتوقع"
        elif replacing_table_df.iloc[3]['Replacement (AR)'] == "الحصاد" or replacing_table_df.iloc[3]['Replacement (AR)'] == "انتهى مؤخرا":
            attendue_ou_rien = ""
        else:
            attendue_ou_rien = ""

        replacementTextKeyPairs = {'$ADMIN1 AR$': replacing_table_df.iloc[0]['Replacement (AR)'],
                                   '$ADMIN2 AR$': replacing_table_df.iloc[1]['Replacement (AR)'],
                                   '$reference year AR$': replacing_table_df.iloc[2]['Replacement (AR)'],
                                   '$season AR$': replacing_table_df.iloc[3]['Replacement (AR)'],
                                   '$season phase AR$': replacing_table_df.iloc[4]['Replacement (AR)'],
                                   '$local measurement unit AR$': replacing_table_df.iloc[5]['Replacement (AR)'],
                                   '$currency AR$': replacing_table_df.iloc[7]['Replacement (AR)'],
                                   '$MIN AMOUNT AR$': replacing_table_df.iloc[8]['Replacement (AR)'],
                                   '$MAX AMOUNT AR$': replacing_table_df.iloc[9]['Replacement (AR)'],
                                   '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement (AR)'],
                                   '$local vegetables AR$': replacing_table_df.iloc[11]['Replacement (AR)'],
                                   '$local fruits AR$': replacing_table_df.iloc[12]['Replacement (AR)'],
                                   '$expected or nothing AR$': expected_or_nothing,

                                   '$ADMIN1$': replacing_table_df.iloc[0]['Replacement (EN)'],
                                   '$ADMIN2$': replacing_table_df.iloc[1]['Replacement (EN)'],
                                   '$reference year$': replacing_table_df.iloc[2]['Replacement (EN)'],
                                   '$season$': replacing_table_df.iloc[3]['Replacement (EN)'],
                                   '$season phase$': replacing_table_df.iloc[4]['Replacement (EN)'],
                                   '$local measurement unit$': replacing_table_df.iloc[5]['Replacement (EN)'],
                                   '$currency$': replacing_table_df.iloc[7]['Replacement (EN)'],
                                   '$MIN AMOUNT$': replacing_table_df.iloc[8]['Replacement (EN)'],
                                   '$MAX AMOUNT$': replacing_table_df.iloc[9]['Replacement (EN)'],
                                   '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement (EN)'],
                                   '$local vegetables$': replacing_table_df.iloc[11]['Replacement (EN)'],
                                   '$local fruits$': replacing_table_df.iloc[12]['Replacement (EN)'],
                                   '$expected or nothing$': expected_or_nothing
                              }

    elif language == "pt":

        expected_or_nothing = ""
        if replacing_table_df.iloc[4]['Replacement (EN)'] == "Not yet in season" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Land preparation" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Planting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Early growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Maturing":
            expected_or_nothing = "expected"
            attendue_ou_rien = "مُتوقع"
        elif replacing_table_df.iloc[4]['Replacement (EN)'] == "Harvesting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Recently finished":
            expected_or_nothing = ""
            attendue_ou_rien = ""
        else:
            expected_or_nothing = ""
            attendue_ou_rien = ""

        replacementTextKeyPairs = {'$ADMIN1 PT$': replacing_table_df.iloc[0]['Replacement (PT)'],
                                   '$ADMIN2 PT$': replacing_table_df.iloc[1]['Replacement (PT)'],
                                   '$reference year PT$': replacing_table_df.iloc[2]['Replacement (PT)'],
                                   '$season PT$': replacing_table_df.iloc[3]['Replacement (PT)'],
                                   '$season phase PT$': replacing_table_df.iloc[4]['Replacement (PT)'],
                                   '$local measurement unit PT$': replacing_table_df.iloc[5]['Replacement (PT)'],
                                   '$currency PT$': replacing_table_df.iloc[7]['Replacement (PT)'],
                                   '$MIN AMOUNT PT$': replacing_table_df.iloc[8]['Replacement (PT)'],
                                   '$MAX AMOUNT PT$': replacing_table_df.iloc[9]['Replacement (PT)'],
                                   '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement (PT)'],
                                   '$local vegetables PT$': replacing_table_df.iloc[11]['Replacement (PT)'],
                                   '$local fruits PT$': replacing_table_df.iloc[12]['Replacement (PT)'],
                                   '$expected or nothing PT$': expected_or_nothing,

                                   '$ADMIN1$': replacing_table_df.iloc[0]['Replacement (EN)'],
                                   '$ADMIN2$': replacing_table_df.iloc[1]['Replacement (EN)'],
                                   '$reference year$': replacing_table_df.iloc[2]['Replacement (EN)'],
                                   '$season$': replacing_table_df.iloc[3]['Replacement (EN)'],
                                   '$season phase$': replacing_table_df.iloc[4]['Replacement (EN)'],
                                   '$local measurement unit$': replacing_table_df.iloc[5]['Replacement (EN)'],
                                   '$currency$': replacing_table_df.iloc[7]['Replacement (EN)'],
                                   '$MIN AMOUNT$': replacing_table_df.iloc[8]['Replacement (EN)'],
                                   '$MAX AMOUNT$': replacing_table_df.iloc[9]['Replacement (EN)'],
                                   '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement (EN)'],
                                   '$local vegetables$': replacing_table_df.iloc[11]['Replacement (EN)'],
                                   '$local fruits$': replacing_table_df.iloc[12]['Replacement (EN)'],
                                   '$expected or nothing$': expected_or_nothing
                              }            
    elif language == "es":
        
        expected_or_nothing = ""
        if replacing_table_df.iloc[4]['Replacement (EN)'] == "Not yet in season" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Land preparation" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Planting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Early growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Maturing":
            expected_or_nothing = "expected"
            attendue_ou_rien = "esperado"
        
        elif replacing_table_df.iloc[4]['Replacement (EN)'] == "Harvesting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Recently finished":
            expected_or_nothing = ""
            attendue_ou_rien = ""
        
        else:
            expected_or_nothing = ""
            attendue_ou_rien = ""

            replacementTextKeyPairs = {'$ADMIN1 ES$': replacing_table_df.iloc[0]['Replacement (ES)'],
                                       '$ADMIN2 ES$': replacing_table_df.iloc[1]['Replacement (ES)'],
                                       '$reference year ES$': replacing_table_df.iloc[2]['Replacement (ES)'],
                                       '$season ES$': replacing_table_df.iloc[3]['Replacement (ES)'],
                                       '$season phase ES$': replacing_table_df.iloc[4]['Replacement (ES)'],
                                       '$local measurement unit ES$': replacing_table_df.iloc[5]['Replacement (ES)'],
                                       '$currency ES$': replacing_table_df.iloc[7]['Replacement (ES)'],
                                       '$MIN AMOUNT ES$': replacing_table_df.iloc[8]['Replacement (ES)'],
                                       '$MAX AMOUNT ES$': replacing_table_df.iloc[9]['Replacement (ES)'],
                                       '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement (ES)'],
                                       '$local vegetables ES$': replacing_table_df.iloc[11]['Replacement (ES)'],
                                       '$local fruits ES$': replacing_table_df.iloc[12]['Replacement (ES)'],
                                       '$expected or nothing ES$': attendue_ou_rien,
            
                                       '$ADMIN1$': replacing_table_df.iloc[0]['Replacement (EN)'],
                                       '$ADMIN2$': replacing_table_df.iloc[1]['Replacement (EN)'],
                                       '$reference year$': replacing_table_df.iloc[2]['Replacement (EN)'],
                                       '$season$': replacing_table_df.iloc[3]['Replacement (EN)'],
                                       '$season phase$': replacing_table_df.iloc[4]['Replacement (EN)'],
                                       '$local measurement unit$': replacing_table_df.iloc[5]['Replacement (EN)'],
                                       '$currency$': replacing_table_df.iloc[7]['Replacement (EN)'],
                                       '$MIN AMOUNT$': replacing_table_df.iloc[8]['Replacement (EN)'],
                                       '$MAX AMOUNT$': replacing_table_df.iloc[9]['Replacement (EN)'],
                                       '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement (EN)'],
                                       '$local vegetables$': replacing_table_df.iloc[11]['Replacement (EN)'],
                                       '$local fruits$': replacing_table_df.iloc[12]['Replacement (EN)'],
                                       '$expected or nothing$': expected_or_nothing
                                    }   

    # Iterate over each column and row in the worksheet to find and replace text based on the dictionary
    for i in range(number_of_columns):
        for k in range(number_of_rows):
            # Get the value of the cell at (i+1, k+1)
            cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
            #print(cellValue)

            
            # Iterate through each key in the replacement dictionary
            for key in replacementTextKeyPairs.keys():
                # If the key is found in the cell value and the cell value is not None
                if key in str(cellValue) and str(cellValue) != None:
                    # Replace the key with its corresponding value in the dictionary
                    newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                    #newCellValue = replacementTextKeyPairs.get(key)
                    # Update the cell with the new value
                    worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                    # Update cellValue for any subsequent replacements within the same cell
                    cellValue = newCellValue
    # Hide the "Additional information" sheet
    workbook["Additional information"].sheet_state = 'hidden'
    # Save the changes to the workbook
    workbook.save(questionnaire_file)

    

def sort_crop_list_by_selection(questionnaire_file):
    crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)

    columns = list(crop_list_df.columns.values.tolist())
    #print(columns)

    if 'Label (FR)' not in columns and 'Label (AR)' not in columns and 'Label (PT)' not in columns and 'Label (ES)' not in columns:

        crop_list_df = crop_list_df[['Select top 10 crops ','GeoPoll code','Dataset code','Label (EN)']]
        q_name = "crp_main"
        q_label = "What has been the main crop that your household has grown for food and income generation in $season$, if any? \n [OPERATOR: SINGLE RESPONSE. \"NO CROP PRODUCTION\" IS IN LAST OPTIONS AT THE END.] \n \n"

        q_name_CROPS_SOLD = "crp_salesmain"
        q_label_CROPS_SOLD = "What has been the main crop that you have sold over the last 3 months?\n[OPERATOR: \"No crop sold in the last 3 months\" APPEARS AT THE END OF THE LIST OF OPTIONS. SELECT ALL CROPS LISTED BY THE RESPONDENT.]\n\n"

        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["survey"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns

        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)

        sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','Dataset code'])
        #print(sorted_crop_list_df)


        sorted_crop_list_df = sorted_crop_list_df.reset_index()
        sorted_crop_list_df['GeoPoll code'] = sorted_crop_list_df.index + 1
        #print(sorted_crop_list_df['GeoPoll code'])
        
        

        sorted_crop_list_df['combined'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (EN)']
        sorted_crop_list = sorted_crop_list_df[['combined']].values.tolist()
        additional_crop_list = [["91)No crop production"], ["92)DON'T KNOW"], ["93)REFUSED"]]
        sorted_crop_list_additional = sorted_crop_list + additional_crop_list
        #print(sorted_crop_list)
        final_sorted_crop_list = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional]))
        
        sorted_crop_list_df['combined_DATASETCODES'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Dataset code'].astype(str)
        sorted_crop_list_DATASETCODES = sorted_crop_list_df[['combined_DATASETCODES']].values.tolist()
        #print(sorted_crop_list_DATASETCODES)
        additional_crop_list_DATASETCODES = [["91)777"], ["92)888"], ["93)999"]]
        sorted_crop_list_additional_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_DATASETCODES
        final_sorted_crop_list_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_DATASETCODES]))
        
        
        
        #print(final_sorted_crop_list)
        #print(final_sorted_crop_list_DATASETCODES)
        #print(crop_list_df)


        #CROPS SOLD
        sorted_crop_list_CROPS_SOLD = list(sorted_crop_list)
        additional_CROPS_SOLD = [["92)DON'T KNOW"], ["93)REFUSED"]]
        sorted_crop_list_CROPS_SOLD.extend(additional_CROPS_SOLD)
        final_sorted_crop_list_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_CROPS_SOLD]))
        #print(sorted_crop_list_CROPS_SOLD)
        #print(final_sorted_crop_list_CROPS_SOLD)
        
        additional_crop_list_CROPS_SOLD_DATASETCODES = [["92)888"], ["93)999"]]
        sorted_crop_list_additional_CROPS_SOLD_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_CROPS_SOLD_DATASETCODES
        final_sorted_crop_list__CROPS_SOLD_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_CROPS_SOLD_DATASETCODES]))
        
        


        # Iterate over the columns and rows, search
        # for the text and replace

        for k in range(number_of_rows):

            questionValue = str(worksheet[get_column_letter(2)+str(k+1)].value)
            labelValue = str(worksheet[get_column_letter(4)+str(k+1)].value)
                #print(cellValue)

            if q_name == str(questionValue) and str(questionValue) != None:
                  newLabelValue = q_label + final_sorted_crop_list
                  #newCellValue = replacementTextKeyPairs.get(key)
                  worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue)

            if q_name_CROPS_SOLD in str(questionValue) and str(questionValue) != None:
                  newLabelValue = q_label_CROPS_SOLD + final_sorted_crop_list_CROPS_SOLD
                  #newCellValue = replacementTextKeyPairs.get(key)
                  worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue)
                    
    
    
    
        replacementTextKeyPairs = {'$CROP CODES$': final_sorted_crop_list_DATASETCODES,
                                  '$CROP SOLD CODES$' : final_sorted_crop_list__CROPS_SOLD_DATASETCODES}
        for i in range(number_of_columns):
            for k in range(number_of_rows):

                cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue

        workbook["Crop list"].sheet_state = 'hidden'            
        workbook.save(questionnaire_file)
        return n_of_choices
    
    
    elif 'Label (FR)' in columns:
        crop_list_df = crop_list_df[['Select top 10 crops ','GeoPoll code','Dataset code','Label (EN)','Label (FR)']]
        q_name = "crp_main"
        q_label_en = "What has been the main crop that your household has grown for food and income generation in $season$, if any? \n [OPERATOR: SINGLE RESPONSE. \"NO CROP PRODUCTION\" IS IN LAST OPTIONS AT THE END.] \n \n"
        q_label_fr = "Quelle a été la principale culture pratiquée par votre ménage pour se nourrir et générer des revenus pendant la $saison$, si applicable? \n [OPÉRATEUR/OPÉRATRICE: RÉPONSE UNIQUE. «AUCUNE PRODUCTION VÉGÉTALE» EST PARMI LES DERNIÈRES OPTIONS, À LA FIN DE LA LISTE.] \n \n"


        q_name_CROPS_SOLD = "crp_salesmain"
        q_label_en_CROPS_SOLD = "What has been the main crop that you have sold over the last 3 months?\n[OPERATOR: \"No crop sold in the last 3 months\" APPEARS AT THE END OF THE LIST OF OPTIONS. SELECT ALL CROPS LISTED BY THE RESPONDENT.]\n\n"
        q_label_fr_CROPS_SOLD = "Quelle a été la principale récolte que vous avez vendue au cours des 3 derniers mois \n[OPERATEUR /OPÉRATRICE: \"Pas de culture vendue sur les 3 derniers mois\" FIGURE DANS LES DERNIÈRES OPTIONS DE RÉPONSE. SÉLECTIONNER TOUTES LES CULTURES LISTÉES PAR LE RÉPONDANT.]\n \n"


        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["survey"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns



        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)



        sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','GeoPoll code'])
        #print(sorted_crop_list_df)
        
        
        ##########NEWWWWWW########
        sorted_crop_list_df = sorted_crop_list_df.reset_index()
        sorted_crop_list_df['GeoPoll code'] = sorted_crop_list_df.index + 1
        #print(sorted_crop_list_df['GeoPoll code'])
        ##########NEWWWWWW########
        
        
        sorted_crop_list_df['combined (EN)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (EN)']
        sorted_crop_list_df['combined (FR)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (FR)']


        sorted_crop_list_en = sorted_crop_list_df[['combined (EN)']].values.tolist()
        additional_crop_list_en = [["91)No crop production"], ["92)DON'T KNOW"], ["93)REFUSED"]]
        sorted_crop_list_additional_en = sorted_crop_list_en + additional_crop_list_en 
        
        
        sorted_crop_list_fr = sorted_crop_list_df[['combined (FR)']].values.tolist()
        additional_crop_list_fr = [["91)Aucune production végétale"], ["92)NE SAIS PAS"], ["93)REFUSÉ"]]
        sorted_crop_list_additional_fr = sorted_crop_list_fr + additional_crop_list_fr
        
        #print(sorted_crop_list)
        final_sorted_crop_list_en = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en]))
        final_sorted_crop_list_fr = '\n'.join(map(str, [i[0] for i in sorted_crop_list_fr]))

        #print(final_sorted_crop_list)
        #print(crop_list_df)
        
        sorted_crop_list_df['combined_DATASETCODES'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Dataset code'].astype(str)
        sorted_crop_list_DATASETCODES = sorted_crop_list_df[['combined_DATASETCODES']].values.tolist()
        #print(sorted_crop_list_DATASETCODES)
        additional_crop_list_DATASETCODES = [["91)777"], ["92)888"], ["93)999"]]
        sorted_crop_list_additional_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_DATASETCODES
        final_sorted_crop_list_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_DATASETCODES]))


        #CROPS SOLD
        sorted_crop_list_en_CROPS_SOLD = list(sorted_crop_list_en)
        additional_en_CROPS_SOLD = [["92)DON'T KNOW"], ["93)REFUSED"]]
        sorted_crop_list_en_CROPS_SOLD.extend(additional_en_CROPS_SOLD)
        final_sorted_crop_list_en_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en_CROPS_SOLD]))
        #print(sorted_crop_list_en_CROPS_SOLD)
        #print(final_sorted_crop_list_en_CROPS_SOLD)

        sorted_crop_list_fr_CROPS_SOLD = list(sorted_crop_list_fr)
        additional_fr_CROPS_SOLD = [["92)NE SAIT PAS"], ["93)REFUSÉ"]]
        sorted_crop_list_fr_CROPS_SOLD.extend(additional_fr_CROPS_SOLD)
        final_sorted_crop_list_fr_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_fr_CROPS_SOLD]))
        #print(sorted_crop_list_fr_CROPS_SOLD)
        #print(final_sorted_crop_list_fr_CROPS_SOLD)

        additional_crop_list_CROPS_SOLD_DATASETCODES = [["91)888"], ["92)999"]]
        sorted_crop_list_additional_CROPS_SOLD_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_CROPS_SOLD_DATASETCODES
        final_sorted_crop_list__CROPS_SOLD_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_CROPS_SOLD_DATASETCODES]))


        # Iterate over the columns and rows, search
        # for the text and replace

        for k in range(number_of_rows):

            questionValue = str(worksheet[get_column_letter(2)+str(k+1)].value)
            labelValue = str(worksheet[get_column_letter(4)+str(k+1)].value)
                #print(cellValue)

            if q_name == str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en + final_sorted_crop_list_en
                newLabelValue_fr = q_label_fr + final_sorted_crop_list_fr
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_fr)

            if q_name_CROPS_SOLD in str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en_CROPS_SOLD + final_sorted_crop_list_en_CROPS_SOLD
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)

                newLabelValue_fr = q_label_fr_CROPS_SOLD + final_sorted_crop_list_fr_CROPS_SOLD
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_fr)

        replacementTextKeyPairs = {'$CROP CODES$': final_sorted_crop_list_DATASETCODES,
                                  '$CROP SOLD CODES$' : final_sorted_crop_list__CROPS_SOLD_DATASETCODES}
        for i in range(number_of_columns):
            for k in range(number_of_rows):

                cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue

        workbook["Crop list"].sheet_state = 'hidden'
        workbook.save(questionnaire_file)
        return n_of_choices
    
    

    elif 'Label (AR)' in columns:
        crop_list_df = crop_list_df[['Select top 10 crops ','GeoPoll code','Dataset code','Label (EN)','Label (AR)']]
        q_name = "crp_main"
        q_label_en = "What has been the main crop that your household has grown for food and income generation in $season$, if any? \n [OPERATOR: SINGLE RESPONSE. \"NO CROP PRODUCTION\" IS IN LAST OPTIONS AT THE END.] \n \n"
        q_label_ar = "ما هو المحصول الرئيسي الذي تزرعه أسرتك من أجل الغذاء وتوليد الدخل في موسم $ Season $ ، إن وجد؟ \n [المشغل: استجابة واحدة. 'لا إنتاج للمحصول' في آخر الخيارات في النهاية.] \n \n"


        q_name_CROPS_SOLD = "crp_salesmain"
        q_label_en_CROPS_SOLD = "What has been the main crop that you have sold over the last 3 months?\n[OPERATOR: SELECT ALL CROPS LISTED BY THE RESPONDENT.]\n\n"
        q_label_ar_CROPS_SOLD = "ما هو المحصول الرئيسي الذي بعته خلال الأشهر الثلاثة الماضية؟ \n [حدد جميع المحاصيل التي سجلها المجيب.] \n \n"


        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["survey"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns



        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)



        sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','GeoPoll code'])
        #print(sorted_crop_list_df)
        
        
        
        ##########NEWWWWWW########
        sorted_crop_list_df = sorted_crop_list_df.reset_index()
        sorted_crop_list_df['GeoPoll code'] = sorted_crop_list_df.index + 1
        print(sorted_crop_list_df['GeoPoll code'])
        ##########NEWWWWWW########
        
        
        sorted_crop_list_df['combined (EN)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (EN)']
        sorted_crop_list_df['combined (AR)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (AR)']

        
        sorted_crop_list_en = sorted_crop_list_df[['combined (EN)']].values.tolist()
        additional_crop_list_en = [["91)No crop production"], ["92)DON'T KNOW"], ["93)REFUSED"]]
        sorted_crop_list_additional_en = sorted_crop_list_en + additional_crop_list_en
        
        sorted_crop_list_ar = sorted_crop_list_df[['combined (AR)']].values.tolist()
        additional_crop_list_ar = [["91)لا إنتاج المحاصيل"], ["92)لا أعرف"], ["93)رفض"]]
        sorted_crop_list_additional_ar = sorted_crop_list_ar + additional_crop_list_ar
        
        
        
        #print(sorted_crop_list)
        final_sorted_crop_list_en = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_en]))
        final_sorted_crop_list_ar = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_ar]))

        #print(final_sorted_crop_list)
        #print(crop_list_df)

        sorted_crop_list_df['combined_DATASETCODES'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Dataset code'].astype(str)
        sorted_crop_list_DATASETCODES = sorted_crop_list_df[['combined_DATASETCODES']].values.tolist()
        #print(sorted_crop_list_DATASETCODES)
        additional_crop_list_DATASETCODES = [["91)777"], ["92)888"], ["93)999"]]
        sorted_crop_list_additional_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_DATASETCODES
        final_sorted_crop_list_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_DATASETCODES]))
        
        
        

        #CROPS SOLD
        sorted_crop_list_en_CROPS_SOLD = list(sorted_crop_list_en)
        additional_en_CROPS_SOLD = [["91)DON'T KNOW"], ["92)REFUSED"]]
        sorted_crop_list_en_CROPS_SOLD.extend(additional_en_CROPS_SOLD)
        final_sorted_crop_list_en_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en_CROPS_SOLD]))
        #print(sorted_crop_list_en_CROPS_SOLD)
        #print(final_sorted_crop_list_en_CROPS_SOLD)

        sorted_crop_list_ar_CROPS_SOLD = list(sorted_crop_list_ar)
        additional_ar_CROPS_SOLD = [["91) لا اعرف"],["92)رفض الإجابة"]]
                                     
        sorted_crop_list_ar_CROPS_SOLD.extend(additional_ar_CROPS_SOLD)
        final_sorted_crop_list_ar_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_ar_CROPS_SOLD]))
        #print(sorted_crop_list_ar_CROPS_SOLD)
        #print(final_sorted_crop_list_ar_CROPS_SOLD)
        
        
        additional_crop_list_CROPS_SOLD_DATASETCODES = [["91)888"], ["92)999"]]
        sorted_crop_list_additional_CROPS_SOLD_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_CROPS_SOLD_DATASETCODES
        final_sorted_crop_list__CROPS_SOLD_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_CROPS_SOLD_DATASETCODES]))
    
             
        
        #OLD VERSION
        # sorted_crop_list_en = sorted_crop_list_df[['combined (EN)']].values.tolist()
        # sorted_crop_list_ar = sorted_crop_list_df[['combined (AR)']].values.tolist()
        # #print(sorted_crop_list)
        # final_sorted_crop_list_en = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en]))
        # final_sorted_crop_list_ar = '\n'.join(map(str, [i[0] for i in sorted_crop_list_ar]))

        # #print(final_sorted_crop_list)
        # #print(crop_list_df)


        # #CROPS SOLD
        # sorted_crop_list_en_CROPS_SOLD = list(sorted_crop_list_en)
        # additional_en_CROPS_SOLD = [["92)DON'T KNOW"], ["93)REFUSED"]]
        # sorted_crop_list_en_CROPS_SOLD.extend(additional_en_CROPS_SOLD)
        # final_sorted_crop_list_en_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en_CROPS_SOLD]))
        # #print(sorted_crop_list_en_CROPS_SOLD)
        # #print(final_sorted_crop_list_en_CROPS_SOLD)

        # sorted_crop_list_ar_CROPS_SOLD = list(sorted_crop_list_ar)
        # additional_ar_CROPS_SOLD = [["92) لا اعرف"],["93)رفض الإجابة"]]
                                     
        # sorted_crop_list_ar_CROPS_SOLD.extend(additional_ar_CROPS_SOLD)
        # final_sorted_crop_list_ar_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_ar_CROPS_SOLD]))
        # #print(sorted_crop_list_ar_CROPS_SOLD)
        # #print(final_sorted_crop_list_ar_CROPS_SOLD)


        # # Iterate over the columns and rows, search
        # # for the text and replace
        
        ## OLD VERSION

        for k in range(number_of_rows):

            questionValue = str(worksheet[get_column_letter(2)+str(k+1)].value)
            labelValue = str(worksheet[get_column_letter(4)+str(k+1)].value)
                #print(cellValue)

            if q_name == str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en + final_sorted_crop_list_en
                newLabelValue_ar = q_label_ar + final_sorted_crop_list_ar
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_ar)

            if q_name_CROPS_SOLD in str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en_CROPS_SOLD + final_sorted_crop_list_en_CROPS_SOLD
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)

                newLabelValue_ar = q_label_ar_CROPS_SOLD + final_sorted_crop_list_ar_CROPS_SOLD
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_ar)


        replacementTextKeyPairs = {'$CROP CODES$': final_sorted_crop_list_DATASETCODES,
                                  '$CROP SOLD CODES$' : final_sorted_crop_list__CROPS_SOLD_DATASETCODES}
        for i in range(number_of_columns):
            for k in range(number_of_rows):

                cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue

        workbook["Crop list"].sheet_state = 'hidden'
        workbook.save(questionnaire_file)
        return n_of_choices
    
    
    

    elif 'Label (ES)' in columns:
        crop_list_df = crop_list_df[['Select top 10 crops ','GeoPoll code','Dataset code','Label (EN)','Label (ES)']]
        q_name = "crp_main"
        q_label_en = "What has been the main crop that your household has grown for food and income generation in $season$, if any? \n [OPERATOR: SINGLE RESPONSE. \"NO CROP PRODUCTION\" IS IN LAST OPTIONS AT THE END.] \n \n"
        q_label_es = "¿Cuál ha sido el principal producto que ha vendido en los últimos 3 meses?\n\n[OPERADOR: ""Ningún cultivo vendido en los últimos 3 meses"" APARECE AL FINAL DE LA LISTA DE OPCIONES.  ES POSIBLE SELECCIONAR MÚLTIPLES RESPUESTAS. SELECCIONE TODOS LOS CULTIVOS QUE MENCIONA EL ENCUESTADO.]\n\n"

        q_name_CROPS_SOLD = "crp_salesmain"
        q_label_en_CROPS_SOLD = "What has been the main crop that you have sold over the last 3 months?\n[OPERATOR: \"No crop sold in the last 3 months\" APPEARS AT THE END OF THE LIST OF OPTIONS. SELECT ALL CROPS LISTED BY THE RESPONDENT.]\n\n"
        q_label_es_CROPS_SOLD = "¿Cuál ha sido el principal producto que ha vendido en los últimos 3 meses?\n\n[OPERADOR: ES POSIBLE SELECCIONAR MÚLTIPLES RESPUESTAS. SELECCIONE TODOS LOS CULTIVOS QUE MENCIONA EL ENCUESTADO.]\n\n"


        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["survey"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns



        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)



        sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','GeoPoll code'])
        #print(sorted_crop_list_df)
        
        
        ##########NEWWWWWW########
        sorted_crop_list_df = sorted_crop_list_df.reset_index()
        sorted_crop_list_df['GeoPoll code'] = sorted_crop_list_df.index + 1
        #print(sorted_crop_list_df['GeoPoll code'])
        ##########NEWWWWWW########
        
        
        sorted_crop_list_df['combined (EN)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (EN)']
        sorted_crop_list_df['combined (ES)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (ES)']

        sorted_crop_list_en = sorted_crop_list_df[['combined (EN)']].values.tolist()
        additional_crop_list_en = [["91)No crop production"], ["92)DON'T KNOW"], ["93)REFUSED"]]
        sorted_crop_list_additional_en = sorted_crop_list_en + additional_crop_list_en
        
        
        sorted_crop_list_es = sorted_crop_list_df[['combined (ES)']].values.tolist()
        additional_crop_list_es = [["91)Sin producción de cultivos"], ["92)no lo sé"], ["93)RECHAZAR"]]
        sorted_crop_list_additional_es = sorted_crop_list_es + additional_crop_list_es
        
        
        #print(sorted_crop_list)
        final_sorted_crop_list_en = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_en]))
        final_sorted_crop_list_es = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_es]))

        #print(final_sorted_crop_list)
        #print(crop_list_df)


        
        sorted_crop_list_df['combined_DATASETCODES'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Dataset code'].astype(str)
        sorted_crop_list_DATASETCODES = sorted_crop_list_df[['combined_DATASETCODES']].values.tolist()
        #print(sorted_crop_list_DATASETCODES)
        additional_crop_list_DATASETCODES = [["91)777"], ["92)888"], ["93)999"]]
        sorted_crop_list_additional_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_DATASETCODES
        final_sorted_crop_list_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_DATASETCODES]))
                
        
        
        
        #CROPS SOLD
        sorted_crop_list_en_CROPS_SOLD = list(sorted_crop_list_en)
        additional_en_CROPS_SOLD = [["91)DON'T KNOW"], ["92)REFUSED"]]
        sorted_crop_list_en_CROPS_SOLD.extend(additional_en_CROPS_SOLD)
        final_sorted_crop_list_en_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en_CROPS_SOLD]))
        #print(sorted_crop_list_en_CROPS_SOLD)
        #print(final_sorted_crop_list_en_CROPS_SOLD)

        sorted_crop_list_es_CROPS_SOLD = list(sorted_crop_list_es)
        additional_es_CROPS_SOLD = [["91)NO SÉ"], ["92)RECHAZAR"]]
        sorted_crop_list_es_CROPS_SOLD.extend(additional_es_CROPS_SOLD)
        final_sorted_crop_list_es_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_es_CROPS_SOLD]))
        
        
        
        additional_crop_list_CROPS_SOLD_DATASETCODES = [["91)888"], ["92)999"]]
        sorted_crop_list_additional_CROPS_SOLD_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_CROPS_SOLD_DATASETCODES
        final_sorted_crop_list__CROPS_SOLD_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_CROPS_SOLD_DATASETCODES]))
        
        
        # sorted_crop_list_en = sorted_crop_list_df[['combined (EN)']].values.tolist()
        # sorted_crop_list_es = sorted_crop_list_df[['combined (ES)']].values.tolist()
        # #print(sorted_crop_list)
        # final_sorted_crop_list_en = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en]))
        # final_sorted_crop_list_es = '\n'.join(map(str, [i[0] for i in sorted_crop_list_es]))

        # #print(final_sorted_crop_list)
        # #print(crop_list_df)


        # #CROPS SOLD
        # sorted_crop_list_en_CROPS_SOLD = list(sorted_crop_list_en)
        # additional_en_CROPS_SOLD = [["92)DON'T KNOW"], ["93)REFUSED"]]
        # sorted_crop_list_en_CROPS_SOLD.extend(additional_en_CROPS_SOLD)
        # final_sorted_crop_list_en_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en_CROPS_SOLD]))
        # #print(sorted_crop_list_en_CROPS_SOLD)
        # #print(final_sorted_crop_list_en_CROPS_SOLD)

        # sorted_crop_list_es_CROPS_SOLD = list(sorted_crop_list_es)
        # additional_es_CROPS_SOLD = [["92)NO SÉ"], ["93)RECHAZAR"]]
        # sorted_crop_list_es_CROPS_SOLD.extend(additional_es_CROPS_SOLD)
        # final_sorted_crop_list_es_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_es_CROPS_SOLD]))


        # # Iterate over the columns and rows, search
        # # for the text and replace

        for k in range(number_of_rows):

            questionValue = str(worksheet[get_column_letter(2)+str(k+1)].value)
            labelValue = str(worksheet[get_column_letter(4)+str(k+1)].value)
                #print(cellValue)

            if q_name == str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en + final_sorted_crop_list_en
                newLabelValue_es = q_label_es + final_sorted_crop_list_es
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_es)

            if q_name_CROPS_SOLD in str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en_CROPS_SOLD + final_sorted_crop_list_en_CROPS_SOLD
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)

                newLabelValue_es = q_label_es_CROPS_SOLD + final_sorted_crop_list_es_CROPS_SOLD
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_es)
                
        replacementTextKeyPairs = {'$CROP CODES$': final_sorted_crop_list_DATASETCODES,
                                  '$CROP SOLD CODES$' : final_sorted_crop_list__CROPS_SOLD_DATASETCODES}
        for i in range(number_of_columns):
            for k in range(number_of_rows):

                cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue

        workbook["Crop list"].sheet_state = 'hidden'
        workbook.save(questionnaire_file)
        return n_of_choices        


    
    

    elif 'Label (PT)' in columns:
        crop_list_df = crop_list_df[['Select top 10 crops ','GeoPoll code','Dataset code','Label (EN)','Label (PT)']]
        q_name = "crp_main"
        q_label_en = "What has been the main crop that your household has grown for food and income generation in $season$, if any? \n [OPERATOR: SINGLE RESPONSE. \"NO CROP PRODUCTION\" IS IN LAST OPTIONS AT THE END.] \n \n"
        q_label_es = "Qual a principal cultura que o seu agregado familiar cultivou para a produção de alimentos ou geração de renda na $season PT$?\n\n[OPERADOR(A): RESPOSTA ÚNICA. ""NENHUMA PRODUÇÃO AGRÍCOLA"" APARECE NO FINAL DA LISTA DAS OPÇÕES.]\n\n"

        q_name_CROPS_SOLD = "crp_salesmain"
        q_label_en_CROPS_SOLD = "What has been the main crop that you have sold over the last 3 months?\n[OPERATOR: \"No crop sold in the last 3 months\" APPEARS AT THE END OF THE LIST OF OPTIONS. SELECT ALL CROPS LISTED BY THE RESPONDENT.]\n\n"
        q_label_es_CROPS_SOLD = "Qual a principal cultura que o(a) senhor(a) vendeu nos últimos 3 meses?\n\n[OPERADOR(A) RESPOSTA ÚNICA.]\n\n"


        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["survey"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns



        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)



        sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','GeoPoll code'])
        #print(sorted_crop_list_df)
        
        
        ##########NEWWWWWW########
        sorted_crop_list_df = sorted_crop_list_df.reset_index()
        sorted_crop_list_df['GeoPoll code'] = sorted_crop_list_df.index + 1
        #print(sorted_crop_list_df['GeoPoll code'])
        ##########NEWWWWWW########
        
        
        sorted_crop_list_df['combined (EN)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (EN)']
        sorted_crop_list_df['combined (PT)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (PT)']

        sorted_crop_list_en = sorted_crop_list_df[['combined (EN)']].values.tolist()
        additional_crop_list_en = [["91)No crop production"], ["92)DON'T KNOW"], ["93)REFUSED"]]
        sorted_crop_list_additional_en = sorted_crop_list_en + additional_crop_list_en
        
        
        sorted_crop_list_es = sorted_crop_list_df[['combined (PT)']].values.tolist()
        additional_crop_list_es = [["91)Sin producción de cultivos"], ["92)no lo sé"], ["93)RECHAZAR"]]
        sorted_crop_list_additional_es = sorted_crop_list_es + additional_crop_list_es
        
        
        #print(sorted_crop_list)
        final_sorted_crop_list_en = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_en]))
        final_sorted_crop_list_es = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_es]))

        #print(final_sorted_crop_list)
        #print(crop_list_df)


        
        sorted_crop_list_df['combined_DATASETCODES'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Dataset code'].astype(str)
        sorted_crop_list_DATASETCODES = sorted_crop_list_df[['combined_DATASETCODES']].values.tolist()
        #print(sorted_crop_list_DATASETCODES)
        additional_crop_list_DATASETCODES = [["91)777"], ["92)888"], ["93)999"]]
        sorted_crop_list_additional_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_DATASETCODES
        final_sorted_crop_list_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_DATASETCODES]))
                
        
        
        
        #CROPS SOLD
        sorted_crop_list_en_CROPS_SOLD = list(sorted_crop_list_en)
        additional_en_CROPS_SOLD = [["91)DON'T KNOW"], ["92)REFUSED"]]
        sorted_crop_list_en_CROPS_SOLD.extend(additional_en_CROPS_SOLD)
        final_sorted_crop_list_en_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en_CROPS_SOLD]))
        #print(sorted_crop_list_en_CROPS_SOLD)
        #print(final_sorted_crop_list_en_CROPS_SOLD)

        sorted_crop_list_es_CROPS_SOLD = list(sorted_crop_list_es)
        additional_es_CROPS_SOLD = [["91)NO SÉ"], ["92)RECHAZAR"]]
        sorted_crop_list_es_CROPS_SOLD.extend(additional_es_CROPS_SOLD)
        final_sorted_crop_list_es_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_es_CROPS_SOLD]))
        
        
        
        additional_crop_list_CROPS_SOLD_DATASETCODES = [["91)888"], ["92)999"]]
        sorted_crop_list_additional_CROPS_SOLD_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_CROPS_SOLD_DATASETCODES
        final_sorted_crop_list__CROPS_SOLD_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_CROPS_SOLD_DATASETCODES]))
        
        
        # sorted_crop_list_en = sorted_crop_list_df[['combined (EN)']].values.tolist()
        # sorted_crop_list_es = sorted_crop_list_df[['combined (ES)']].values.tolist()
        # #print(sorted_crop_list)
        # final_sorted_crop_list_en = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en]))
        # final_sorted_crop_list_es = '\n'.join(map(str, [i[0] for i in sorted_crop_list_es]))

        # #print(final_sorted_crop_list)
        # #print(crop_list_df)


        # #CROPS SOLD
        # sorted_crop_list_en_CROPS_SOLD = list(sorted_crop_list_en)
        # additional_en_CROPS_SOLD = [["92)DON'T KNOW"], ["93)REFUSED"]]
        # sorted_crop_list_en_CROPS_SOLD.extend(additional_en_CROPS_SOLD)
        # final_sorted_crop_list_en_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en_CROPS_SOLD]))
        # #print(sorted_crop_list_en_CROPS_SOLD)
        # #print(final_sorted_crop_list_en_CROPS_SOLD)

        # sorted_crop_list_es_CROPS_SOLD = list(sorted_crop_list_es)
        # additional_es_CROPS_SOLD = [["92)NO SÉ"], ["93)RECHAZAR"]]
        # sorted_crop_list_es_CROPS_SOLD.extend(additional_es_CROPS_SOLD)
        # final_sorted_crop_list_es_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_es_CROPS_SOLD]))


        # # Iterate over the columns and rows, search
        # # for the text and replace

        for k in range(number_of_rows):

            questionValue = str(worksheet[get_column_letter(2)+str(k+1)].value)
            labelValue = str(worksheet[get_column_letter(4)+str(k+1)].value)
                #print(cellValue)

            if q_name == str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en + final_sorted_crop_list_en
                newLabelValue_es = q_label_es + final_sorted_crop_list_es
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_es)

            if q_name_CROPS_SOLD in str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en_CROPS_SOLD + final_sorted_crop_list_en_CROPS_SOLD
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)

                newLabelValue_es = q_label_es_CROPS_SOLD + final_sorted_crop_list_es_CROPS_SOLD
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_es)
                
        replacementTextKeyPairs = {'$CROP CODES$': final_sorted_crop_list_DATASETCODES,
                                  '$CROP SOLD CODES$' : final_sorted_crop_list__CROPS_SOLD_DATASETCODES}
        for i in range(number_of_columns):
            for k in range(number_of_rows):

                cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue

        workbook["Crop list"].sheet_state = 'hidden'
        workbook.save(questionnaire_file)
        return n_of_choices        

    
# Function to check all questions in two Excel files
def check_all_questions(questionnaire_file, template_file, result_file):
     # Read Excel files and create DataFrames for questionnaire and template
    questionnaire_df = pd.read_excel(pd.ExcelFile(questionnaire_file), sheet_name='survey', skiprows=2).dropna(subset=['Q Name'])
    questionnaire_df
    
    template_df = pd.read_excel(pd.ExcelFile(template_file), sheet_name='survey', skiprows=2).dropna(subset=['Q Name'])
    writer = pd.ExcelWriter(result_file, engine = 'xlsxwriter')
    
# Function to detect the language
    language = detect_language(questionnaire_file)
    result_brief = ""
    result_details = ""
    
# Perform an inner join on the two DataFrames using 'Q Name' as the key    
    outer_join = pd.merge(template_df, questionnaire_df, how='inner', on=['Q Name', 'Q Name'])
    
# Below code segments perform comparisons for different columns like 'Suggested Qname', 'English', 'Length', etc.
# For each column, a new DataFrame is created, and the match/mismatch is identified
# The comparison results are written to Excel, and messages are appended to result_brief and result_details    


# Create an empty DataFrame to store comparison results
    comparaison_result = pd.DataFrame()
    
# Store the 'Q Name' column from 'outer_join' DataFrame into 'question_name'
    comparaison_result['question_name'] = outer_join['Q Name']

# Create a DataFrame for the comparison of 'Suggested Qname'
    comparaison_result_SQName = pd.DataFrame()
    
# Copy 'Q Name', 'Suggested Qname_x', and 'Suggested Qname_y' columns from 'outer_join'
    comparaison_result_SQName['Q Name'] = outer_join['Q Name']
    comparaison_result_SQName['Suggested Qname_x'] = outer_join['Suggested Qname_x']
    comparaison_result_SQName['Suggested Qname_y'] = outer_join['Suggested Qname_y']
    
# Check if 'Suggested Qname_x' is equal to 'Suggested Qname_y', and store the result
    comparaison_result_SQName['Suggested Qname_match'] = np.where(outer_join['Suggested Qname_x'] == outer_join['Suggested Qname_y'], 'True', 'False')
    
# Remove rows where both 'Suggested Qname_x' and 'Suggested Qname_y' are NaN
    comparaison_result_SQName = comparaison_result_SQName.dropna(subset=['Suggested Qname_x','Suggested Qname_y'], how='all')

# Write the comparison results to an Excel file with the specified sheet name
    comparaison_result_SQName.to_excel(writer, sheet_name='Suggested Qname_match', engine='xlsxwriter')
    
# Retrieve a list of question names where 'Suggested Qname_match' is False
    comparaison_result_SQName_list = list(comparaison_result_SQName.loc[((comparaison_result_SQName["Suggested Qname_match"] == "False")),['Q Name']].values.tolist())

# If there are differences in 'Suggested Qname', update result_brief and result_details with the differences
    if len(comparaison_result_SQName_list) > 0:
        message = "\ndetecting differences in column: Suggested Name\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_SQName_list)

# Similar to the above, create a DataFrame for comparison of 'English' columns
    comparaison_result_English = pd.DataFrame()
    comparaison_result_English['Q Name'] = outer_join['Q Name']
    comparaison_result_English['English_x'] = outer_join['English_x']
    comparaison_result_English['English_y'] = outer_join['English_y']
    comparaison_result_English['English_match'] = np.where(outer_join['English_x'] == outer_join['English_y'], 'True', 'False')
    comparaison_result_English = comparaison_result_English.dropna(subset=['English_x','English_y'], how='all')
    comparaison_result_English.to_excel(writer, sheet_name='English_match', engine='xlsxwriter')
    comparaison_result_English_list = list(comparaison_result_English.loc[((comparaison_result_English["English_match"] == "False")),['Q Name']].values.tolist())
 
# Similar to the above, create a DataFrame for comparison of 'English' columns   
    if len(comparaison_result_English_list) > 0:
        message = "\ndetecting differences in column: English\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_English_list)
    comparaison_result = pd.DataFrame()
    comparaison_result['question_name'] = outer_join['Q Name']
    
    
# Re-create an empty DataFrame 'comparaison_result' and store the 'Q Name' column again
# Note: This line is a repetition of the first two lines,   
    
    comparaison_result_Length = pd.DataFrame()
    comparaison_result_Length['Q Name'] = outer_join['Q Name']
    comparaison_result_Length['Length_x'] = outer_join['Length_x']
    comparaison_result_Length['Length_y'] = outer_join['Length_y']
    comparaison_result_Length['Length_match'] = np.where(outer_join['Length_x'] == outer_join['Length_y'], 'True', 'False')
    comparaison_result_Length = comparaison_result_Length.dropna(subset=['Length_x','Length_y'], how='all')
    comparaison_result_Length.to_excel(writer, sheet_name='Length_match', engine='xlsxwriter')
    comparaison_result_Length_list = list(comparaison_result_Length.loc[((comparaison_result_Length["Length_match"] == "False")),['Q Name']].values.tolist())

    if len(comparaison_result_Length_list) > 0:
        message = "\ndetecting differences in column: Length\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_Length_list)
        
        
        
    comparaison_result_QType = pd.DataFrame()
    comparaison_result_QType['Q Name'] = outer_join['Q Name']
    comparaison_result_QType['Q Type_x'] = outer_join['Q Type_x']
    comparaison_result_QType['Q Type_y'] = outer_join['Q Type_y']
    comparaison_result_QType['Q Type_match'] = np.where(outer_join['Q Type_x'] == outer_join['Q Type_y'], 'True', 'False')
    comparaison_result_QType = comparaison_result_QType.dropna(subset=['Q Type_x','Q Type_y'], how='all')
    comparaison_result_QType.to_excel(writer, sheet_name='Q Type_match', engine='xlsxwriter')
    comparaison_result_QType_list = list(comparaison_result_QType.loc[((comparaison_result_QType["Q Type_match"] == "False")),['Q Name']].values.tolist())

    if len(comparaison_result_QType_list) > 0:
        message = "\ndetecting differences in column: Q Type\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_QType_list)
        
        
    comparaison_result_Randomize = pd.DataFrame()
    comparaison_result_Randomize['Q Name'] = outer_join['Q Name']
    comparaison_result_Randomize['Randomize_x'] = outer_join['Randomize_x']
    comparaison_result_Randomize['Randomize_y'] = outer_join['Randomize_y']
    comparaison_result_Randomize['Randomize_match'] = np.where(outer_join['Randomize_x'] == outer_join['Randomize_y'], 'True', 'False')
    comparaison_result_Randomize = comparaison_result_Randomize.dropna(subset=['Randomize_x','Randomize_y'], how='all')
    comparaison_result_Randomize.to_excel(writer, sheet_name='Randomize_match', engine='xlsxwriter')
    comparaison_result_Randomize_list = list(comparaison_result_Randomize.loc[((comparaison_result_Randomize["Randomize_match"] == "False")),['Q Name']].values.tolist())

    if len(comparaison_result_Randomize_list) > 0:
        message = "\ndetecting differences in column: Randomize\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_Randomize_list)
        
    comparaison_result_Conditional = pd.DataFrame()
    comparaison_result_Conditional['Q Name'] = outer_join['Q Name']
    comparaison_result_Conditional['Conditional_x'] = outer_join['Conditional_x']
    comparaison_result_Conditional['Conditional_y'] = outer_join['Conditional_y']
    comparaison_result_Conditional['Conditional_match'] = np.where(outer_join['Conditional_x'] == outer_join['Conditional_y'], 'True', 'False')
    comparaison_result_Conditional = comparaison_result_Conditional.dropna(subset=['Conditional_x','Conditional_y'], how='all')
    comparaison_result_Conditional.to_excel(writer, sheet_name='Conditional_match', engine='xlsxwriter')
    comparaison_result_Conditional_list = list(comparaison_result_Conditional.loc[((comparaison_result_Conditional["Conditional_match"] == "False")),['Q Name']].values.tolist())

    if len(comparaison_result_Conditional_list) > 0:
        message = "\ndetecting differences in column: Conditional\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_Conditional_list)
        
        
        
    comparaison_result_ProgrammingInstructions = pd.DataFrame()
    comparaison_result_ProgrammingInstructions['Q Name'] = outer_join['Q Name']
    comparaison_result_ProgrammingInstructions['Programming Instructions_x'] = outer_join['Programming Instructions_x']
    comparaison_result_ProgrammingInstructions['Programming Instructions_y'] = outer_join['Programming Instructions_y']
    comparaison_result_ProgrammingInstructions['Programming Instructions_match'] = np.where(outer_join['Programming Instructions_x'] == outer_join['Programming Instructions_y'], 'True', 'False')
    comparaison_result_ProgrammingInstructions = comparaison_result_ProgrammingInstructions.dropna(subset=['Programming Instructions_x','Programming Instructions_y'], how='all')
    comparaison_result_ProgrammingInstructions.to_excel(writer, sheet_name='Programming Instructions_match', engine='xlsxwriter')
    comparaison_result_ProgrammingInstructions_list = list(comparaison_result_ProgrammingInstructions.loc[((comparaison_result_ProgrammingInstructions["Programming Instructions_match"] == "False")),['Q Name']].values.tolist())

    if len(comparaison_result_ProgrammingInstructions_list) > 0:
        message = "\ndetecting differences in column: Programming Instructions\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_ProgrammingInstructions_list)
        
        
        
        
    comparaison_result_SkipPattern = pd.DataFrame()
    comparaison_result_SkipPattern['Q Name'] = outer_join['Q Name']
    comparaison_result_SkipPattern['Skip Pattern_x'] = outer_join['Skip Pattern_x']
    comparaison_result_SkipPattern['Skip Pattern_y'] = outer_join['Skip Pattern_y']
    comparaison_result_SkipPattern['Skip Pattern_match'] = np.where(outer_join['Skip Pattern_x'] == outer_join['Skip Pattern_y'], 'True', 'False')
    comparaison_result_SkipPattern = comparaison_result_SkipPattern.dropna(subset=['Skip Pattern_x','Skip Pattern_y'], how='all')
    comparaison_result_SkipPattern.to_excel(writer, sheet_name='Skip Pattern_match', engine='xlsxwriter')
    comparaison_result_SkipPattern_list = list(comparaison_result_SkipPattern.loc[((comparaison_result_SkipPattern["Skip Pattern_match"] == "False")),['Q Name']].values.tolist())

    if len(comparaison_result_SkipPattern_list) > 0:
        message = "\ndetecting differences in column: Skip Pattern\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_SkipPattern_list)
        
        
        
        
    comparaison_result_Codes = pd.DataFrame()
    comparaison_result_Codes['Q Name'] = outer_join['Q Name']
    comparaison_result_Codes['Codes_x'] = outer_join['Codes_x']
    comparaison_result_Codes['Codes_y'] = outer_join['Codes_y']
    comparaison_result_Codes['Codes_match'] = np.where(outer_join['Codes_x'] == outer_join['Codes_y'], 'True', 'False')
    comparaison_result_Codes = comparaison_result_Codes.dropna(subset=['Codes_x','Codes_y'], how='all')
    comparaison_result_Codes.to_excel(writer, sheet_name='Codes_match', engine='xlsxwriter')
    comparaison_result_Codes_list = list(comparaison_result_Codes.loc[((comparaison_result_Codes["Codes_match"] == "False")),['Q Name']].values.tolist())

    if len(comparaison_result_Codes_list) > 0:
        message = "\ndetecting differences in column: Codes\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_Codes_list)
        
        
        
    comparaison_result_Defaultskippatternsconditional  = pd.DataFrame()
    comparaison_result_Defaultskippatternsconditional['Q Name'] = outer_join['Q Name']
    comparaison_result_Defaultskippatternsconditional['Default skip patterns & conditional _x'] = outer_join['Default skip patterns & conditional _x']
    comparaison_result_Defaultskippatternsconditional['Default skip patterns & conditional _y'] = outer_join['Default skip patterns & conditional _y']
    comparaison_result_Defaultskippatternsconditional['Default skip patterns & conditional _match'] = np.where(outer_join['Default skip patterns & conditional _x'] == outer_join['Default skip patterns & conditional _y'], 'True', 'False')
    comparaison_result_Defaultskippatternsconditional  = comparaison_result_Defaultskippatternsconditional.dropna(subset=['Default skip patterns & conditional _x','Default skip patterns & conditional _y'], how='all')
    comparaison_result_Defaultskippatternsconditional.to_excel(writer, sheet_name='Default skip patterns_match', engine='xlsxwriter')
    comparaison_result_Defaultskippatternsconditional_list = list(comparaison_result_Defaultskippatternsconditional.loc[((comparaison_result_Defaultskippatternsconditional["Default skip patterns & conditional _match"] == "False")),['Q Name']].values.tolist())

    if len(comparaison_result_Defaultskippatternsconditional_list) > 0:
        message = "\ndetecting differences in column: Default skip patterns & conditional \n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_Defaultskippatternsconditional_list)
        
        
        
        
    #comparaison_result_Specifyskippatternvariable = pd.DataFrame()
    #comparaison_result_Specifyskippatternvariable['Q Name'] = outer_join['Q Name']
    #comparaison_result_Specifyskippatternvariable['Specify skip pattern variable (from blue text)_x'] = outer_join['Specify skip pattern variable (from blue text)_x']
    #comparaison_result_Specifyskippatternvariable['Specify skip pattern variable (from blue text)_y'] = outer_join['Specify skip pattern variable (from blue text)_y']
    #comparaison_result_Specifyskippatternvariable['Specify skip pattern variable_match'] = np.where(outer_join['Specify skip pattern variable (from blue text)_x'] == outer_join['Specify skip pattern variable (from blue text)_y'], 'True', 'False')
    #comparaison_result_Specifyskippatternvariable = comparaison_result_Specifyskippatternvariable.dropna(subset=['Specify skip pattern variable (from blue text)_x','Specify skip pattern variable (from blue text)_y'], how='all')
    #comparaison_result_Specifyskippatternvariable.to_excel(writer, sheet_name='Specify skip pattern_match', engine='xlsxwriter')
    #comparaison_result_Specifyskippatternvariable_list = list(comparaison_result_Specifyskippatternvariable.loc[((comparaison_result_Specifyskippatternvariable["Specify skip pattern variable_match"] == "False")),['Q Name']].values.tolist())

    #if len(comparaison_result_Specifyskippatternvariable_list) > 0:
    #    message = "\ndetecting differences in column: Specify skip pattern variable (from blue text)\n"
    #    result_brief = result_brief + message
    #    result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_Specifyskippatternvariable_list)
        
        
        
    comparaison_result_Mandatory = pd.DataFrame()
    comparaison_result_Mandatory['Q Name'] = outer_join['Q Name']
    comparaison_result_Mandatory['Mandatory_x'] = outer_join['Mandatory_x']
    comparaison_result_Mandatory['Mandatory_y'] = outer_join['Mandatory_y']
    comparaison_result_Mandatory['Mandatory_match'] = np.where(outer_join['Mandatory_x'] == outer_join['Mandatory_y'], 'True', 'False')
    comparaison_result_Mandatory = comparaison_result_Mandatory.dropna(subset=['Mandatory_x','Mandatory_y'], how='all')
    comparaison_result_Mandatory.to_excel(writer, sheet_name='Mandatory_match', engine='xlsxwriter')
    comparaison_result_Mandatory_list = list(comparaison_result_Mandatory.loc[((comparaison_result_Mandatory["Mandatory_match"] == "False")),['Q Name']].values.tolist())

    if len(comparaison_result_Mandatory_list) > 0:
        message = "\ndetecting differences in column: Mandatory\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_Mandatory_list)    
    
    
    comparaison_result_Corequestionsonly = pd.DataFrame()
    comparaison_result_Corequestionsonly['Q Name'] = outer_join['Q Name']
    comparaison_result_Corequestionsonly['Core questions only_x'] = outer_join['Core questions only_x']
    comparaison_result_Corequestionsonly['Core questions only_y'] = outer_join['Core questions only_y']
    comparaison_result_Corequestionsonly['Core questions only_match'] = np.where(outer_join['Core questions only_x'] == outer_join['Core questions only_y'], 'True', 'False')
    comparaison_result_Corequestionsonly = comparaison_result_Corequestionsonly.dropna(subset=['Core questions only_x','Core questions only_y'], how='all')
    comparaison_result_Corequestionsonly.to_excel(writer, sheet_name='Core questions only_match', engine='xlsxwriter')
    comparaison_result_Corequestionsonly_list = list(comparaison_result_Corequestionsonly.loc[((comparaison_result_Corequestionsonly["Core questions only_match"] == "False")),['Q Name']].values.tolist())

    if len(comparaison_result_Corequestionsonly_list) > 0:
        message = "\ndetecting differences in column: Core questions only\n"
        result_brief = result_brief + message
        result_details = result_details + message + '; '.join(str(e) for e in comparaison_result_Corequestionsonly_list)
        
        
    
    #writer.save()
    #writer.close()
    return result_brief, result_details

    