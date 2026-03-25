import pandas as pd
import xlrd
#import arcpy
import openpyxl
from pandas import ExcelWriter
from datetime import datetime
import os
import numpy as np
#from arcgis.gis import GIS
#from arcgis.features import SpatialDataFrame
from shutil import copyfile
from openpyxl.utils.cell import get_column_letter
import urllib.request, json

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font



def detect_language(file_name):
    language = ""
    
    if "_en_" in file_name.lower():
        language = "en"
    elif "_fr_" in file_name.lower():
        language = "fr"
    elif "_es_" in file_name.lower():
        language = "es"
    elif "_ar_" in file_name.lower():
        language = "ar"
    elif "_pt_" in file_name.lower():
        language = "pt"        
        
    else:
        print("Please include the language in the filename: %s " % file_name)
        language = "en"
    return language


def detect_template(template_version,file_name):
    language = detect_language(file_name)
    
    if language == "en":
        template_questionnaire_file = r'household_questionnaire_geopoll_EN_template_' + template_version + '_ISO3.xlsx'
    elif language == "fr":
        template_questionnaire_file = r'household_questionnaire_geopoll_FR_template_' + template_version + '_ISO3.xlsx'
    elif language == "es":
        template_questionnaire_file = r'household_questionnaire_geopoll_ES_template_' + template_version + '_ISO3.xlsx'
    elif language == "ar":
        template_questionnaire_file = r'household_questionnaire_geopoll_AR_template_' + template_version + '_ISO3.xlsx'
    elif language == "pt":
        template_questionnaire_file = r'household_questionnaire_geopoll_PT_template_' + template_version + '_ISO3.xlsx'
    else:
        print("Please include the language in the filename: %s " % country_template)
    return template_questionnaire_file


def importallsheets(in_excel, out_gdb):
    print("max_counter: %s" % max_counter)
    
    counter = 0
    workbook = xlrd.open_workbook(in_excel)
    
    sheets = [sheet.name for sheet in workbook.sheets()]

    print('{} sheets found: {}'.format(len(sheets), ','.join(sheets)))
    
    for sheet in sheets:
        counter +=1
        
        if counter <= max_counter:
            out_table = os.path.join(
                out_gdb,
                arcpy.ValidateTableName(
                    "{0}".format(sheet),
                    out_gdb))

            print('Converting {} to {}'.format(sheet, out_table))

            arcpy.ExcelToTable_conversion(in_excel, out_table, sheet)


def make_attribute_dict(fc, code_field, value_field):
    attdict = {}
    
    with arcpy.da.SearchCursor(fc, [code_field, value_field]) as cursor:
        for row in cursor:
            attdict[row[0]] = row[1]
    return attdict

def fix_category_formatting(category):
     return category.replace("[","(").replace("]",")").replace("(specify)","").replace("/ ",", ").capitalize().replace("adps","ADPs").replace("idp","IDP").replace("covid","COVID").replace(" , ",", ").replace("staplec","staple")

def count_number_of_questions_qname(questionnaire_file):
    questionnaire_df = pd.read_excel(pd.ExcelFile(questionnaire_file), sheet_name='survey',skiprows=2)
    questionnaire_df = questionnaire_df[['Q Name']]
    list_questions = questionnaire_df.dropna().values.tolist()
    list_questions = [item for sublist in list_questions for item in sublist]
    list_questions = [s.strip() for s in list_questions]
    n_of_questions = len(list_questions)
    return list_questions, n_of_questions

def count_number_of_questions_sqname(questionnaire_file):
    # Read the Excel file and access the 'survey' sheet, skipping the first 2 rows
    questionnaire_df = pd.read_excel(pd.ExcelFile(questionnaire_file), sheet_name='survey',skiprows=2)
    # Filter the dataframe to only include the 'Suggested Qname' column
    questionnaire_df = questionnaire_df[['Suggested Qname']]
    # Remove rows with NaN values and convert the DataFrame into a list
    list_questions = questionnaire_df.dropna().values.tolist()
    # Flatten the list
    list_questions = [item for sublist in list_questions for item in sublist]
    # Remove any leading/trailing whitespace from the strings in the list
    list_questions = [s.strip() for s in list_questions]
    # Count the number of questions
    n_of_questions = len(list_questions)
    # Return the cleaned list of questions and the count of questions
    return list_questions, n_of_questions


from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

def highlight_differences_in_qname(questionnaire_file, previousround_questionnaire_file):
    # Load Excel data
    curr_df = pd.read_excel(questionnaire_file, sheet_name='survey', skiprows=2)
    prev_df = pd.read_excel(previousround_questionnaire_file, sheet_name='survey', skiprows=2)

    # Clean column names
    curr_df.columns = curr_df.columns.str.strip()
    prev_df.columns = prev_df.columns.str.strip()

    # Drop missing Q Name rows and index by it
    curr_df = curr_df.dropna(subset=['Q Name']).set_index('Q Name')
    prev_df = prev_df.dropna(subset=['Q Name']).set_index('Q Name')

    # Load workbook and worksheet
    wb = load_workbook(questionnaire_file)
    ws = wb['survey']

    # Mapping of headers to Excel column letters
    header_row = 3
    header_map = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=header_row, column=col).value
        if header:
            header_map[header.strip()] = get_column_letter(col)

    # Fields to check
    fields_to_check = [
        "Suggested Qname", "English", "Q Type", "Conditional",
        "Programming Instructions", "Codes", "Skip Pattern", "GeoPoll Comments",
        "Default skip patterns & conditional", "Specify skip pattern variable (from blue text)",
        "Estimated percentage of repondents", "Mandatory", "Duration (sec)",
        "Weighted duration", "Core questions only"
    ]

    # Highlight color
    #yellow_fill = PatternFill(start_color="fff8b3", end_color="fff8b3", fill_type="solid")
    yellow_fill = PatternFill(start_color="E0B0FF", end_color="E0B0FF", fill_type="solid")

    missing_columns = []

    # Iterate over each row
    for row in range(4, ws.max_row + 1):
        qname = ws[f"B{row}"].value
        if not qname or qname not in curr_df.index or qname not in prev_df.index:
            continue

        for field in fields_to_check:
            if field not in header_map:
                if field not in missing_columns:
                    missing_columns.append(field)
                continue

            col_letter = header_map[field]

            curr_val = str(curr_df.at[qname, field]).strip() if field in curr_df.columns and pd.notna(curr_df.at[qname, field]) else ""
            prev_val = str(prev_df.at[qname, field]).strip() if field in prev_df.columns and pd.notna(prev_df.at[qname, field]) else ""

            if curr_val != prev_val:
                ws[f"{col_letter}{row}"].fill = yellow_fill

    wb.save(questionnaire_file)

    if missing_columns:
        print("⚠️ The following columns were not found and skipped:")
        for field in missing_columns:
            print(f"  - {field}")

    print("✅ Highlighted all programming differences successfully (in purple).")

def highlight_differences_in_qname_OLD_stopped_using_10December2025(questionnaire_file, previousround_questionnaire_file):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    # Load current and previous data
    curr_df = pd.read_excel(questionnaire_file, sheet_name='survey', skiprows=2)
    prev_df = pd.read_excel(previousround_questionnaire_file, sheet_name='survey', skiprows=2)

    # Create lookup by 'Q Name'
    prev_lookup = prev_df.set_index('Q Name')[['Suggested Qname', 'English']].dropna(how='all')

    # Load workbook and worksheet
    wb = load_workbook(questionnaire_file)
    ws = wb['survey']

    # Define yellow highlight only
    yellow_fill = PatternFill(start_color="fff8b3", end_color="fff8b3", fill_type="solid")

    # Iterate through rows (starting from row 4 after skiprows+header)
    for row in range(4, ws.max_row + 1):
        qname = ws[f"B{row}"].value  # 'Q Name'
        suggested = ws[f"C{row}"]
        english = ws[f"D{row}"]

        if qname and qname in prev_lookup.index:
            prev_suggested = prev_lookup.at[qname, 'Suggested Qname']
            prev_english = prev_lookup.at[qname, 'English']

            # Highlight Suggested Qname if different
            if str(suggested.value).strip() != str(prev_suggested).strip():
                suggested.fill = yellow_fill

            # Highlight English label if different
            if str(english.value).strip() != str(prev_english).strip():
                english.fill = yellow_fill

    wb.save(questionnaire_file)
    print("✅ Highlighted only differences in Suggested Qname and English label (yellow).")

    #OLD PART OF THE SCRIPT
    # Read 'Q Name' and 'Suggested Qname' columns from both files into lists
    #questionnaire_df = pd.read_excel(pd.ExcelFile(questionnaire_file), sheet_name='survey', skiprows=2)
    #previousround_df = pd.read_excel(pd.ExcelFile(previousround_questionnaire_file), sheet_name='survey', skiprows=2)

    # Extract the 'Q Name' and 'Suggested Qname' columns into lists for the current questionnaire
    #questionnaire_qnames = questionnaire_df['Q Name'].dropna().tolist()
    #questionnaire_suggested_qnames = questionnaire_df['Suggested Qname'].dropna().tolist()
    
    # Extract the 'Q Name' and 'Suggested Qname' columns into lists for the previous round questionnaire
    #previousround_qnames = previousround_df['Q Name'].dropna().tolist()
    #previousround_suggested_qnames = previousround_df['Suggested Qname'].dropna().tolist()

    # Create lookup dictionaries using 'Q Name' as key, if 'name' column is missing
    #prev_qname_dict = dict(zip(previousround_df['Q Name'], previousround_df['Q Name']))
    #prev_sugg_dict = dict(zip(previousround_df['Q Name'], previousround_df['Suggested Qname']))
    
    # Open the questionnaire file with openpyxl
    #wb = load_workbook(filename=questionnaire_file)
    #ws = wb['survey']

    # Define a green color fill to highlight cells
    #green_fill = PatternFill(start_color="9dff9d", end_color="9dff9d", fill_type="solid")

    # Highlight cells that are in both questionnaire_qnames and previousround_qnames for 'Q Name' column
    #for row in ws.iter_rows(min_row=4, min_col=2, max_col=2):  # updated column number for 'Q Name'
    #    for cell in row:
    #        if cell.value in previousround_qnames:
    #            cell.fill = green_fill

    # Highlight cells that are in both questionnaire_suggested_qnames and previousround_suggested_qnames for 'Suggested Qname' column
    #for row in ws.iter_rows(min_row=4, min_col=3, max_col=3):  # updated column number for 'Suggested Qname'
    #    for cell in row:
    #        if cell.value in previousround_suggested_qnames:
    #            cell.fill = green_fill

    


from openpyxl.styles import Font

def create_question_changes_sheet(
    wb,
    unique_country_questionnaire_qname,
    unique_template_questionnaire_qname,
    unique_country_questionnaire,
    unique_template_questionnaire
):
    # Create a new worksheet in the workbook
    sheet = wb.create_sheet("Question Changes")

    # Introductory note with clear explanation
    sheet["A1"] = (
        "This sheet highlights differences in question names (Q Name and Suggested Q Name) "
        "between the current country questionnaire and the previous round (if available). "
        "Use this to quickly identify new additions, removals, or edits."
    )
    sheet["A1"].font = Font(bold=True, italic=True)

    # Set column headers
    sheet["A3"] = "Q Name in Country Only"
    sheet["B3"] = "Q Name Missing in Country (Present in Previous)"
    sheet["D3"] = "Suggested Q Name in Country Only"
    sheet["E3"] = "Suggested Q Name Missing in Country (Present in Previous)"
    for col in ["A3", "B3", "D3", "E3"]:
        sheet[col].font = Font(bold=True)

    # Fill in column A with QNames added in country version
    for i, q in enumerate(unique_country_questionnaire_qname, start=4):
        sheet.cell(row=i, column=1).value = q

    # Fill in column B with QNames missing in country but present in previous/template
    for i, q in enumerate(unique_template_questionnaire_qname, start=4):
        sheet.cell(row=i, column=2).value = q

    # Fill in column D with Suggested QNames added in country version
    for i, q in enumerate(unique_country_questionnaire, start=4):
        sheet.cell(row=i, column=4).value = q

    # Fill in column E with Suggested QNames missing in country
    for i, q in enumerate(unique_template_questionnaire, start=4):
        sheet.cell(row=i, column=5).value = q

    # Adjust column widths
    for col in ['A', 'B', 'D', 'E']:
        max_len = max((len(str(cell.value)) for cell in sheet[col] if cell.value), default=20)
        sheet.column_dimensions[col].width = max_len + 5

    return wb

def create_question_changes_sheet_VOLD(wb, unique_country_questionnaire_qname, unique_template_questionnaire_qname, unique_country_questionnaire, unique_template_questionnaire):
    # Create a new sheet called "Question changes"
    new_sheet = wb.create_sheet("Question changes")

    # Modify the values for A1 and apply bold and italic formatting
    new_sheet["A1"] = "***In the case of a Previous Questionnaire validation, the 'Template Questionnaire' refers to the Previous round questionnaire***"
    font = Font(bold=True, italic=True)
    new_sheet["A1"].font = font

    # Paste the outputs of Message 7 and 8 into separate columns in "Question changes"
    new_sheet["A3"] = "QName - Country questionnaire:"
    new_sheet["B3"] = "QName - Template questionnaire only:"

    # Apply bold formatting to cells A3 and B3
    bold_font = Font(bold=True)
    new_sheet["A3"].font = bold_font
    new_sheet["B3"].font = bold_font
    new_sheet["D3"].font = bold_font
    new_sheet["E3"].font = bold_font

    # Paste the values vertically in column A (QName Field) starting from row 4
    for i, question in enumerate(unique_country_questionnaire_qname, start=4):
        new_sheet.cell(row=i, column=1).value = question

    # Paste the values vertically in column B (QName Field) starting from row 4
    for i, question in enumerate(unique_template_questionnaire_qname, start=4):
        new_sheet.cell(row=i, column=2).value = question

    # Paste the outputs of Message 5 and 6 into separate columns in "Question changes"
    new_sheet["D3"] = "SuggestedQName - Country questionnaire:"
    new_sheet["E3"] = "SuggestedQName - Template questionnaire:"

    # Paste the values vertically in column D (SuggestedQName Field) starting from row 4
    for i, question in enumerate(unique_country_questionnaire, start=4):
        new_sheet.cell(row=i, column=4).value = question

    # Paste the values vertically in column E (SuggestedQName Field) starting from row 4
    for i, question in enumerate(unique_template_questionnaire, start=4):
        new_sheet.cell(row=i, column=5).value = question

    # Auto format column width for cells A1, A3, B3, D3, and E3
    columns_to_format = ['A', 'B', 'D', 'E']

    # Determine the maximum length for each column
    max_lengths = [max(len(str(new_sheet[column + '3'].value)), len(column)) for column in columns_to_format]

    # Auto format column width for cells A1, A3, B3, D3, and E3
    for column in columns_to_format:
        adjusted_width = (max_lengths[columns_to_format.index(column)] + 2) * 1.2
        new_sheet.column_dimensions[column].width = adjusted_width

    return wb


def read_questionnaire(input_questionnaire_file, output_filename):
    ##this section of the script reads the survey excel file and creates an excel file with multiple sheets:
    ### each sheet contains the coded value and description for a "Single choice" or "Open Ended-Select All That Apply" question.
    ###moreover, it creates and populates several lists that will be used later for defining each field of the final table names, types and domains

    #print("Opening questionnaire DF")
    quest_df = pd.read_excel(open(input_questionnaire_file, 'rb'), sheet_name='survey',skiprows=2)

# Check if the survey includes other languages besides English and print an appropriate message
    if 'French' in quest_df:
        print("Version in French and English")
        languages = ["French","English"]
    elif 'Arabic' in quest_df:
        print("Version in Arabic and English")
        languages = ["Arabic","English"]
    elif 'Spanish' in quest_df:
        print("Version in Spanish and English")
        languages = ["Spanish","English"]
    elif 'Portuguese' in quest_df:
        print("Version in Portuguese and English")
        languages = ["Portuguese","English"]
        
    else:
        print("Version in English only")
        languages = ["English"]
        
# Iterate through languages to create Excel files with appropriate naming convention        
    for language in languages:
        filename = output_filename[:-5] + "_" +  language.lower()[:2] + ".xlsx"
        writer = pd.ExcelWriter(filename, engine='xlsxwriter')
        #create a list of all possible numbering
        numbering = ["%s)" % n for n in range(1,200)] ## 1), 2), ... 200)
        
        # Initialize containers to store the results
        dict_derived_fieldnames = {} #this dict will group all derived fields in case of "Select All That Apply" type of questions
        field_names_list = [] ##this list will contain all fields of the final table
        text_type_fields = [] ##this list will contain all fields of the final table with TEXT type
        range_type_fields = [] ##this list will contain all fields of the final table storing RANGE data  (will be LONG type)
        double_type_fields = [] ##this list will contain all fields of the final table storing DOUBLE data
        ##iterate the following for each row (so each question of the questionnaire)
        all_derived_fieldnames = []
        all_answers_with_other_option = [] ##this list will contain all "Other: specify" fields
        
         # Filters the dataframe to rows where the 'Suggested Qname' is not null
        quest_df = quest_df[quest_df['Suggested Qname'].notna()]

        for index, row in quest_df.iterrows():
            question_name = []
            question_type = []
            try:
                first_derived_fieldname = "" #the name of the first derived field will be the main of the domain table
                derived_fieldnames = []
                codes_and_labels = []
                categories = str(row[language]).replace("\t","")
                question_name = row['Suggested Qname'].strip()  #Q Name
                question_type = row['Q Type']
                programming_instructions = row['Programming Instructions'] #this field contains coded values for crop_main
                #print("\n\n----%s----" % question_name)
                #only for questions with pre-defined categories need domains
                if question_type in ("StartRecording","Single Choice","Open Ended-Single Choice", "Open Ended - Single Choice", "Open Ended-Select All That Apply",
                                     "Select All That Apply","Open Ended - Select All That Apply "):
                    if question_name == 'crp_main': #for this question only, coded values should be taken from field programming_instructions
                        programming_lines = programming_instructions.splitlines()
                        for programming_line in programming_lines:
                            if ")" in programming_line:
                                #print(programming_line)
                                index, category = programming_line.split(")")
                                category = fix_category_formatting(category)
                                codes_and_labels.append([index, category])
                    else:
                        #find all numbering present in the category string
                        numbering_in_text = [n for n in numbering if n in categories]
                        #print(numbering_in_text)
                        ##the following loop creates a list "codes_and_labels" with all available codes&labels for each question
                        for index in range(0,len(numbering_in_text)):
                            start = categories.find(numbering_in_text[index]) + len(numbering_in_text[index])
                            try:
                                end = categories.find(numbering_in_text[index + 1])
                                substring = categories[start:end].strip()
                            except:
                                # it fails during the last loop -> the last option is usually at the end of the string
                                substring = categories[start:].strip()
                            #print(substring)
                            category = fix_category_formatting(substring)
                            codes_and_labels.append([index +1, category])

                    if question_type not in ["Open Ended-Select All That Apply","Select All That Apply","Open Ended - Select All That Apply "]:
                        #so questions with NO derived fields
                        field_names_list.append(question_name.strip())
                        codes_and_labels_df = pd.DataFrame(codes_and_labels, columns=['code', 'label'])
                        codes_and_labels_df.to_excel(writer, sheet_name=question_name)
                    else:
                        #so questions with derived fields
                        numbering_in_qname = [n for n in numbering if n in question_name]
                        for index in range(0, len(numbering_in_qname)):
                            start = question_name.find(numbering_in_qname[index]) + len(numbering_in_qname[index])
                            try:
                                end = question_name.find(numbering_in_qname[index + 1])
                                derived_field_name = question_name[start:end].strip()
                            except:
                                # it fails during the last loop -> the last option is usually at the end of the string
                                derived_field_name = question_name[start:].strip()
                            all_derived_fieldnames.append(derived_field_name)
                            derived_fieldnames.append(derived_field_name)
                            field_names_list.append(derived_field_name)
                            if derived_field_name[-6:] == "_other": ##this field will need to be STRING - with no domain (since it's a 'other specify')
                                all_answers_with_other_option.append(derived_field_name)
                            if index == 0:
                                first_derived_fieldname = derived_field_name
                                codes_and_labels_df = pd.DataFrame(codes_and_labels, columns=['code', 'label'])
                                #codes_and_labels_df.to_excel(writer, sheet_name=derived_field_name) #we don't need domain table for derived fields, since they will use YES NOT domain table
                        dict_derived_fieldnames[first_derived_fieldname] = all_derived_fieldnames
                elif question_type == "Range":
                    #these questions will be associated to LONG type fields
                    field_names_list.append(question_name.strip())
                    range_type_fields.append(question_name)
                elif question_type == "Open Ended":
                    #these questions will be associated to TEXT type fields
                    if not pd.isnull(question_name): #NaN rows we want to skip (i.e. OptIn question without a name in the survey)
                        field_names_list.append(question_name)
                        text_type_fields.append(question_name)
                else:
                    #print("QUESTION SKIPPED ---------", question_name, question_type)
                    pass
            except Exception as e:
                print("Failed question %s (type: %s) for error: %s " % (question_name, question_type,e))

        #adding Yes No table (for derived fields domain)
        d = {1:"Yes",0:"No"}
        yesno_df = pd.DataFrame(d.items(), columns=['code', 'label'])
        yesno_df.to_excel(writer, sheet_name='yes_no')
        #print("Saving codes and labels %s" % coded_values_file)

        #print(all_answers_with_other_option)

        list_of_yes_no_fields = []
        for derived_field in all_derived_fieldnames:
            if derived_field not in all_answers_with_other_option:
                list_of_yes_no_fields.append(derived_field)

        derived_fields_df = pd.DataFrame(list_of_yes_no_fields)
        derived_fields_df.to_excel(writer, sheet_name='derived_fields')
        #writer.save()
    return languages

def detect_enumerator(geopoll_or_kobo_template):
    enumerator = ""
    if "kobo" in geopoll_or_kobo_template.lower():
        country_df.columns = country_df.columns.str.replace("[/]", "")
        enumerator = "kobo"
    elif "geopoll" in geopoll_or_kobo_template.lower():
        enumerator = "geopoll"
    else:
        print("Please include Geopoll or Kobo inside the filename: %s " % geopoll_or_kobo_template)
        enumerator = "Please include Geopoll or Kobo inside the filename"
    return enumerator

def insert_sheet_with_adm2_reference_old(questionnaire_file,adm0_iso3):
    # Connect to the ArcGIS online service
    gis = GIS("https://hqfao.maps.arcgis.com")
    # Retrieve the layer using its item ID
    item = gis.content.get('3596c3ad318849068eda21517ade30be')
    flayer = item.layers[0]
    # Query the layer to obtain the administrative level 2 data for the given ISO3 country code
    query = "adm0_ISO3 = '" + adm0_iso3 +"'"
    sdf = flayer.query(where=query).sdf
    
    # Remove unnecessary columns
    del sdf['OBJECTID']
    #del sdf['validity']
    del sdf['Shape__Area']
    del sdf['Shape__Length']
    del sdf['SHAPE']
    #print(sdf.head())
    excel_book = openpyxl.load_workbook(questionnaire_file)
    with pd.ExcelWriter(questionnaire_file, engine='openpyxl') as writer:
        writer.book = excel_book
        sdf.to_excel(writer, 'ADMIN info', index=False)
        writer.save()

    return sdf


def insert_sheet_with_adm_reference(questionnaire_file,admin_level,adm0_iso3):
    if admin_level == "Admin 0":
        admin_url = "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/Administrative_Boundaries_Reference_(view_layer)/FeatureServer/2/query?where=adm0_ISO3%20%3D%20'" + adm0_iso3 + "'&outFields=adm0_name,adm0_name_local,adm0_ISO3_2d,adm0_ISO3&returnGeometry=false&outSR=4326&f=json"
    elif admin_level == "Admin 1":
        admin_url = "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/Administrative_Boundaries_Reference_(view_layer)/FeatureServer/1/query?where=adm0_ISO3%20%3D%20'" + adm0_iso3 + "'&outFields=adm1_name,adm1_name_local,adm1_pcode,adm0_name,adm0_ISO3&returnGeometry=false&outSR=4326&f=json"
    elif admin_level == "Admin 2":
        admin_url = "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/Administrative_Boundaries_Reference_(view_layer)/FeatureServer/0/query?where=adm0_ISO3%20%3D%20'" + adm0_iso3 + "'&outFields=adm2_name,adm2_name_local,adm2_pcode,adm1_name,adm1_pcode,adm0_name,adm0_ISO3&returnGeometry=false&outSR=4326&f=json"
    elif admin_level == "Admin 3":
        admin_url ="https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/Reference_Admin_3/FeatureServer/0/query?where=adm0_ISO3%20%3D%20'"+ adm0_iso3 + "'&outFields=adm3_name,adm3_name_local,adm3_pcode,adm2_name,adm2_pcode,adm1_name,adm1_pcode,adm0_name,adm0_ISO3&returnGeometry=false&outSR=4326&f=json"

    with urllib.request.urlopen(admin_url) as admin_url:
        data = json.loads(admin_url.read().decode())
        #print(type(data))

    df = pd.json_normalize(data['features'])
    #print (df.head())

    if admin_level == "Admin 0":
        df.rename(columns={'attributes.adm0_name': 'adm0_name',
                           'attributes.adm0_name_local': 'adm0_name_local',
                           'attributes.adm0_ISO3_2d': 'adm0_ISO3_2d',
                           'attributes.adm0_ISO3': 'adm0_ISO3'}, inplace=True)
    elif admin_level == "Admin 1":
        df.rename(columns={'attributes.adm0_name': 'adm0_name',
                           'attributes.adm0_ISO3': 'adm0_ISO3',
                           'attributes.adm1_name': 'adm1_name',
                           'attributes.adm1_name_local': 'adm1_name_local',
                           'attributes.adm1_pcode': 'adm1_pcode'}, inplace=True)

    elif admin_level == "Admin 2":
        df.rename(columns={'attributes.adm0_name': 'adm0_name',
                           'attributes.adm0_ISO3': 'adm0_ISO3',
                           'attributes.adm1_name': 'adm1_name',
                           'attributes.adm1_pcode': 'adm1_pcode',
                           'attributes.adm2_name': 'adm2_name',
                           'attributes.adm2_name_local': 'adm2_name_local',
                           'attributes.adm2_pcode': 'adm2_pcode'}, inplace=True)

    elif admin_level == "Admin 3":
        df.rename(columns={'attributes.adm0_name': 'adm0_name',
                           'attributes.adm0_ISO3': 'adm0_ISO3',
                           'attributes.adm1_name': 'adm1_name',
                           'attributes.adm1_pcode': 'adm1_pcode',
                           'attributes.adm2_name': 'adm2_name',
                           'attributes.adm2_pcode': 'adm2_pcode',
                           'attributes.adm3_name': 'adm3_name',
                           'attributes.adm3_name_local': 'adm3_name_local',
                           'attributes.adm3_pcode': 'adm3_pcode'}, inplace=True)


    excel_book = openpyxl.load_workbook(questionnaire_file)
    with pd.ExcelWriter(questionnaire_file, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name=admin_level + ' info', index=False)
        #writer.save()
    return df


from openpyxl.utils import get_column_letter

def update_questionnaire(template_file, country_file):
    keys = ['$ADMIN1$', '$ADMIN2$', '$reference year$', '$season$', '$season phase $',
            '$local measurement unit$', '$currency$', '$MIN AMOUNT$', '$MAX AMOUNT$',
            '$THRESHOLD$', '$local vegetables$', '$local fruits$', '$expected or nothing$']
    
    # Load the questionnaire template from the specified file, skipping the first 3 rows.
    template_data = pd.read_excel(template_file,'survey', header=2)
    # Extract only the 'Q Name' and 'English' columns and drop any rows with missing values.
    selected_data = template_data[['Q Name', 'English']]
    selected_data=selected_data.dropna()
    
    # Filter rows where the 'English' column contains any of the specified keys.
    mask = selected_data['English'].apply(lambda x: any(key in str(x) for key in keys))
    selected_data = selected_data[mask]
    #print(selected_data["Q Name"])
    
    # Load the country questionnaire with openpyxl
    workbook = openpyxl.load_workbook(country_file)
    workbook.sheetnames
    worksheet = workbook["survey"]
    
    # Determine the number of rows in the worksheet.
    number_of_rows = worksheet.max_row 
    
    # Iterate over each row in the worksheet.
    for k in range(number_of_rows):
        # Extract the question value from the second column of the current row.
        questionValue = str(worksheet[get_column_letter(2)+str(k+1)].value)
        
        # Compare the question value with the 'Q Name' column of the selected_data DataFrame.
        for index, row in selected_data.iterrows():
            if questionValue == row['Q Name']:

                # If there's a match, update the value in the fourth column with the 'English' value from the template.
                worksheet[get_column_letter(4)+str(k+1)] = str(row['English'])
    workbook.save(country_file)




def find_and_replace_strings_in_df(questionnaire_file):
    language = detect_language(questionnaire_file)
    replacing_table_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Additional information', skiprows=1)

    workbook = openpyxl.load_workbook(questionnaire_file)
    workbook.sheetnames
    worksheet = workbook["survey"]
    number_of_rows = worksheet.max_row #Number of Rows
    number_of_columns = worksheet.max_column #Number of Columns

    if language == "en":
        replacing_table_df = replacing_table_df[['Original','Replacement']]
        replacement_list = list(replacing_table_df['Original'].values.tolist())
        #print(replacement_list)

        # Determine the 'expected_or_nothing' value based on the 'Replacement' value at index 4
        expected_or_nothing = ""
        if replacing_table_df.iloc[4]['Replacement'] == "Not yet in season" or replacing_table_df.iloc[4]['Replacement'] == "Land preparation" or replacing_table_df.iloc[4]['Replacement'] == "Planting" or replacing_table_df.iloc[4]['Replacement'] == "Early growing" or replacing_table_df.iloc[4]['Replacement'] == "Growing" or replacing_table_df.iloc[4]['Replacement'] == "Maturing":
            expected_or_nothing = "expected"
        elif replacing_table_df.iloc[4]['Replacement'] == "Harvesting" or replacing_table_df.iloc[4]['Replacement'] == "Recently finished":
            expected_or_nothing = ""
        else:
            expected_or_nothing = ""

        replacementTextKeyPairs = {'$ADMIN1$': replacing_table_df.iloc[0]['Replacement'],
                                   '$ADMIN2$': replacing_table_df.iloc[1]['Replacement'],
                                   '$reference year$': replacing_table_df.iloc[2]['Replacement'],
                                   '$season$': replacing_table_df.iloc[3]['Replacement'],
                                   '$season phase $': replacing_table_df.iloc[4]['Replacement'],
                                   '$local measurement unit$': replacing_table_df.iloc[5]['Replacement'],
                                   '$currency$': replacing_table_df.iloc[7]['Replacement'],
                                   '$MIN AMOUNT$': replacing_table_df.iloc[8]['Replacement'],
                                   '$MAX AMOUNT$': replacing_table_df.iloc[9]['Replacement'],
                                   '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement'],
                                   '$local vegetables$': replacing_table_df.iloc[11]['Replacement'],
                                   '$local fruits$': replacing_table_df.iloc[12]['Replacement'],
                                   '$expected or nothing$': expected_or_nothing
                              }

        

    elif language == "fr":
        
        print("FRENCH")

        replacing_table_df = replacing_table_df[['Original','Replacement (EN)','Replacement (FR)']]
        replacement_list = list(replacing_table_df['Original'].values.tolist())
        
        #print(replacing_table_df)
        #print(replacement_list)
        
        
        ##Translating values
        
        #Translating Reference Year
        reference_year_EN = ""
        if replacing_table_df.iloc[2]['Replacement (FR)'] == "l'année dernière":
            reference_year_EN = "last year"
        elif replacing_table_df.iloc[2]['Replacement (FR)'] == "une année normale":
            reference_year_EN = "a normal year"
        else:
            reference_year_EN = replacing_table_df.iloc[2]['Replacement (FR)']
        

        #Translating season phase
        season_phase_EN = ""
        if replacing_table_df.iloc[4]['Replacement (FR)'] == "Pas encore en saison":
            season_phase_EN = "Not yet in season"
        elif replacing_table_df.iloc[4]['Replacement (FR)'] == "Préparation du terrain":
            season_phase_EN = "Land preparation"
        elif replacing_table_df.iloc[4]['Replacement (FR)'] == "Semis / plantation":
            season_phase_EN = "Planting"            
        elif replacing_table_df.iloc[4]['Replacement (FR)'] == "Début de la croissance":
            season_phase_EN = "Early growing"
        elif replacing_table_df.iloc[4]['Replacement (FR)'] == "Croissance":
            season_phase_EN = "Growing"
        elif replacing_table_df.iloc[4]['Replacement (FR)'] == "Proche de la récolte":
            season_phase_EN = "Maturing"
        elif replacing_table_df.iloc[4]['Replacement (FR)'] == "Récolte":
            season_phase_EN = "Harvesting"
        elif replacing_table_df.iloc[4]['Replacement (FR)'] == "Post-récolte":
            season_phase_EN = "Recently finished"
        else:
            season_phase_EN = replacing_table_df.iloc[4]['Replacement (FR)']
            
            
        expected_or_nothing = ""
        attendue_ou_rien = ""
        if replacing_table_df.iloc[4]['Replacement (EN)'] == "Not yet in season" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Land preparation" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Planting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Early growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Maturing":
            expected_or_nothing = "expected"
            attendue_ou_rien = "attendue"
        elif replacing_table_df.iloc[4]['Replacement (EN)'] == "Harvesting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Recently finished":
            expected_or_nothing = ""
            attendue_ou_rien = ""
        else:
            expected_or_nothing = ""
            attendue_ou_rien = ""


        replacementTextKeyPairs = {'$ADMIN1$': replacing_table_df.iloc[0]['Replacement (FR)'],
                                   '$ADMIN2$': replacing_table_df.iloc[1]['Replacement (FR)'],
                                   '$reference year$': reference_year_EN,
                                   '$reference year FR$': replacing_table_df.iloc[2]['Replacement (FR)'],
                                   '$season$': replacing_table_df.iloc[3]['Replacement (EN)'],
                                   '$saison$': replacing_table_df.iloc[3]['Replacement (FR)'],
                                   '$season phase $': replacing_table_df.iloc[4]['Replacement (FR)'],
                                   '$local measurement unit$': replacing_table_df.iloc[5]['Replacement (EN)'],
                                   '$unité de mesure locale$': replacing_table_df.iloc[5]['Replacement (FR)'],
                                   '$currency$': replacing_table_df.iloc[7]['Replacement (FR)'],
                                   '$MIN AMOUNT$': replacing_table_df.iloc[8]['Replacement (FR)'],
                                   '$MAX AMOUNT$': replacing_table_df.iloc[9]['Replacement (FR)'],
                                   '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement (FR)'],
                                   '$local vegetables$': replacing_table_df.iloc[11]['Replacement (FR)'],
                                   '$local fruits$': replacing_table_df.iloc[12]['Replacement (FR)'],
                                   '$expected or nothing$': expected_or_nothing,
                                   '$attendue ou rien$': attendue_ou_rien
                                    }
        


    elif language == "ar":

        expected_or_nothing = ""
        if replacing_table_df.iloc[4]['Replacement (EN)'] == "Not yet in season" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Land preparation" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Planting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Early growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Maturing":
            expected_or_nothing = "expected"
        elif replacing_table_df.iloc[4]['Replacement (EN)'] == "Harvesting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Recently finished":
            expected_or_nothing = ""
        else:
            expected_or_nothing = ""

        attendue_ou_rien = ""
        if replacing_table_df.iloc[4]['Replacement (AR)'] == "ليس في الموسم بعد" or replacing_table_df.iloc[4]['Replacement (AR)'] == "إعداد الأرض" or replacing_table_df.iloc[4]['Replacement (AR)'] == "الزرع" or replacing_table_df.iloc[4]['Replacement (AR)'] == "النمو المبكر" or replacing_table_df.iloc[4]['Replacement (AR)'] == "النمو" or replacing_table_df.iloc[4]['Replacement (AR)'] == "النضوج":
            attendue_ou_rien = "مُتوقع"
        elif replacing_table_df.iloc[3]['Replacement (AR)'] == "الحصاد" or replacing_table_df.iloc[3]['Replacement (AR)'] == "انتهى مؤخرا":
            attendue_ou_rien = ""
        else:
            attendue_ou_rien = ""

        replacementTextKeyPairs = {'$ADMIN1 AR$': replacing_table_df.iloc[0]['Replacement (AR)'],
                                   '$ADMIN2 AR$': replacing_table_df.iloc[1]['Replacement (AR)'],
                                   '$reference year AR$': replacing_table_df.iloc[2]['Replacement (AR)'],
                                   '$season AR$': replacing_table_df.iloc[3]['Replacement (AR)'],
                                   '$season phase AR$': replacing_table_df.iloc[4]['Replacement (AR)'],
                                   '$local measurement unit AR$': replacing_table_df.iloc[5]['Replacement (AR)'],
                                   '$currency AR$': replacing_table_df.iloc[7]['Replacement (AR)'],
                                   '$MIN AMOUNT AR$': replacing_table_df.iloc[8]['Replacement (AR)'],
                                   '$MAX AMOUNT AR$': replacing_table_df.iloc[9]['Replacement (AR)'],
                                   '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement (AR)'],
                                   '$local vegetables AR$': replacing_table_df.iloc[11]['Replacement (AR)'],
                                   '$local fruits AR$': replacing_table_df.iloc[12]['Replacement (AR)'],
                                   '$expected or nothing AR$': expected_or_nothing,

                                   '$ADMIN1$': replacing_table_df.iloc[0]['Replacement (EN)'],
                                   '$ADMIN2$': replacing_table_df.iloc[1]['Replacement (EN)'],
                                   '$reference year$': replacing_table_df.iloc[2]['Replacement (EN)'],
                                   '$season$': replacing_table_df.iloc[3]['Replacement (EN)'],
                                   '$season phase$': replacing_table_df.iloc[4]['Replacement (EN)'],
                                   '$local measurement unit$': replacing_table_df.iloc[5]['Replacement (EN)'],
                                   '$currency$': replacing_table_df.iloc[7]['Replacement (EN)'],
                                   '$MIN AMOUNT$': replacing_table_df.iloc[8]['Replacement (EN)'],
                                   '$MAX AMOUNT$': replacing_table_df.iloc[9]['Replacement (EN)'],
                                   '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement (EN)'],
                                   '$local vegetables$': replacing_table_df.iloc[11]['Replacement (EN)'],
                                   '$local fruits$': replacing_table_df.iloc[12]['Replacement (EN)'],
                                   '$expected or nothing$': expected_or_nothing
                              }

    elif language == "pt":

        expected_or_nothing = ""
        if replacing_table_df.iloc[4]['Replacement (EN)'] == "Not yet in season" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Land preparation" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Planting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Early growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Maturing":
            expected_or_nothing = "expected"
            attendue_ou_rien = "مُتوقع"
        elif replacing_table_df.iloc[4]['Replacement (EN)'] == "Harvesting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Recently finished":
            expected_or_nothing = ""
            attendue_ou_rien = ""
        else:
            expected_or_nothing = ""
            attendue_ou_rien = ""

        replacementTextKeyPairs = {'$ADMIN1 PT$': replacing_table_df.iloc[0]['Replacement (PT)'],
                                   '$ADMIN2 PT$': replacing_table_df.iloc[1]['Replacement (PT)'],
                                   '$reference year PT$': replacing_table_df.iloc[2]['Replacement (PT)'],
                                   '$season PT$': replacing_table_df.iloc[3]['Replacement (PT)'],
                                   '$season phase PT$': replacing_table_df.iloc[4]['Replacement (PT)'],
                                   '$local measurement unit PT$': replacing_table_df.iloc[5]['Replacement (PT)'],
                                   '$currency PT$': replacing_table_df.iloc[7]['Replacement (PT)'],
                                   '$MIN AMOUNT PT$': replacing_table_df.iloc[8]['Replacement (PT)'],
                                   '$MAX AMOUNT PT$': replacing_table_df.iloc[9]['Replacement (PT)'],
                                   '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement (PT)'],
                                   '$local vegetables PT$': replacing_table_df.iloc[11]['Replacement (PT)'],
                                   '$local fruits PT$': replacing_table_df.iloc[12]['Replacement (PT)'],
                                   '$expected or nothing PT$': expected_or_nothing,

                                   '$ADMIN1$': replacing_table_df.iloc[0]['Replacement (EN)'],
                                   '$ADMIN2$': replacing_table_df.iloc[1]['Replacement (EN)'],
                                   '$reference year$': replacing_table_df.iloc[2]['Replacement (EN)'],
                                   '$season$': replacing_table_df.iloc[3]['Replacement (EN)'],
                                   '$season phase$': replacing_table_df.iloc[4]['Replacement (EN)'],
                                   '$local measurement unit$': replacing_table_df.iloc[5]['Replacement (EN)'],
                                   '$currency$': replacing_table_df.iloc[7]['Replacement (EN)'],
                                   '$MIN AMOUNT$': replacing_table_df.iloc[8]['Replacement (EN)'],
                                   '$MAX AMOUNT$': replacing_table_df.iloc[9]['Replacement (EN)'],
                                   '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement (EN)'],
                                   '$local vegetables$': replacing_table_df.iloc[11]['Replacement (EN)'],
                                   '$local fruits$': replacing_table_df.iloc[12]['Replacement (EN)'],
                                   '$expected or nothing$': expected_or_nothing
                              }            
    elif language == "es":
        
        expected_or_nothing = ""
        if replacing_table_df.iloc[4]['Replacement (EN)'] == "Not yet in season" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Land preparation" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Planting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Early growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Growing" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Maturing":
            expected_or_nothing = "expected"
            attendue_ou_rien = "esperado"
        
        elif replacing_table_df.iloc[4]['Replacement (EN)'] == "Harvesting" or replacing_table_df.iloc[4]['Replacement (EN)'] == "Recently finished":
            expected_or_nothing = ""
            attendue_ou_rien = ""
        
        else:
            expected_or_nothing = ""
            attendue_ou_rien = ""

            replacementTextKeyPairs = {'$ADMIN1 ES$': replacing_table_df.iloc[0]['Replacement (ES)'],
                                       '$ADMIN2 ES$': replacing_table_df.iloc[1]['Replacement (ES)'],
                                       '$reference year ES$': replacing_table_df.iloc[2]['Replacement (ES)'],
                                       '$season ES$': replacing_table_df.iloc[3]['Replacement (ES)'],
                                       '$season phase ES$': replacing_table_df.iloc[4]['Replacement (ES)'],
                                       '$local measurement unit ES$': replacing_table_df.iloc[5]['Replacement (ES)'],
                                       '$currency ES$': replacing_table_df.iloc[7]['Replacement (ES)'],
                                       '$MIN AMOUNT ES$': replacing_table_df.iloc[8]['Replacement (ES)'],
                                       '$MAX AMOUNT ES$': replacing_table_df.iloc[9]['Replacement (ES)'],
                                       '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement (ES)'],
                                       '$local vegetables ES$': replacing_table_df.iloc[11]['Replacement (ES)'],
                                       '$local fruits ES$': replacing_table_df.iloc[12]['Replacement (ES)'],
                                       '$expected or nothing ES$': attendue_ou_rien,
            
                                       '$ADMIN1$': replacing_table_df.iloc[0]['Replacement (EN)'],
                                       '$ADMIN2$': replacing_table_df.iloc[1]['Replacement (EN)'],
                                       '$reference year$': replacing_table_df.iloc[2]['Replacement (EN)'],
                                       '$season$': replacing_table_df.iloc[3]['Replacement (EN)'],
                                       '$season phase$': replacing_table_df.iloc[4]['Replacement (EN)'],
                                       '$local measurement unit$': replacing_table_df.iloc[5]['Replacement (EN)'],
                                       '$currency$': replacing_table_df.iloc[7]['Replacement (EN)'],
                                       '$MIN AMOUNT$': replacing_table_df.iloc[8]['Replacement (EN)'],
                                       '$MAX AMOUNT$': replacing_table_df.iloc[9]['Replacement (EN)'],
                                       '$THRESHOLD$': replacing_table_df.iloc[10]['Replacement (EN)'],
                                       '$local vegetables$': replacing_table_df.iloc[11]['Replacement (EN)'],
                                       '$local fruits$': replacing_table_df.iloc[12]['Replacement (EN)'],
                                       '$expected or nothing$': expected_or_nothing
                                    }   

    # Iterate over each column and row in the worksheet to find and replace text based on the dictionary
    for i in range(number_of_columns):
        for k in range(number_of_rows):
            # Get the value of the cell at (i+1, k+1)
            cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
            #print(cellValue)

            
            # Iterate through each key in the replacement dictionary
            for key in replacementTextKeyPairs.keys():
                # If the key is found in the cell value and the cell value is not None
                if key in str(cellValue) and str(cellValue) != None:
                    # Replace the key with its corresponding value in the dictionary
                    newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                    #newCellValue = replacementTextKeyPairs.get(key)
                    # Update the cell with the new value
                    worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                    # Update cellValue for any subsequent replacements within the same cell
                    cellValue = newCellValue
    # Hide the "Additional information" sheet
    workbook["Additional information"].sheet_state = 'hidden'
    # Save the changes to the workbook
    workbook.save(questionnaire_file)

    

def sort_crop_list_by_selection(questionnaire_file):
    crop_list_df = pd.read_excel(pd.ExcelFile(questionnaire_file), 'Crop list', skiprows=2)

    columns = list(crop_list_df.columns.values.tolist())
    #print(columns)

    if 'Label (FR)' not in columns and 'Label (AR)' not in columns and 'Label (PT)' not in columns and 'Label (ES)' not in columns:

        crop_list_df = crop_list_df[['Select top 10 crops ','GeoPoll code','Dataset code','Label (EN)']]
        q_name = "crp_main"
        q_label = "What has been the main crop that your household has grown for food and income generation in $season$, if any? \n [OPERATOR: SINGLE RESPONSE. \"NO CROP PRODUCTION\" IS IN LAST OPTIONS AT THE END.] \n \n"

        q_name_CROPS_SOLD = "crp_salesmain"
        q_label_CROPS_SOLD = "What has been the main crop that you have sold over the last 3 months?\n[OPERATOR: \"No crop sold in the last 3 months\" APPEARS AT THE END OF THE LIST OF OPTIONS. SELECT ALL CROPS LISTED BY THE RESPONDENT.]\n\n"

        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["survey"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns

        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)

        sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','Dataset code'])
        #print(sorted_crop_list_df)


        sorted_crop_list_df = sorted_crop_list_df.reset_index()
        sorted_crop_list_df['GeoPoll code'] = sorted_crop_list_df.index + 1
        #print(sorted_crop_list_df['GeoPoll code'])
        
        

        sorted_crop_list_df['combined'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (EN)']
        sorted_crop_list = sorted_crop_list_df[['combined']].values.tolist()
        additional_crop_list = [["91)No crop production"], ["92)DON'T KNOW"], ["93)REFUSED"]]
        sorted_crop_list_additional = sorted_crop_list + additional_crop_list
        #print(sorted_crop_list)
        final_sorted_crop_list = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional]))
        
        sorted_crop_list_df['combined_DATASETCODES'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Dataset code'].astype(str)
        sorted_crop_list_DATASETCODES = sorted_crop_list_df[['combined_DATASETCODES']].values.tolist()
        #print(sorted_crop_list_DATASETCODES)
        additional_crop_list_DATASETCODES = [["91)777"], ["92)888"], ["93)999"]]
        sorted_crop_list_additional_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_DATASETCODES
        final_sorted_crop_list_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_DATASETCODES]))
        
        
        
        #print(final_sorted_crop_list)
        #print(final_sorted_crop_list_DATASETCODES)
        #print(crop_list_df)


        #CROPS SOLD
        sorted_crop_list_CROPS_SOLD = list(sorted_crop_list)
        additional_CROPS_SOLD = [["92)DON'T KNOW"], ["93)REFUSED"]]
        sorted_crop_list_CROPS_SOLD.extend(additional_CROPS_SOLD)
        final_sorted_crop_list_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_CROPS_SOLD]))
        #print(sorted_crop_list_CROPS_SOLD)
        #print(final_sorted_crop_list_CROPS_SOLD)
        
        additional_crop_list_CROPS_SOLD_DATASETCODES = [["92)888"], ["93)999"]]
        sorted_crop_list_additional_CROPS_SOLD_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_CROPS_SOLD_DATASETCODES
        final_sorted_crop_list__CROPS_SOLD_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_CROPS_SOLD_DATASETCODES]))
        
        


        # Iterate over the columns and rows, search
        # for the text and replace

        for k in range(number_of_rows):

            questionValue = str(worksheet[get_column_letter(2)+str(k+1)].value)
            labelValue = str(worksheet[get_column_letter(4)+str(k+1)].value)
                #print(cellValue)

            if q_name == str(questionValue) and str(questionValue) != None:
                  newLabelValue = q_label + final_sorted_crop_list
                  #newCellValue = replacementTextKeyPairs.get(key)
                  worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue)

            if q_name_CROPS_SOLD in str(questionValue) and str(questionValue) != None:
                  newLabelValue = q_label_CROPS_SOLD + final_sorted_crop_list_CROPS_SOLD
                  #newCellValue = replacementTextKeyPairs.get(key)
                  worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue)
                    
    
    
    
        replacementTextKeyPairs = {'$CROP CODES$': final_sorted_crop_list_DATASETCODES,
                                  '$CROP SOLD CODES$' : final_sorted_crop_list__CROPS_SOLD_DATASETCODES}
        for i in range(number_of_columns):
            for k in range(number_of_rows):

                cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue

        workbook["Crop list"].sheet_state = 'hidden'            
        workbook.save(questionnaire_file)
        return n_of_choices
    
    
    elif 'Label (FR)' in columns:
        crop_list_df = crop_list_df[['Select top 10 crops ','GeoPoll code','Dataset code','Label (EN)','Label (FR)']]
        q_name = "crp_main"
        q_label_en = "What has been the main crop that your household has grown for food and income generation in $season$, if any? \n [OPERATOR: SINGLE RESPONSE. \"NO CROP PRODUCTION\" IS IN LAST OPTIONS AT THE END.] \n \n"
        q_label_fr = "Quelle a été la principale culture pratiquée par votre ménage pour se nourrir et générer des revenus pendant la $saison$, si applicable? \n [OPÉRATEUR/OPÉRATRICE: RÉPONSE UNIQUE. «AUCUNE PRODUCTION VÉGÉTALE» EST PARMI LES DERNIÈRES OPTIONS, À LA FIN DE LA LISTE.] \n \n"


        q_name_CROPS_SOLD = "crp_salesmain"
        q_label_en_CROPS_SOLD = "What has been the main crop that you have sold over the last 3 months?\n[OPERATOR: \"No crop sold in the last 3 months\" APPEARS AT THE END OF THE LIST OF OPTIONS. SELECT ALL CROPS LISTED BY THE RESPONDENT.]\n\n"
        q_label_fr_CROPS_SOLD = "Quelle a été la principale récolte que vous avez vendue au cours des 3 derniers mois \n[OPERATEUR /OPÉRATRICE: \"Pas de culture vendue sur les 3 derniers mois\" FIGURE DANS LES DERNIÈRES OPTIONS DE RÉPONSE. SÉLECTIONNER TOUTES LES CULTURES LISTÉES PAR LE RÉPONDANT.]\n \n"


        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["survey"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns



        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)



        sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','GeoPoll code'])
        #print(sorted_crop_list_df)
        
        
        ##########NEWWWWWW########
        sorted_crop_list_df = sorted_crop_list_df.reset_index()
        sorted_crop_list_df['GeoPoll code'] = sorted_crop_list_df.index + 1
        #print(sorted_crop_list_df['GeoPoll code'])
        ##########NEWWWWWW########
        
        
        sorted_crop_list_df['combined (EN)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (EN)']
        sorted_crop_list_df['combined (FR)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (FR)']


        sorted_crop_list_en = sorted_crop_list_df[['combined (EN)']].values.tolist()
        additional_crop_list_en = [["91)No crop production"], ["92)DON'T KNOW"], ["93)REFUSED"]]
        sorted_crop_list_additional_en = sorted_crop_list_en + additional_crop_list_en 
        
        
        sorted_crop_list_fr = sorted_crop_list_df[['combined (FR)']].values.tolist()
        additional_crop_list_fr = [["91)Aucune production végétale"], ["92)NE SAIS PAS"], ["93)REFUSÉ"]]
        sorted_crop_list_additional_fr = sorted_crop_list_fr + additional_crop_list_fr
        
        #print(sorted_crop_list)
        final_sorted_crop_list_en = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en]))
        final_sorted_crop_list_fr = '\n'.join(map(str, [i[0] for i in sorted_crop_list_fr]))

        #print(final_sorted_crop_list)
        #print(crop_list_df)
        
        sorted_crop_list_df['combined_DATASETCODES'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Dataset code'].astype(str)
        sorted_crop_list_DATASETCODES = sorted_crop_list_df[['combined_DATASETCODES']].values.tolist()
        #print(sorted_crop_list_DATASETCODES)
        additional_crop_list_DATASETCODES = [["91)777"], ["92)888"], ["93)999"]]
        sorted_crop_list_additional_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_DATASETCODES
        final_sorted_crop_list_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_DATASETCODES]))


        #CROPS SOLD
        sorted_crop_list_en_CROPS_SOLD = list(sorted_crop_list_en)
        additional_en_CROPS_SOLD = [["92)DON'T KNOW"], ["93)REFUSED"]]
        sorted_crop_list_en_CROPS_SOLD.extend(additional_en_CROPS_SOLD)
        final_sorted_crop_list_en_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en_CROPS_SOLD]))
        #print(sorted_crop_list_en_CROPS_SOLD)
        #print(final_sorted_crop_list_en_CROPS_SOLD)

        sorted_crop_list_fr_CROPS_SOLD = list(sorted_crop_list_fr)
        additional_fr_CROPS_SOLD = [["92)NE SAIT PAS"], ["93)REFUSÉ"]]
        sorted_crop_list_fr_CROPS_SOLD.extend(additional_fr_CROPS_SOLD)
        final_sorted_crop_list_fr_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_fr_CROPS_SOLD]))
        #print(sorted_crop_list_fr_CROPS_SOLD)
        #print(final_sorted_crop_list_fr_CROPS_SOLD)

        additional_crop_list_CROPS_SOLD_DATASETCODES = [["91)888"], ["92)999"]]
        sorted_crop_list_additional_CROPS_SOLD_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_CROPS_SOLD_DATASETCODES
        final_sorted_crop_list__CROPS_SOLD_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_CROPS_SOLD_DATASETCODES]))


        # Iterate over the columns and rows, search
        # for the text and replace

        for k in range(number_of_rows):

            questionValue = str(worksheet[get_column_letter(2)+str(k+1)].value)
            labelValue = str(worksheet[get_column_letter(4)+str(k+1)].value)
                #print(cellValue)

            if q_name == str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en + final_sorted_crop_list_en
                newLabelValue_fr = q_label_fr + final_sorted_crop_list_fr
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_fr)

            if q_name_CROPS_SOLD in str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en_CROPS_SOLD + final_sorted_crop_list_en_CROPS_SOLD
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)

                newLabelValue_fr = q_label_fr_CROPS_SOLD + final_sorted_crop_list_fr_CROPS_SOLD
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_fr)

        replacementTextKeyPairs = {'$CROP CODES$': final_sorted_crop_list_DATASETCODES,
                                  '$CROP SOLD CODES$' : final_sorted_crop_list__CROPS_SOLD_DATASETCODES}
        for i in range(number_of_columns):
            for k in range(number_of_rows):

                cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue

        workbook["Crop list"].sheet_state = 'hidden'
        workbook.save(questionnaire_file)
        return n_of_choices
    
    

    elif 'Label (AR)' in columns:
        crop_list_df = crop_list_df[['Select top 10 crops ','GeoPoll code','Dataset code','Label (EN)','Label (AR)']]
        q_name = "crp_main"
        q_label_en = "What has been the main crop that your household has grown for food and income generation in $season$, if any? \n [OPERATOR: SINGLE RESPONSE. \"NO CROP PRODUCTION\" IS IN LAST OPTIONS AT THE END.] \n \n"
        q_label_ar = "ما هو المحصول الرئيسي الذي تزرعه أسرتك من أجل الغذاء وتوليد الدخل في موسم $ Season $ ، إن وجد؟ \n [المشغل: استجابة واحدة. 'لا إنتاج للمحصول' في آخر الخيارات في النهاية.] \n \n"


        q_name_CROPS_SOLD = "crp_salesmain"
        q_label_en_CROPS_SOLD = "What has been the main crop that you have sold over the last 3 months?\n[OPERATOR: SELECT ALL CROPS LISTED BY THE RESPONDENT.]\n\n"
        q_label_ar_CROPS_SOLD = "ما هو المحصول الرئيسي الذي بعته خلال الأشهر الثلاثة الماضية؟ \n [حدد جميع المحاصيل التي سجلها المجيب.] \n \n"


        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["survey"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns



        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)



        sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','GeoPoll code'])
        #print(sorted_crop_list_df)
        
        
        
        ##########NEWWWWWW########
        sorted_crop_list_df = sorted_crop_list_df.reset_index()
        sorted_crop_list_df['GeoPoll code'] = sorted_crop_list_df.index + 1
        print(sorted_crop_list_df['GeoPoll code'])
        ##########NEWWWWWW########
        
        
        sorted_crop_list_df['combined (EN)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (EN)']
        sorted_crop_list_df['combined (AR)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (AR)']

        
        sorted_crop_list_en = sorted_crop_list_df[['combined (EN)']].values.tolist()
        additional_crop_list_en = [["91)No crop production"], ["92)DON'T KNOW"], ["93)REFUSED"]]
        sorted_crop_list_additional_en = sorted_crop_list_en + additional_crop_list_en
        
        sorted_crop_list_ar = sorted_crop_list_df[['combined (AR)']].values.tolist()
        additional_crop_list_ar = [["91)لا إنتاج المحاصيل"], ["92)لا أعرف"], ["93)رفض"]]
        sorted_crop_list_additional_ar = sorted_crop_list_ar + additional_crop_list_ar
        
        
        
        #print(sorted_crop_list)
        final_sorted_crop_list_en = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_en]))
        final_sorted_crop_list_ar = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_ar]))

        #print(final_sorted_crop_list)
        #print(crop_list_df)

        sorted_crop_list_df['combined_DATASETCODES'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Dataset code'].astype(str)
        sorted_crop_list_DATASETCODES = sorted_crop_list_df[['combined_DATASETCODES']].values.tolist()
        #print(sorted_crop_list_DATASETCODES)
        additional_crop_list_DATASETCODES = [["91)777"], ["92)888"], ["93)999"]]
        sorted_crop_list_additional_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_DATASETCODES
        final_sorted_crop_list_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_DATASETCODES]))
        
        
        

        #CROPS SOLD
        sorted_crop_list_en_CROPS_SOLD = list(sorted_crop_list_en)
        additional_en_CROPS_SOLD = [["91)DON'T KNOW"], ["92)REFUSED"]]
        sorted_crop_list_en_CROPS_SOLD.extend(additional_en_CROPS_SOLD)
        final_sorted_crop_list_en_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en_CROPS_SOLD]))
        #print(sorted_crop_list_en_CROPS_SOLD)
        #print(final_sorted_crop_list_en_CROPS_SOLD)

        sorted_crop_list_ar_CROPS_SOLD = list(sorted_crop_list_ar)
        additional_ar_CROPS_SOLD = [["91) لا اعرف"],["92)رفض الإجابة"]]
                                     
        sorted_crop_list_ar_CROPS_SOLD.extend(additional_ar_CROPS_SOLD)
        final_sorted_crop_list_ar_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_ar_CROPS_SOLD]))
        #print(sorted_crop_list_ar_CROPS_SOLD)
        #print(final_sorted_crop_list_ar_CROPS_SOLD)
        
        
        additional_crop_list_CROPS_SOLD_DATASETCODES = [["91)888"], ["92)999"]]
        sorted_crop_list_additional_CROPS_SOLD_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_CROPS_SOLD_DATASETCODES
        final_sorted_crop_list__CROPS_SOLD_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_CROPS_SOLD_DATASETCODES]))
    
             
        
        #OLD VERSION
        # sorted_crop_list_en = sorted_crop_list_df[['combined (EN)']].values.tolist()
        # sorted_crop_list_ar = sorted_crop_list_df[['combined (AR)']].values.tolist()
        # #print(sorted_crop_list)
        # final_sorted_crop_list_en = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en]))
        # final_sorted_crop_list_ar = '\n'.join(map(str, [i[0] for i in sorted_crop_list_ar]))

        # #print(final_sorted_crop_list)
        # #print(crop_list_df)


        # #CROPS SOLD
        # sorted_crop_list_en_CROPS_SOLD = list(sorted_crop_list_en)
        # additional_en_CROPS_SOLD = [["92)DON'T KNOW"], ["93)REFUSED"]]
        # sorted_crop_list_en_CROPS_SOLD.extend(additional_en_CROPS_SOLD)
        # final_sorted_crop_list_en_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en_CROPS_SOLD]))
        # #print(sorted_crop_list_en_CROPS_SOLD)
        # #print(final_sorted_crop_list_en_CROPS_SOLD)

        # sorted_crop_list_ar_CROPS_SOLD = list(sorted_crop_list_ar)
        # additional_ar_CROPS_SOLD = [["92) لا اعرف"],["93)رفض الإجابة"]]
                                     
        # sorted_crop_list_ar_CROPS_SOLD.extend(additional_ar_CROPS_SOLD)
        # final_sorted_crop_list_ar_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_ar_CROPS_SOLD]))
        # #print(sorted_crop_list_ar_CROPS_SOLD)
        # #print(final_sorted_crop_list_ar_CROPS_SOLD)


        # # Iterate over the columns and rows, search
        # # for the text and replace
        
        ## OLD VERSION

        for k in range(number_of_rows):

            questionValue = str(worksheet[get_column_letter(2)+str(k+1)].value)
            labelValue = str(worksheet[get_column_letter(4)+str(k+1)].value)
                #print(cellValue)

            if q_name == str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en + final_sorted_crop_list_en
                newLabelValue_ar = q_label_ar + final_sorted_crop_list_ar
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_ar)

            if q_name_CROPS_SOLD in str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en_CROPS_SOLD + final_sorted_crop_list_en_CROPS_SOLD
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)

                newLabelValue_ar = q_label_ar_CROPS_SOLD + final_sorted_crop_list_ar_CROPS_SOLD
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_ar)


        replacementTextKeyPairs = {'$CROP CODES$': final_sorted_crop_list_DATASETCODES,
                                  '$CROP SOLD CODES$' : final_sorted_crop_list__CROPS_SOLD_DATASETCODES}
        for i in range(number_of_columns):
            for k in range(number_of_rows):

                cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue

        workbook["Crop list"].sheet_state = 'hidden'
        workbook.save(questionnaire_file)
        return n_of_choices
    
    
    

    elif 'Label (ES)' in columns:
        crop_list_df = crop_list_df[['Select top 10 crops ','GeoPoll code','Dataset code','Label (EN)','Label (ES)']]
        q_name = "crp_main"
        q_label_en = "What has been the main crop that your household has grown for food and income generation in $season$, if any? \n [OPERATOR: SINGLE RESPONSE. \"NO CROP PRODUCTION\" IS IN LAST OPTIONS AT THE END.] \n \n"
        q_label_es = "¿Cuál ha sido el principal producto que ha vendido en los últimos 3 meses?\n\n[OPERADOR: ""Ningún cultivo vendido en los últimos 3 meses"" APARECE AL FINAL DE LA LISTA DE OPCIONES.  ES POSIBLE SELECCIONAR MÚLTIPLES RESPUESTAS. SELECCIONE TODOS LOS CULTIVOS QUE MENCIONA EL ENCUESTADO.]\n\n"

        q_name_CROPS_SOLD = "crp_salesmain"
        q_label_en_CROPS_SOLD = "What has been the main crop that you have sold over the last 3 months?\n[OPERATOR: \"No crop sold in the last 3 months\" APPEARS AT THE END OF THE LIST OF OPTIONS. SELECT ALL CROPS LISTED BY THE RESPONDENT.]\n\n"
        q_label_es_CROPS_SOLD = "¿Cuál ha sido el principal producto que ha vendido en los últimos 3 meses?\n\n[OPERADOR: ES POSIBLE SELECCIONAR MÚLTIPLES RESPUESTAS. SELECCIONE TODOS LOS CULTIVOS QUE MENCIONA EL ENCUESTADO.]\n\n"


        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["survey"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns



        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)



        sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','GeoPoll code'])
        #print(sorted_crop_list_df)
        
        
        ##########NEWWWWWW########
        sorted_crop_list_df = sorted_crop_list_df.reset_index()
        sorted_crop_list_df['GeoPoll code'] = sorted_crop_list_df.index + 1
        #print(sorted_crop_list_df['GeoPoll code'])
        ##########NEWWWWWW########
        
        
        sorted_crop_list_df['combined (EN)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (EN)']
        sorted_crop_list_df['combined (ES)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (ES)']

        sorted_crop_list_en = sorted_crop_list_df[['combined (EN)']].values.tolist()
        additional_crop_list_en = [["91)No crop production"], ["92)DON'T KNOW"], ["93)REFUSED"]]
        sorted_crop_list_additional_en = sorted_crop_list_en + additional_crop_list_en
        
        
        sorted_crop_list_es = sorted_crop_list_df[['combined (ES)']].values.tolist()
        additional_crop_list_es = [["91)Sin producción de cultivos"], ["92)no lo sé"], ["93)RECHAZAR"]]
        sorted_crop_list_additional_es = sorted_crop_list_es + additional_crop_list_es
        
        
        #print(sorted_crop_list)
        final_sorted_crop_list_en = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_en]))
        final_sorted_crop_list_es = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_es]))

        #print(final_sorted_crop_list)
        #print(crop_list_df)


        
        sorted_crop_list_df['combined_DATASETCODES'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Dataset code'].astype(str)
        sorted_crop_list_DATASETCODES = sorted_crop_list_df[['combined_DATASETCODES']].values.tolist()
        #print(sorted_crop_list_DATASETCODES)
        additional_crop_list_DATASETCODES = [["91)777"], ["92)888"], ["93)999"]]
        sorted_crop_list_additional_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_DATASETCODES
        final_sorted_crop_list_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_DATASETCODES]))
                
        
        
        
        #CROPS SOLD
        sorted_crop_list_en_CROPS_SOLD = list(sorted_crop_list_en)
        additional_en_CROPS_SOLD = [["91)DON'T KNOW"], ["92)REFUSED"]]
        sorted_crop_list_en_CROPS_SOLD.extend(additional_en_CROPS_SOLD)
        final_sorted_crop_list_en_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en_CROPS_SOLD]))
        #print(sorted_crop_list_en_CROPS_SOLD)
        #print(final_sorted_crop_list_en_CROPS_SOLD)

        sorted_crop_list_es_CROPS_SOLD = list(sorted_crop_list_es)
        additional_es_CROPS_SOLD = [["91)NO SÉ"], ["92)RECHAZAR"]]
        sorted_crop_list_es_CROPS_SOLD.extend(additional_es_CROPS_SOLD)
        final_sorted_crop_list_es_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_es_CROPS_SOLD]))
        
        
        
        additional_crop_list_CROPS_SOLD_DATASETCODES = [["91)888"], ["92)999"]]
        sorted_crop_list_additional_CROPS_SOLD_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_CROPS_SOLD_DATASETCODES
        final_sorted_crop_list__CROPS_SOLD_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_CROPS_SOLD_DATASETCODES]))
        
        
        # sorted_crop_list_en = sorted_crop_list_df[['combined (EN)']].values.tolist()
        # sorted_crop_list_es = sorted_crop_list_df[['combined (ES)']].values.tolist()
        # #print(sorted_crop_list)
        # final_sorted_crop_list_en = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en]))
        # final_sorted_crop_list_es = '\n'.join(map(str, [i[0] for i in sorted_crop_list_es]))

        # #print(final_sorted_crop_list)
        # #print(crop_list_df)


        # #CROPS SOLD
        # sorted_crop_list_en_CROPS_SOLD = list(sorted_crop_list_en)
        # additional_en_CROPS_SOLD = [["92)DON'T KNOW"], ["93)REFUSED"]]
        # sorted_crop_list_en_CROPS_SOLD.extend(additional_en_CROPS_SOLD)
        # final_sorted_crop_list_en_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en_CROPS_SOLD]))
        # #print(sorted_crop_list_en_CROPS_SOLD)
        # #print(final_sorted_crop_list_en_CROPS_SOLD)

        # sorted_crop_list_es_CROPS_SOLD = list(sorted_crop_list_es)
        # additional_es_CROPS_SOLD = [["92)NO SÉ"], ["93)RECHAZAR"]]
        # sorted_crop_list_es_CROPS_SOLD.extend(additional_es_CROPS_SOLD)
        # final_sorted_crop_list_es_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_es_CROPS_SOLD]))


        # # Iterate over the columns and rows, search
        # # for the text and replace

        for k in range(number_of_rows):

            questionValue = str(worksheet[get_column_letter(2)+str(k+1)].value)
            labelValue = str(worksheet[get_column_letter(4)+str(k+1)].value)
                #print(cellValue)

            if q_name == str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en + final_sorted_crop_list_en
                newLabelValue_es = q_label_es + final_sorted_crop_list_es
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_es)

            if q_name_CROPS_SOLD in str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en_CROPS_SOLD + final_sorted_crop_list_en_CROPS_SOLD
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)

                newLabelValue_es = q_label_es_CROPS_SOLD + final_sorted_crop_list_es_CROPS_SOLD
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_es)
                
        replacementTextKeyPairs = {'$CROP CODES$': final_sorted_crop_list_DATASETCODES,
                                  '$CROP SOLD CODES$' : final_sorted_crop_list__CROPS_SOLD_DATASETCODES}
        for i in range(number_of_columns):
            for k in range(number_of_rows):

                cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue

        workbook["Crop list"].sheet_state = 'hidden'
        workbook.save(questionnaire_file)
        return n_of_choices        


    
    

    elif 'Label (PT)' in columns:
        crop_list_df = crop_list_df[['Select top 10 crops ','GeoPoll code','Dataset code','Label (EN)','Label (PT)']]
        q_name = "crp_main"
        q_label_en = "What has been the main crop that your household has grown for food and income generation in $season$, if any? \n [OPERATOR: SINGLE RESPONSE. \"NO CROP PRODUCTION\" IS IN LAST OPTIONS AT THE END.] \n \n"
        q_label_es = "Qual a principal cultura que o seu agregado familiar cultivou para a produção de alimentos ou geração de renda na $season PT$?\n\n[OPERADOR(A): RESPOSTA ÚNICA. ""NENHUMA PRODUÇÃO AGRÍCOLA"" APARECE NO FINAL DA LISTA DAS OPÇÕES.]\n\n"

        q_name_CROPS_SOLD = "crp_salesmain"
        q_label_en_CROPS_SOLD = "What has been the main crop that you have sold over the last 3 months?\n[OPERATOR: \"No crop sold in the last 3 months\" APPEARS AT THE END OF THE LIST OF OPTIONS. SELECT ALL CROPS LISTED BY THE RESPONDENT.]\n\n"
        q_label_es_CROPS_SOLD = "Qual a principal cultura que o(a) senhor(a) vendeu nos últimos 3 meses?\n\n[OPERADOR(A) RESPOSTA ÚNICA.]\n\n"


        workbook = openpyxl.load_workbook(questionnaire_file)
        workbook.sheetnames
        worksheet = workbook["survey"]

        number_of_rows = worksheet.max_row #Number of Rows
        number_of_columns = worksheet.max_column #Number of Columns



        list_choices = crop_list_df['Select top 10 crops '].dropna().values.tolist()
        n_of_choices = len(list_choices)



        sorted_crop_list_df = crop_list_df.sort_values(by = ['Select top 10 crops ','GeoPoll code'])
        #print(sorted_crop_list_df)
        
        
        ##########NEWWWWWW########
        sorted_crop_list_df = sorted_crop_list_df.reset_index()
        sorted_crop_list_df['GeoPoll code'] = sorted_crop_list_df.index + 1
        #print(sorted_crop_list_df['GeoPoll code'])
        ##########NEWWWWWW########
        
        
        sorted_crop_list_df['combined (EN)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (EN)']
        sorted_crop_list_df['combined (PT)'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Label (PT)']

        sorted_crop_list_en = sorted_crop_list_df[['combined (EN)']].values.tolist()
        additional_crop_list_en = [["91)No crop production"], ["92)DON'T KNOW"], ["93)REFUSED"]]
        sorted_crop_list_additional_en = sorted_crop_list_en + additional_crop_list_en
        
        
        sorted_crop_list_es = sorted_crop_list_df[['combined (PT)']].values.tolist()
        additional_crop_list_es = [["91)Sin producción de cultivos"], ["92)no lo sé"], ["93)RECHAZAR"]]
        sorted_crop_list_additional_es = sorted_crop_list_es + additional_crop_list_es
        
        
        #print(sorted_crop_list)
        final_sorted_crop_list_en = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_en]))
        final_sorted_crop_list_es = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_es]))

        #print(final_sorted_crop_list)
        #print(crop_list_df)


        
        sorted_crop_list_df['combined_DATASETCODES'] = sorted_crop_list_df['GeoPoll code'].astype(str)+')'+sorted_crop_list_df['Dataset code'].astype(str)
        sorted_crop_list_DATASETCODES = sorted_crop_list_df[['combined_DATASETCODES']].values.tolist()
        #print(sorted_crop_list_DATASETCODES)
        additional_crop_list_DATASETCODES = [["91)777"], ["92)888"], ["93)999"]]
        sorted_crop_list_additional_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_DATASETCODES
        final_sorted_crop_list_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_DATASETCODES]))
                
        
        
        
        #CROPS SOLD
        sorted_crop_list_en_CROPS_SOLD = list(sorted_crop_list_en)
        additional_en_CROPS_SOLD = [["91)DON'T KNOW"], ["92)REFUSED"]]
        sorted_crop_list_en_CROPS_SOLD.extend(additional_en_CROPS_SOLD)
        final_sorted_crop_list_en_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en_CROPS_SOLD]))
        #print(sorted_crop_list_en_CROPS_SOLD)
        #print(final_sorted_crop_list_en_CROPS_SOLD)

        sorted_crop_list_es_CROPS_SOLD = list(sorted_crop_list_es)
        additional_es_CROPS_SOLD = [["91)NO SÉ"], ["92)RECHAZAR"]]
        sorted_crop_list_es_CROPS_SOLD.extend(additional_es_CROPS_SOLD)
        final_sorted_crop_list_es_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_es_CROPS_SOLD]))
        
        
        
        additional_crop_list_CROPS_SOLD_DATASETCODES = [["91)888"], ["92)999"]]
        sorted_crop_list_additional_CROPS_SOLD_DATASETCODES = sorted_crop_list_DATASETCODES + additional_crop_list_CROPS_SOLD_DATASETCODES
        final_sorted_crop_list__CROPS_SOLD_DATASETCODES = '\n'.join(map(str, [i[0] for i in sorted_crop_list_additional_CROPS_SOLD_DATASETCODES]))
        
        
        # sorted_crop_list_en = sorted_crop_list_df[['combined (EN)']].values.tolist()
        # sorted_crop_list_es = sorted_crop_list_df[['combined (ES)']].values.tolist()
        # #print(sorted_crop_list)
        # final_sorted_crop_list_en = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en]))
        # final_sorted_crop_list_es = '\n'.join(map(str, [i[0] for i in sorted_crop_list_es]))

        # #print(final_sorted_crop_list)
        # #print(crop_list_df)


        # #CROPS SOLD
        # sorted_crop_list_en_CROPS_SOLD = list(sorted_crop_list_en)
        # additional_en_CROPS_SOLD = [["92)DON'T KNOW"], ["93)REFUSED"]]
        # sorted_crop_list_en_CROPS_SOLD.extend(additional_en_CROPS_SOLD)
        # final_sorted_crop_list_en_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_en_CROPS_SOLD]))
        # #print(sorted_crop_list_en_CROPS_SOLD)
        # #print(final_sorted_crop_list_en_CROPS_SOLD)

        # sorted_crop_list_es_CROPS_SOLD = list(sorted_crop_list_es)
        # additional_es_CROPS_SOLD = [["92)NO SÉ"], ["93)RECHAZAR"]]
        # sorted_crop_list_es_CROPS_SOLD.extend(additional_es_CROPS_SOLD)
        # final_sorted_crop_list_es_CROPS_SOLD = '\n'.join(map(str, [i[0] for i in sorted_crop_list_es_CROPS_SOLD]))


        # # Iterate over the columns and rows, search
        # # for the text and replace

        for k in range(number_of_rows):

            questionValue = str(worksheet[get_column_letter(2)+str(k+1)].value)
            labelValue = str(worksheet[get_column_letter(4)+str(k+1)].value)
                #print(cellValue)

            if q_name == str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en + final_sorted_crop_list_en
                newLabelValue_es = q_label_es + final_sorted_crop_list_es
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_es)

            if q_name_CROPS_SOLD in str(questionValue) and str(questionValue) != None:
                newLabelValue_en = q_label_en_CROPS_SOLD + final_sorted_crop_list_en_CROPS_SOLD
                #newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(4)+str(k+1)] = str(newLabelValue_en)

                newLabelValue_es = q_label_es_CROPS_SOLD + final_sorted_crop_list_es_CROPS_SOLD
                worksheet[get_column_letter(6)+str(k+1)] = str(newLabelValue_es)
                
        replacementTextKeyPairs = {'$CROP CODES$': final_sorted_crop_list_DATASETCODES,
                                  '$CROP SOLD CODES$' : final_sorted_crop_list__CROPS_SOLD_DATASETCODES}
        for i in range(number_of_columns):
            for k in range(number_of_rows):

                cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
                #print(cellValue)

                for key in replacementTextKeyPairs.keys():

                    if key in str(cellValue) and str(cellValue) != None:
                        newCellValue = cellValue.replace(key,str(replacementTextKeyPairs.get(key)))
                        worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue)
                        cellValue = newCellValue

        workbook["Crop list"].sheet_state = 'hidden'
        workbook.save(questionnaire_file)
        return n_of_choices        


def identify_added_optional_questions(template_qnames, country_qnames):
    """
    Identify optional questions (starting with 'o_') added in the country questionnaire 
    that are not present in the template.
    """
    template_optionals = {q for q in template_qnames if q.startswith('o_')}
    country_optionals = {q for q in country_qnames if q.startswith('o_')}
    
    added_optionals = sorted(country_optionals - template_optionals)
    return added_optionals



import re
from collections import defaultdict

def group_question_differences(question_list):
    """
    Groups question codes based on predefined patterns (e.g., prefixes like 'fcs_', 'shock_', etc.)
    """
    groups = defaultdict(list)

    # Define patterns and their readable group names
    pattern_map = {
        r'^fcs_': 'Food Consumption Score (FCS)',
        r'^shock_': 'Shocks & Hazards',
        r'^rcsi_': 'Reduced Coping Strategies Index (rCSI)',
        r'^cs_': 'Coping Strategies',
        r'^crp_': 'Crops',
        r'^ls_': 'Livestock',
        r'^fish_': 'Fisheries',
        r'^future_int_': 'Interview Scheduling',
        r'^hh_asset_': 'Household Productive Assets',
        r'^hh_wealth_': 'Household Wealth',
        r'^hh_admin': 'Admin Area (hh)',
        r'^hh_': 'Household Info (General)',
        r'^calldispo_': 'Call Disposition',
        r'^covid_': 'COVID-19 Impacts',
        r'^phone_': 'Phone Metadata',
        r'^o_': 'Observations / External'
    }

    for q in question_list:
        matched = False
        for pattern, label in pattern_map.items():
            if re.match(pattern, q):
                groups[label].append(q)
                matched = True
                break
        if not matched:
            groups['🟠 Other / Unmatched'].append(q)

    return dict(groups)


import pandas as pd
import numpy as np
from pandas import ExcelWriter
import re

def check_all_questions(
    questionnaire_file,
    template_file,
    result_file,
    previousround_questionnaire="no",
    previous_questionnaire_file=None
):
    # Determine reference file (template vs. previous round)
    using_previous = previousround_questionnaire.strip().lower() == "yes"
    reference_file = previous_questionnaire_file if using_previous else template_file
    suffix_ref = "previous" if using_previous else "template"
    suffix_current = "current"

    # Load data
    questionnaire_df = pd.read_excel(questionnaire_file, sheet_name='survey', skiprows=2)
    reference_df = pd.read_excel(reference_file, sheet_name='survey', skiprows=2)

    # Clean Q Name column
    questionnaire_df = questionnaire_df.dropna(subset=['Q Name'])
    reference_df = reference_df.dropna(subset=['Q Name'])
    questionnaire_df['Q Name'] = questionnaire_df['Q Name'].astype(str)
    reference_df['Q Name'] = reference_df['Q Name'].astype(str)

    # Merge on Q Name
    merged = pd.merge(reference_df, questionnaire_df, on='Q Name', suffixes=(f"_{suffix_ref}", f"_{suffix_current}"), how='inner')

    # Identify common columns to compare
    shared_columns = set(reference_df.columns).intersection(set(questionnaire_df.columns))
    shared_columns.discard('Q Name')

    writer = ExcelWriter(result_file, engine='xlsxwriter')
    summary_data = []
    result_brief = ""
    result_details = ""

    # Compare columns
    for col in sorted(shared_columns):
        col_ref = f"{col}_{suffix_ref}"
        col_curr = f"{col}_{suffix_current}"
        match_col = f"{col}_match"

        if col_ref not in merged.columns or col_curr not in merged.columns:
            continue

        merged[match_col] = np.where(
            merged[col_ref].astype(str).str.strip() == merged[col_curr].astype(str).str.strip(),
            'True', 'False'
        )

        df_mismatch = merged[merged[match_col] == 'False'][['Q Name', col_ref, col_curr]]

        if not df_mismatch.empty:
            safe_sheet_name = re.sub(r'[^A-Za-z0-9 ]+', '', col)[:31]
            df_mismatch.to_excel(writer, sheet_name=safe_sheet_name, index=False)
            message = f"\nDetected differences in: {col} ({len(df_mismatch)} mismatches)\n"
            result_brief += message
            result_details += message + "; ".join(df_mismatch['Q Name'].astype(str).tolist()) + "\n"
            summary_data.append([col, len(df_mismatch)])
        else:
            summary_data.append([col, 0])

    # Create summary sheet
    summary_df = pd.DataFrame(summary_data, columns=['Column', 'Mismatch_Count'])
    summary_df.to_excel(writer, sheet_name='Summary', index=False)

    writer.close()

    if summary_df['Mismatch_Count'].sum() == 0:
        result_brief = "✅ No differences detected between reference and questionnaire."

    return result_brief, result_details




def check_all_questions_unified(
    questionnaire_file,
    template_file,
    result_file,
    previousround_questionnaire="no",
    previous_questionnaire_file=None
):
    import pandas as pd
    import numpy as np
    import re
    from pandas import ExcelWriter

    using_previous = previousround_questionnaire.strip().lower() == "yes"
    reference_file = previous_questionnaire_file if using_previous else template_file
    suffix_ref = "Previous" if using_previous else "Template"
    suffix_curr = "Current"

    questionnaire_df = pd.read_excel(questionnaire_file, sheet_name='survey', skiprows=2)
    reference_df = pd.read_excel(reference_file, sheet_name='survey', skiprows=2)

    questionnaire_df = questionnaire_df.dropna(subset=['Q Name'])
    reference_df = reference_df.dropna(subset=['Q Name'])
    questionnaire_df['Q Name'] = questionnaire_df['Q Name'].astype(str)
    reference_df['Q Name'] = reference_df['Q Name'].astype(str)

    merged = pd.merge(reference_df, questionnaire_df, on='Q Name', suffixes=(f"_{suffix_ref}", f"_{suffix_curr}"), how='inner')
    shared_columns = set(reference_df.columns).intersection(set(questionnaire_df.columns))
    shared_columns.discard('Q Name')

    summary_data = []
    all_mismatches = []

    for col in sorted(shared_columns):
        col_ref = f"{col}_{suffix_ref}"
        col_curr = f"{col}_{suffix_curr}"
        if col_ref not in merged.columns or col_curr not in merged.columns:
            continue

        merged[f"{col}_match"] = np.where(
            merged[col_ref].astype(str).str.strip() == merged[col_curr].astype(str).str.strip(),
            'True', 'False'
        )

        df_mismatch = merged[merged[f"{col}_match"] == 'False'][['Q Name', col_ref, col_curr]]
        if not df_mismatch.empty:
            df_mismatch = df_mismatch.rename(columns={
                col_ref: 'Previous Value',
                col_curr: 'Current Value'
            })
            df_mismatch.insert(1, 'Field Miss matched', col)
            all_mismatches.append(df_mismatch)
            summary_data.append([col, len(df_mismatch)])
        else:
            summary_data.append([col, 0])

    with ExcelWriter(result_file, engine='xlsxwriter') as writer:
        if all_mismatches:
            mismatches_df = pd.concat(all_mismatches, ignore_index=True)
            mismatches_df.to_excel(writer, sheet_name='All_Mismatches', index=False)
            result_brief = f"❌ Detected {sum(x[1] for x in summary_data)} mismatches across {len(all_mismatches)} fields."
            result_details = mismatches_df.to_string(index=False)
        else:
            result_brief = "✅ No differences detected between reference and questionnaire."
            result_details = ""

        pd.DataFrame(summary_data, columns=['Field', 'Mismatch Count']).to_excel(writer, sheet_name='Summary', index=False)

    return result_brief, result_details